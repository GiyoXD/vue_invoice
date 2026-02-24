import openpyxl
import sys

def main():
    try:
        path = r"c:\Users\JPZ031127\Desktop\project\vue_invoice_project\temp_uploads\MAD INVOICE.xlsx"
        print(f"Loading {path}")
        wb = openpyxl.load_workbook(path, data_only=True)
        print("Sheets:", wb.sheetnames)
        
        for sheet_name in wb.sheetnames:
            if "detail" in sheet_name.lower() or "packing" in sheet_name.lower():
                print(f"\n--- Sheet: {sheet_name} ---")
                ws = wb[sheet_name]
                print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
                for r in range(1, min(ws.max_row + 1, 100)):
                    row_data = []
                    for c in range(1, min(ws.max_column + 1, 30)):
                        val = ws.cell(row=r, column=c).value
                        # also print formula if available
                        if val is not None:
                            row_data.append(f"{c}:{repr(val)}")
                    if row_data:
                        print(f"Row {r}: " + " | ".join(row_data))
                        
                # Let's also check formulas
        print("\nNow loading with formulas")
        wb_formula = openpyxl.load_workbook(path, data_only=False)
        for sheet_name in wb_formula.sheetnames:
            if "detail" in sheet_name.lower() or "packing" in sheet_name.lower():
                print(f"\n--- Sheet: {sheet_name} (Formulas) ---")
                ws = wb_formula[sheet_name]
                for r in range(max(1, ws.max_row - 50), ws.max_row + 1):
                    row_data = []
                    for c in range(1, min(ws.max_column + 1, 30)):
                        val = ws.cell(row=r, column=c).value
                        if val is not None and isinstance(val, str) and val.startswith("="):
                            row_data.append(f"{c}:{repr(val)}")
                    if row_data:
                        print(f"Row {r}: " + " | ".join(row_data))
                        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
