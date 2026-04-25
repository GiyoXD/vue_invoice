# core/invoice_generator/generate_invoice.py
import os
import json
import pickle
import argparse
import sys
import time
import datetime
import logging
import traceback
from pathlib import Path
from typing import Optional, Dict, Any, List
import openpyxl
import re
import ast


# Keep your existing imports
from core.invoice_generator.config.config_loader import BundledConfigLoader
from core.invoice_generator.builders.workbook_builder import WorkbookBuilder
from core.invoice_generator.builders.deep_sheet_builder import DeepSheetBuilder
from core.invoice_generator.processors.single_table_processor import SingleTableProcessor
from core.invoice_generator.processors.multi_table_processor import MultiTableProcessor

from core.invoice_generator.utils.print_area_config import configure_print_area
from core.invoice_generator.utils.generation_session import GenerationSession
from core.invoice_generator.resolvers import InvoiceAssetResolver
from core.utils.file_lock import ensure_file_unlocked

logger = logging.getLogger(__name__)



# --- Constants for Blueprints ---
from core.system_config import sys_config

DEFAULT_TEMPLATE_DIR = sys_config.templates_dir
DEFAULT_CONFIG_DIR = sys_config.registry_dir


# --- Helper Functions ---









def run_invoice_generation(
    input_data_path: Path,
    output_path: Path,
    template_dir: Optional[Path] = None,
    config_dir: Optional[Path] = None,
    daf_mode: bool = False,
    custom_mode: bool = False,
    explicit_config_path: Optional[Path] = None,
    explicit_template_path: Optional[Path] = None,
    input_data_dict: Optional[Dict[str, Any]] = None,
    return_bytes: bool = False,
    enable_auto_fit: bool = True
):
    """
    Library entry point for invoice generation. 
    Uses GenerationSession context manager to ensure robust error handling.
    """
    # 0. FORCE CLEAR SESSION LOG (Per User Request)
    try:
        from core.logger_config import clear_session_log
        clear_session_log()
    except Exception:
        pass

    # 1. Resolve Paths
    input_data_path, output_path, template_dir, config_dir = _resolve_generation_paths(
        input_data_path, output_path, template_dir, config_dir
    )

    # 2. Initialize Context
    ctx = _initialize_context(
        input_data_path, output_path, template_dir, config_dir,
        daf_mode, custom_mode, enable_auto_fit, explicit_config_path, explicit_template_path, input_data_dict
    )

    # === CORE GENERATION LOGIC WITH MONITOR ===
    # Using 'meta_args' compatible dict for monitor (removing argparse dep)
    monitor_args = {
        "DAF": daf_mode,
        "custom": custom_mode,
        "input_data_file": str(input_data_path),
        "configdir": str(config_dir)
    }

    with GenerationSession(output_path, args=monitor_args, input_data=ctx.invoice_data) as session:
        logger.info("=== Starting Invoice Generation (Orchestrated) ===")
        
        # 3. Execution Pipeline
        try:
            _load_resources(ctx)
        except ValueError as e:
            logger.error(f"[_load_resources] Failed to load resources: {e}")
            raise
        monitor_paths = {
            'template': str(ctx.paths.get('template', 'unknown')),
            'config': str(ctx.paths.get('config', 'unknown'))
        }
        session.update_logs(header_info={"resolved_paths": monitor_paths})

        _prepare_workbooks(ctx)
        
        _process_sheets(ctx, session)
        
        _inject_unknown_sheets(ctx)
        
    # 4. Build dynamic output filename based on sheets & invoice_id
        _build_output_filename(ctx)
        
        if return_bytes:
            import io
            logger.info("Applying Print Area & Page Setup...")

            # Only apply print settings to sheets defined in the config
            configured_sheets = set(ctx.config_loader.get_sheets_to_process())

            for sheet in ctx.output_workbook.sheetnames:
                if sheet not in configured_sheets:
                    continue
                try:
                    ws = ctx.output_workbook[sheet]
                    if ws is None: continue
                    max_col = _count_layout_columns(ctx, sheet)
                    configure_print_area(ws, max_col_override=max_col)
                except Exception as e:
                    logger.error(f"Print setup failed for '{sheet}': {e}")
                    
            logger.info("Saving workbook to in-memory buffer")
            buffer = io.BytesIO()
            ctx.output_workbook.save(buffer)
            buffer.seek(0)
            
            # Cleanup
            if ctx.template_workbook: ctx.template_workbook.close()
            if ctx.output_workbook: ctx.output_workbook.close()
            
            return ctx.output_path.name, buffer.getvalue()
        else:
            _finalize(ctx)

    return ctx.output_path


# --- Internal Pipeline Structures ---

class GeneratorContext:
    """Holds state for the invoice generation pipeline."""
    def __init__(self, input_path: Path, output_path: Path, invoice_data: Dict):
        self.input_path = input_path
        self.output_path = output_path
        self.invoice_data = invoice_data
        
        # Paths
        self.template_dir: Optional[Path] = None
        self.config_dir: Optional[Path] = None
        self.paths: Dict[str, Path] = {}
        
        # Config & Resources
        self.config_loader: Optional[BundledConfigLoader] = None
        self.template_workbook: Optional[openpyxl.Workbook] = None
        self.output_workbook: Optional[openpyxl.Workbook] = None
        
        # Flags
        self.daf_mode = False
        self.custom_mode = False
        self.enable_auto_fit = True
        
        # Derived
        self.final_grand_total_pallets = 0


def _initialize_context(
    input_path: Path, output_path: Path, 
    template_dir: Path, config_dir: Path,
    daf_mode: bool, custom_mode: bool, enable_auto_fit: bool,
    manual_config: Optional[Path], manual_template: Optional[Path],
    data_dict: Optional[Dict]
) -> GeneratorContext:
    invoice_data = data_dict or {}
    if not invoice_data:
        logger.warning("No input data dictionary provided.")

    ctx = GeneratorContext(input_path, output_path, invoice_data)
    ctx.template_dir = template_dir
    ctx.config_dir = config_dir
    ctx.daf_mode = daf_mode
    ctx.custom_mode = custom_mode
    ctx.enable_auto_fit = enable_auto_fit
    
    # Pre-resolve known manual paths
    if manual_config: ctx.paths['config'] = manual_config.resolve()
    if manual_template: ctx.paths['template'] = manual_template.resolve()
    
    return ctx


def _load_resources(ctx: GeneratorContext):
    """Stage 1: Resolve assets and load configuration."""
    # A. Resolve Paths
    resolver = InvoiceAssetResolver(base_config_dir=ctx.config_dir, base_template_dir=ctx.template_dir)
    assets = resolver.resolve_assets_for_input_file(str(ctx.input_path))
    
    if assets:
        if 'config' not in ctx.paths:
            ctx.paths['config'] = assets.config_path
            logger.info(f"Using resolved config: {ctx.paths['config']}")
        if 'template' not in ctx.paths:
            ctx.paths['template'] = assets.template_path
            logger.info(f"Using resolved template: {ctx.paths['template']}")
            
    ctx.paths['data'] = ctx.input_path

    # Validation
    if 'config' not in ctx.paths or 'template' not in ctx.paths:
         raise FileNotFoundError(f"Could not resolve config/template for '{ctx.input_path.name}'")

    # B. Load Config
    try:
        ctx.config_loader = BundledConfigLoader(ctx.paths['config'])
    except Exception as e:
        raise RuntimeError(f"Failed to load configuration: {e}") from e

    # C. Calculate Grand Total Pallets
    # Read the pre-calculated total from footer_data.grand_total (set by data_parser)
    footer_data = ctx.invoice_data.get('footer_data', {}) or {}
    grand_total = footer_data.get('grand_total')
    
    if not grand_total or 'col_pallet_count' not in grand_total:
        raise ValueError("CRITICAL: Missing grand total pallet count in parsed data. Cannot generate invoice.")
        
    gt_pallets = grand_total.get('col_pallet_count', 0)
    
    if gt_pallets:
        ctx.final_grand_total_pallets = int(gt_pallets)
        logger.info(f"Grand total pallets from footer_data: {ctx.final_grand_total_pallets}")
    else:
        ctx.final_grand_total_pallets = 0
        logger.warning("⚠ No pallet count found in footer_data.grand_total.col_pallet_count")


def _prepare_workbooks(ctx: GeneratorContext):
    """Stage 2: Load Template and Build Output Workbook."""
    ctx.output_path.parent.mkdir(parents=True, exist_ok=True)
    
    logger.info(f"Initializing clean output workbook (JSON-only mode)")
    
    # REQUIRED: Load JSON template config
    json_config = ctx.config_loader.get_template_json_config()
    if not json_config:
        error_msg = f"CRITICAL: No JSON template found for client/config. JSON templates are now REQUIRED. (Missing *_template.json?)"
        logger.critical(error_msg)
        raise ValueError(error_msg)

    # Initialize clean workbook
    ctx.output_workbook = openpyxl.Workbook()
    default_ws = ctx.output_workbook.active
    if default_ws: ctx.output_workbook.remove(default_ws)
    
    # Create sheets defined in JSON config
    for sheet_name in json_config.keys():
        ctx.output_workbook.create_sheet(sheet_name)
        logger.info(f"Created sheet '{sheet_name}' from JSON template")
        
    # Set template_workbook to refer to output_workbook 
    # (since we are creating from scratch, they are effectively the same object in this new flow)
    # This satisfies processors that expect a template_workbook object, although they should rely on JSON.
    ctx.template_workbook = ctx.output_workbook

    # Deep Sheet Injection
    try:
        DeepSheetBuilder.build(ctx.output_workbook, ctx.invoice_data)
    except Exception as e:
        logger.warning(f"DeepSheet injection skipped: {e}")


def _process_sheets(ctx: GeneratorContext, session: GenerationSession):
    """Stage 3: Iterate and process each configured sheet."""
    sheets_config = ctx.config_loader.get_sheets_to_process()
    sheets_to_process = [s for s in sheets_config if s in ctx.output_workbook.sheetnames]
    
    if not sheets_to_process:
        raise ValueError("No valid sheets found to process.")

    # Mock args for processors (removing argparse dependency logic)
    # Processors expect an object with .DAF and .custom flags
    class ProcessorFlags:
        def __init__(self, daf, custom, enable_auto_fit):
            self.DAF = daf
            self.custom = custom
            self.enable_auto_fit = enable_auto_fit
    
    proc_args = ProcessorFlags(ctx.daf_mode, ctx.custom_mode, ctx.enable_auto_fit)

    for sheet_name in sheets_to_process:
        logger.info(f"Processing sheet '{sheet_name}'")
        try:
            tmpl_ws = ctx.template_workbook[sheet_name]
            out_ws = ctx.output_workbook[sheet_name]
            sheet_conf = ctx.config_loader.get_sheet_config(sheet_name)
            ds_type = ctx.config_loader.get_data_source_type(sheet_name)
            
            if not ds_type:
                continue

            processor = _get_processor(
                ds_type, tmpl_ws, out_ws, sheet_name, sheet_conf, 
                ctx.config_loader, ctx.invoice_data, proc_args, ctx.final_grand_total_pallets,
                ctx.template_workbook, ctx.output_workbook 
            )

            if processor and processor.process():
                 session.log_success(sheet_name)
                 if hasattr(processor, 'header_info'):
                     session.update_logs(header_info=processor.header_info)
            else:
                 session.log_failure(sheet_name, error=RuntimeError("Processor returned False"))

        except Exception as e:
            session.log_failure(sheet_name, error=e)
            raise e


def _get_processor(ds_type, tmpl_ws, out_ws, name, conf, loader, data, args, pallets, tmpl_wb, out_wb):
    """Factory method for processors."""
    # Common kwargs
    kwargs = {
        "template_worksheet": tmpl_ws,
        "output_worksheet": out_ws,
        "sheet_name": name,
        "sheet_config": conf,
        "config_loader": loader,
        "data_source_indicator": ds_type,
        "invoice_data": data,
        "cli_args": args,
        "final_grand_total_pallets": pallets,
        "template_workbook": tmpl_wb,
        "output_workbook": out_wb
    }

    if ds_type in ["processed_tables_multi", "processed_tables", "detail_packing_list"]:
        return MultiTableProcessor(**kwargs)
    elif "aggregation" in ds_type or ds_type in ["DAF_aggregation", "summary_packing_list"]:
        # Fallback to aggregation for unknown/custom types that have 'aggregation' in the name
        return SingleTableProcessor(**kwargs)
    else:
        logger.warning(f"Unknown data source type '{ds_type}', falling back to SingleTableProcessor")
        return SingleTableProcessor(**kwargs)


def _inject_unknown_sheets(ctx: GeneratorContext):
    """
    Stage 3.5: Copy unknown sheets from bundled .xlsx into output workbook.
    
    Unknown sheets are those present in the original bundled .xlsx template
    but NOT defined in the JSON template config. These are preserved as-is,
    maintaining their original visibility state (visible/hidden/veryHidden),
    cell values, styles, merged cells, and dimensions.
    """
    # Derive the bundle directory from the config path
    config_path = Path(ctx.paths.get('config', ''))
    if not config_path.exists():
        logger.debug("No config path found, skipping unknown sheet injection.")
        return
    
    bundle_dir = config_path.parent
    
    # Derive prefix from config filename (e.g. "TEST_VN_config.json" -> "TEST_VN")
    config_stem = config_path.stem  # "TEST_VN_config"
    prefix = config_stem.replace("_config", "")  # "TEST_VN"
    
    # Find matching .xlsx by prefix first, fallback to any .xlsx
    source_xlsx_path = bundle_dir / f"{prefix}.xlsx"
    if not source_xlsx_path.exists():
        xlsx_candidates = list(bundle_dir.glob("*.xlsx"))
        if not xlsx_candidates:
            logger.debug(f"No .xlsx files found in bundle dir: {bundle_dir}")
            return
        source_xlsx_path = xlsx_candidates[0]
    logger.info(f"[Unknown Sheets] Loading source template: {source_xlsx_path.name}")
    
    try:
        source_wb = openpyxl.load_workbook(source_xlsx_path)
    except Exception as e:
        logger.warning(f"[Unknown Sheets] Failed to load source xlsx: {e}")
        return
    
    # Determine which sheets are "configured" (already in output)
    configured_sheets = set(ctx.output_workbook.sheetnames)
    unknown_sheets = [s for s in source_wb.sheetnames if s not in configured_sheets]
    
    if not unknown_sheets:
        logger.info("[Unknown Sheets] No unknown sheets to inject.")
        source_wb.close()
        return
    
    logger.info(f"[Unknown Sheets] Found {len(unknown_sheets)} unknown sheet(s): {unknown_sheets}")
    
    for sheet_name in unknown_sheets:
        try:
            source_ws = source_wb[sheet_name]
            target_ws = ctx.output_workbook.create_sheet(sheet_name)
            _deep_copy_worksheet(source_ws, target_ws)
            logger.info(f"  ✅ Injected '{sheet_name}' (state={source_ws.sheet_state})")
        except Exception as e:
            logger.warning(f"  ⚠ Failed to inject '{sheet_name}': {e}")
    
    source_wb.close()


def _deep_copy_worksheet(source_ws, target_ws):
    """
    Deep copy a worksheet's content from one workbook to another.
    
    Copies cell values, styles, merged cells, column dimensions,
    row dimensions, and sheet visibility state.
    
    Args:
        source_ws: Source worksheet (from bundled .xlsx)
        target_ws: Target worksheet (in output workbook)
    """
    from openpyxl.utils import get_column_letter
    from copy import copy
    
    # 1. Copy cell values and styles
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            
            # Copy style attributes
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.fill = copy(cell.fill)
                target_cell.border = copy(cell.border)
                target_cell.alignment = copy(cell.alignment)
                target_cell.number_format = cell.number_format
                target_cell.protection = copy(cell.protection)
    
    # 2. Copy merged cell ranges
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))
    
    # 3. Copy column dimensions (widths)
    for col_letter, col_dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dim.width
        target_ws.column_dimensions[col_letter].hidden = col_dim.hidden
    
    # 4. Copy row dimensions (heights)
    for row_num, row_dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_num].height = row_dim.height
        target_ws.row_dimensions[row_num].hidden = row_dim.hidden
    
    # 5. Copy sheet visibility state (visible/hidden/veryHidden)
    target_ws.sheet_state = source_ws.sheet_state


def _get_mode_suffix(ctx: GeneratorContext) -> str:
    """
    Returns a filename suffix based on the active generation mode.
    
    Returns:
        ' DAF' if daf_mode is active,
        ' Custom' if custom_mode is active,
        '' for standard mode.
    """
    if ctx.daf_mode:
        return " DAF"
    elif ctx.custom_mode:
        return " Custom"
    return ""


def _build_output_filename(ctx: GeneratorContext):
    """
    Builds a dynamic output filename based on what sheets are present,
    the invoice ID, and the active generation mode.
    
    Maps sheet names to abbreviations:
        - "Contract" -> "CT"
        - "Invoice"  -> "INV"
        - "Packing list" -> "PL"
    
    Result examples:
        - Standard: "CT&INV&PL MT2-26007E.xlsx"
        - Custom:   "CT&INV&PL MT2-26007E Custom.xlsx"
        - DAF:      "CT&INV&PL MT2-26007E DAF.xlsx"
    """
    # Sheet name -> abbreviation mapping (order matters for the prefix)
    SHEET_ABBREVS = [
        ("Contract",     "CT"),
        ("Invoice",      "INV"),
        ("Packing list", "PL"),
    ]
    
    # Build prefix from sheets present in the output workbook
    present_abbrevs = []
    for sheet_name, abbrev in SHEET_ABBREVS:
        if sheet_name in ctx.output_workbook.sheetnames:
            present_abbrevs.append(abbrev)
    
    if not present_abbrevs:
        logger.warning("[Filename] No recognizable sheets found. Using default filename.")
        return
    
    prefix = "&".join(present_abbrevs)
    
    # Extract invoice_id from invoice_data
    inv_no = ""
    if 'invoice_info' in ctx.invoice_data:
        inv_no = ctx.invoice_data['invoice_info'].get('col_inv_no', "") or \
                 ctx.invoice_data['invoice_info'].get('inv_no', "")
    
    if not inv_no:
        # Fallback: try processed_tables_multi
        tables = ctx.invoice_data.get('processed_tables_multi', {})
        table_1 = tables.get('1', {})
        vals = table_1.get('col_inv_no', [])
        if isinstance(vals, list):
            for v in vals:
                if v:
                    inv_no = str(v)
                    break
    
    # Get mode suffix (e.g. " DAF", " Custom", or "")
    mode_suffix = _get_mode_suffix(ctx)
    
    # Compose filename
    if inv_no:
        new_name = f"{prefix} {inv_no}{mode_suffix}.xlsx"
    else:
        new_name = f"{prefix}{mode_suffix}.xlsx"
    
    # Sanitize: remove characters illegal in Windows filenames
    new_name = re.sub(r'[<>:"/\\|?*]', '_', new_name)
    
    new_output_path = ctx.output_path.parent / new_name
    logger.info(f"[Filename] Dynamic output: {new_name}")
    ctx.output_path = new_output_path


def _count_layout_columns(ctx: GeneratorContext, sheet_name: str):
    """
    Count the actual number of Excel columns from the layout config's structure.columns.

    Accounts for parent columns with children (e.g. col_qty_header with
    children col_qty_pcs + col_qty_sf = 2 actual columns, not 3).

    Args:
        ctx: The generator context containing the config loader.
        sheet_name: Name of the sheet to count columns for.

    Returns:
        int column count, or None if no layout structure is defined.
    """
    layout = ctx.config_loader.get_layout_config(sheet_name)
    columns = layout.get('structure', {}).get('columns', [])
    if not columns:
        return None

    count = 0
    for col in columns:
        children = col.get('children', [])
        if children:
            count += len(children)
        else:
            count += 1

    logger.debug(f"[PrintArea] Layout column count for '{sheet_name}': {count}")
    return count


def _finalize(ctx: GeneratorContext):
    """Stage 4: Apply print settings and save."""
    logger.info("Applying Print Area & Page Setup...")

    # Only apply print settings to sheets defined in the config
    configured_sheets = set(ctx.config_loader.get_sheets_to_process())

    for sheet in ctx.output_workbook.sheetnames:
        if sheet not in configured_sheets:
            logger.info(f"Skipping print setup for unconfigured sheet '{sheet}'")
            continue

        try:
            ws = ctx.output_workbook[sheet]
            if ws is None:
                logger.warning(f"Sheet '{sheet}' is in sheetnames but returned None - skipping print setup")
                continue

            # Use the layout config to determine the correct max column
            max_col = _count_layout_columns(ctx, sheet)
            configure_print_area(ws, max_col_override=max_col)
        except Exception as e:
            logger.error(f"Print setup failed for '{sheet}': {e}")


    logger.info(f"Saving workbook to {ctx.output_path}")
    
    # Check for file locks and attempting to kill Excel if needed
    try:
        ensure_file_unlocked(ctx.output_path)
    except Exception as e:
        logger.error(f"File Lock Error: {e}")
        # We raise here because if we can't write, we can't save.
        raise e
        
    ctx.output_workbook.save(ctx.output_path)
    
    # Cleanup
    if ctx.template_workbook: ctx.template_workbook.close()
    if ctx.output_workbook: ctx.output_workbook.close()


def _resolve_generation_paths(
    input_data_path: Path, 
    output_path: Path, 
    template_dir: Optional[Path] = None, 
    config_dir: Optional[Path] = None
) -> tuple[Path, Path, Path, Path]:
    """Resolves all paths and applies default directories if necessary."""
    # Ensure inputs are Path objects
    input_data_path = Path(input_data_path).resolve()
    output_path = Path(output_path).resolve()

    # Apply defaults for blueprint directories
    if template_dir is None:
        template_dir = DEFAULT_TEMPLATE_DIR
        logger.info(f"Using default blueprint template directory: {template_dir}")
    if config_dir is None:
        config_dir = DEFAULT_CONFIG_DIR
        logger.info(f"Using default blueprint config directory: {config_dir}")

    template_dir = Path(template_dir).resolve()
    config_dir = Path(config_dir).resolve()

    return input_data_path, output_path, template_dir, config_dir


def main():
    """CLI Entry point for backward compatibility."""
    parser = argparse.ArgumentParser(description="Generate Invoice CLI")
    parser.add_argument("input_data_file", help="Path to input data file")
    parser.add_argument("-o", "--output", default=None, help="Output path (default: output/ dir)")
    parser.add_argument("-t", "--templatedir", default=None, help="Template dir (defaults to database/blueprints/template)")
    parser.add_argument("-c", "--configdir", default=None, help="Config dir (defaults to database/blueprints/config/bundled)")
    parser.add_argument("--config", help="Explicit path to config file")
    parser.add_argument("--template", help="Explicit path to template file")
    parser.add_argument("--DAF", action="store_true", help="DAF mode")
    parser.add_argument("--custom", action="store_true", help="Custom mode")
    parser.add_argument("--no-auto-fit", action="store_true", help="Disable auto-fit column dimensions")
    parser.add_argument("--debug", action="store_true", help="Debug logging")
    
    args = parser.parse_args()
    
    # Configure Logging for CLI using centralized logger
    from core.logger_config import setup_logging
    from core.system_config import sys_config
    level = logging.DEBUG if args.debug else logging.INFO
    setup_logging(log_dir=sys_config.run_log_dir, level=level)
    
    # Determine Output Path
    if args.output:
        output_path = Path(args.output)
    else:
        # Derive from input stem
        input_stem = Path(args.input_data_file).stem
        output_path = sys_config.output_dir / f"{input_stem}.xlsx"
        # Ensure output dir exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
    
    try:
        # Load data for CLI usage
        cli_data = {}
        try:
            with open(args.input_data_file, 'r', encoding='utf-8') as f:
                cli_data = json.load(f)
        except Exception as e:
            print(f"Failed to load input data file: {e}")
            sys.exit(1)

        run_invoice_generation(
            input_data_path=Path(args.input_data_file),
            output_path=output_path,
            template_dir=Path(args.templatedir) if args.templatedir else None,
            config_dir=Path(args.configdir) if args.configdir else None,
            daf_mode=args.DAF,
            custom_mode=args.custom,
            enable_auto_fit=not args.no_auto_fit,
            explicit_config_path=Path(args.config) if args.config else None,
            explicit_template_path=Path(args.template) if args.template else None,
            input_data_dict=cli_data
        )
        print(f"Successfully generated: {args.output}")
    except Exception as e:
        print(f"Generation failed: {e}")
        # The monitor would have already written the metadata file with the stack trace
        sys.exit(1)

if __name__ == "__main__":
    main()