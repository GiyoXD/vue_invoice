"""
Cell Styler - Applies styles from StyleRegistry to Excel cells

This module handles the actual application of merged styles to openpyxl cells,
translating style dictionaries into openpyxl Font, Alignment, Fill, Border objects.
"""

import logging
from typing import Dict, Any, Optional
from openpyxl.cell import Cell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class CellStyler:
    """
    Applies style definitions to Excel cells.
    
    Translates style dictionaries from StyleRegistry into openpyxl style objects.
    
    Usage:
        styler = CellStyler()
        cell = worksheet['A1']
        
        style = {'format': '0.00', 'bold': True, 'alignment': 'center', 'fill_color': 'CCCCCC'}
        styler.apply(cell, style)
    """
    
    # Border style mapping
    BORDER_STYLES = {
        'thin': 'thin',
        'medium': 'medium',
        'thick': 'thick',
        'double': 'double',
        'hair': 'hair',
        'dashed': 'dashed',
        'dotted': 'dotted'
    }
    
    def apply(self, cell: Cell, style: Dict[str, Any]):
        """
        Apply style dictionary to cell.
        
        Args:
            cell: openpyxl Cell object
            style: Style dictionary from StyleRegistry
        """
        if not style:
            logger.warning(f"warning!!  Cell {cell.coordinate}: NO style dictionary provided!")
            return
        
        # Validate expected style properties
        expected_props = ['alignment', 'format', 'font_name', 'font_size']
        missing_props = [prop for prop in expected_props if prop not in style or style[prop] is None]
        if missing_props:
            logger.warning(f"warning!!  Cell {cell.coordinate}: Missing style properties: {missing_props}")
            logger.warning(f"   → Style dict keys: {list(style.keys())}")
        
        # Apply font properties (bold, italic, size, name)
        self._apply_font(cell, style)
        
        # Apply alignment (horizontal, vertical, wrap)
        self._apply_alignment(cell, style)
        
        # Apply fill color
        self._apply_fill(cell, style)
        
        # Apply borders
        self._apply_borders(cell, style)
        
        # Apply number format
        self._apply_format(cell, style)
    
    def _apply_font(self, cell: Cell, style: Dict[str, Any]):
        """Apply font properties to cell."""
        font_kwargs = {}
        
        # Check for required font properties
        required_font_props = ['font_name', 'font_size']
        missing_font_props = [prop for prop in required_font_props if not style.get(prop)]
        
        if missing_font_props:
            logger.warning(f"warning!!  Cell {cell.coordinate}: Missing required font properties: {missing_font_props}")
            logger.warning(f"   → Available style keys: {list(style.keys())}")
            return
        
        if style.get('bold') is not None:
            font_kwargs['bold'] = style['bold']
        
        if style.get('italic') is not None:
            font_kwargs['italic'] = style['italic']
        
        if style.get('font_size'):
            font_kwargs['size'] = style['font_size']
        
        if style.get('font_name'):
            font_kwargs['name'] = style['font_name']
        
        if font_kwargs:
            cell.font = Font(**font_kwargs)
    
    def _apply_alignment(self, cell: Cell, style: Dict[str, Any]):
        """Apply alignment properties to cell."""
        alignment_kwargs = {}
        
        # Check for required alignment property
        if not style.get('alignment'):
            logger.warning(f"warning!!  Cell {cell.coordinate}: Missing required alignment property")
            logger.warning(f"   → Available style keys: {list(style.keys())}")
            return
        
        if style.get('alignment'):
            align_val = style['alignment']
            if isinstance(align_val, dict):
                # New format: alignment is a dict of properties
                alignment_kwargs.update(align_val)
            else:
                # Legacy format: alignment is just the horizontal string
                alignment_kwargs['horizontal'] = align_val
        
        # Always default to center for vertical alignment
        alignment_kwargs['vertical'] = style.get('vertical_alignment', 'center')
        
        if style.get('wrap_text') is not None:
            alignment_kwargs['wrap_text'] = style['wrap_text']
        
        if alignment_kwargs:
            cell.alignment = Alignment(**alignment_kwargs)
    
    def _apply_fill(self, cell: Cell, style: Dict[str, Any]):
        """Apply fill color to cell."""
        if style.get('fill_color'):
            fill_color = style['fill_color']
            # Remove '#' if present
            if fill_color.startswith('#'):
                fill_color = fill_color[1:]
            
            cell.fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type='solid'
            )
    
    def _apply_borders(self, cell: Cell, style: Dict[str, Any]):
        """Apply border style to cell."""
        border_style_name = style.get('border_style')
        
        if border_style_name:
            # Map style name to openpyxl border style
            openpyxl_style = self.BORDER_STYLES.get(border_style_name, 'thin')
            
            # Create border sides
            side = Side(style=openpyxl_style, color='000000')
            
            # Special case: no_bottom border (for static content rows)
            if border_style_name == 'no_bottom':
                cell.border = Border(
                    left=side,
                    right=side,
                    top=side,
                    bottom=Side(style=None)  # No bottom border
                )
            # Special case: sides_only border (for col_static column)
            elif border_style_name == 'sides_only':
                cell.border = Border(
                    left=side,
                    right=side,
                    top=Side(style=None),     # No top border
                    bottom=Side(style=None)   # No bottom border
                )
            else:
                # Apply to all sides (standard behavior)
                cell.border = Border(
                    left=side,
                    right=side,
                    top=side,
                    bottom=side
                )
        # Note: If border_style not in style dict, no borders are applied
        # This is expected behavior - borders are optional styling
    
    def _apply_format(self, cell: Cell, style: Dict[str, Any]):
        """Apply number format to cell."""
        # Check for required format property
        if not style.get('format'):
            logger.warning(f"warning!!  Cell {cell.coordinate}: Missing required format property")
            logger.warning(f"   → Available style keys: {list(style.keys())}")
            return
        
        if style.get('format'):
            cell.number_format = style['format']
    
    def apply_row_height(self, worksheet, row_num: int, height: Optional[int]):
        """
        Apply row height to a specific row.
        
        Args:
            worksheet: openpyxl Worksheet
            row_num: Row number (1-indexed)
            height: Height in points (None = default)
        """
        if height:
            worksheet.row_dimensions[row_num].height = height
    
    def apply_column_width(self, worksheet, col_letter: str, width: Optional[int]):
        """
        Apply column width to a specific column.
        
        Args:
            worksheet: openpyxl Worksheet
            col_letter: Column letter ('A', 'B', etc.)
            width: Width in characters (None = default)
        """
        if width:
            worksheet.column_dimensions[col_letter].width = width
    
    def apply_to_range(self, worksheet, start_row: int, end_row: int, col_index: int, style: Dict[str, Any]):
        """
        Apply style to a range of cells in a column.
        
        Args:
            worksheet: openpyxl Worksheet
            start_row: Starting row (1-indexed)
            end_row: Ending row (1-indexed)
            col_index: Column index (1-indexed)
            style: Style dictionary
        """
        col_letter = get_column_letter(col_index)
        
        for row in range(start_row, end_row + 1):
            cell = worksheet[f'{col_letter}{row}']
            self.apply(cell, style)
