from logging import log
import logging
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Tuple
import copy

logger = logging.getLogger(__name__)

class TemplateStateBuilder:
    """
    A builder responsible for capturing and restoring the state of a template file.
    This includes the header, footer, and other static content.
    
    Captures the original template state during initialization, before any modifications.
    """
    
    DEBUG = False  # Set to True to enable debug printing

    def __init__(self, worksheet: Worksheet, num_header_cols: int, header_end_row: int, footer_start_row: int, debug: bool = False):
        """
        Initialize and immediately capture template state.
        
        Args:
            worksheet: The worksheet to capture state from
            num_header_cols: Number of header columns
            header_end_row: Last row of the header section
            footer_start_row: First row of the footer section (from template)
            debug: Enable debug printing (default: False)
        """
        self.worksheet = worksheet
        self.header_state: List[List[Dict[str, Any]]] = []
        self.footer_state: List[List[Dict[str, Any]]] = []
        self.header_merged_cells: List[str] = []
        self.footer_merged_cells: List[str] = []
        self.row_heights: Dict[int, float] = {}
        self.column_widths: Dict[int, float] = {}
        self.template_footer_start_row: int = footer_start_row
        self.template_footer_end_row: int = -1
        self.header_end_row = header_end_row
        self.min_row = 1
        self.max_row = self.worksheet.max_row
        self.min_col = 1
        self.num_header_cols = num_header_cols
        self.debug = debug or TemplateStateBuilder.DEBUG  # Use instance or class-level debug flag
        
        if self.debug:
            logger.debug(f"TemplateStateBuilder init: worksheet={worksheet.title}, header_end={header_end_row}, footer_start={footer_start_row}")
        
        # Column mapping: template_col_index -> output_col_index
        # Used to shift template content when columns are filtered/removed
        # Default is 1:1 mapping (no shift)
        self.column_mapping: Dict[int, int] = {}
        
        # Log of text replacements performed
        self.replacements_log: List[Dict[str, str]] = []

        # Store default style objects for comparison
        default_workbook = openpyxl.Workbook()
        default_cell = default_workbook.active['A1']
        self.default_font = default_cell.font
        self.default_fill = default_cell.fill
        self.default_border = default_cell.border
        self.default_alignment = default_cell.alignment
        default_workbook.close() # Close the dummy workbook

        # Calculate max_col based on the maximum column with content in the entire worksheet
        max_col_with_content = 0
        max_row_with_content = 0 # Initialize max_row_with_content
        for r_idx in range(1, self.worksheet.max_row + 1):
            for c_idx in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=r_idx, column=c_idx)
                if self._has_content_or_style(cell):
                    max_col_with_content = max(max_col_with_content, c_idx)
                    max_row_with_content = max(max_row_with_content, r_idx) # Update max_row_with_content
        self.max_col = max(max_col_with_content, self.num_header_cols) # Ensure it's at least num_header_cols
        self.max_row = max(max_row_with_content, self.max_row) # Update self.max_row with max_row_with_content
        
        if self.debug:
            logger.debug(f"Template dimensions: max_col={self.max_col}, max_row={self.max_row}, num_header_cols={self.num_header_cols}")
        
        # Capture template state immediately during initialization
        if self.debug:
            logger.debug(f"Capturing template state during init")
            logger.debug(f"Header: rows 1-{header_end_row}, Footer: rows {footer_start_row}-{self.max_row}")
        self._capture_header(header_end_row)
        self._capture_footer(footer_start_row, self.max_row)
        if self.debug:
            logger.debug(f"State captured: {len(self.header_state)} header rows, {len(self.footer_state)} footer rows")
    
    def set_column_mapping(self, mapping: Dict[int, int]):
        """
        Set the column mapping for restoration.
        
        This allows template content to be shifted when columns are filtered/removed.
        For example, if template has 7 columns but output only uses 5 columns,
        the mapping tells us where each template column should go in the output.
        
        Args:
            mapping: Dict mapping template column index (1-based) to output column index (1-based)
                    Example: {1: 1, 2: 2, 3: None, 4: 3, 5: 4, 6: None, 7: 5}
                    None means the column was removed and content should be skipped
        """
        self.column_mapping = mapping
        if self.debug:
            active_mappings = {k: v for k, v in mapping.items() if v is not None}
            skipped_cols = [k for k, v in mapping.items() if v is None]
            logger.debug(f"Column mapping set: {len(active_mappings)} columns mapped")
            logger.debug(f"  Active: {active_mappings}")
            if skipped_cols:
                logger.debug(f"  Skipped template columns: {skipped_cols}")
    
    def _get_mapped_column(self, template_col: int) -> int:
        """
        Get the output column index for a given template column.
        
        Args:
            template_col: Template column index (1-based)
            
        Returns:
            Output column index (1-based), or None if column was removed
        """
        if not self.column_mapping:
            # No mapping set, use 1:1 mapping
            return template_col
        
        return self.column_mapping.get(template_col, template_col)


    def _has_content_or_style(self, cell) -> bool:
        if cell.value is not None and cell.value != '':
            return True
        # Check if any style is applied (not default)
        if cell.font and not self._is_default_style(cell.font, self.default_font): return True
        if cell.fill and not self._is_default_style(cell.fill, self.default_fill): return True
        if cell.border and not self._is_default_style(cell.border, self.default_border): return True
        if cell.alignment and not self._is_default_style(cell.alignment, self.default_alignment): return True
        return False

    def _is_default_style(self, style_obj, default_obj) -> bool:
        if style_obj is None:
            return True
        if default_obj is None: # Should not happen if default_obj is properly initialized
            return False
        
        # Compare relevant attributes for each style type
        if isinstance(style_obj, Font):
            return (
                style_obj.name == default_obj.name and
                style_obj.size == default_obj.size and
                style_obj.bold == default_obj.bold and
                style_obj.italic == default_obj.italic and
                style_obj.underline == default_obj.underline and
                style_obj.strike == default_obj.strike and
                style_obj.color == default_obj.color
            )
        elif isinstance(style_obj, PatternFill):
            return (
                style_obj.fill_type == default_obj.fill_type and
                style_obj.start_color == default_obj.start_color and
                style_obj.end_color == default_obj.end_color
            )
        elif isinstance(style_obj, Border):
            return (
                style_obj.left == default_obj.left and
                style_obj.right == default_obj.right and
                style_obj.top == default_obj.top and
                style_obj.bottom == default_obj.bottom and
                style_obj.diagonal == default_obj.diagonal
            )
        elif isinstance(style_obj, Alignment):
            return (
                style_obj.horizontal == default_obj.horizontal and
                style_obj.vertical == default_obj.vertical and
                style_obj.text_rotation == default_obj.text_rotation and
                style_obj.wrap_text == default_obj.wrap_text and
                style_obj.shrink_to_fit == default_obj.shrink_to_fit and
                style_obj.indent == default_obj.indent
            )
        
        return False # If type not recognized, assume not default

    def _format_cell_style_info(self, cell_info: Dict[str, Any], cell_coord: str) -> str:
        """Format cell styling information for debug logging."""
        parts = []
        
        if cell_info.get('value'):
            val_str = str(cell_info['value'])
            # Sanitize value to avoid encoding errors
            safe_val = val_str.encode('ascii', 'replace').decode('ascii')
            parts.append(f"value='{safe_val[:30]}{'...' if len(safe_val) > 30 else ''}'")
        
        if cell_info.get('font'):
            font = cell_info['font']
            font_parts = []
            if font.name: font_parts.append(f"name={font.name}")
            if font.size: font_parts.append(f"size={font.size}")
            if font.bold: font_parts.append("bold")
            if font.italic: font_parts.append("italic")
            if font.color and hasattr(font.color, 'rgb'):
                font_parts.append(f"color={font.color.rgb}")
            if font_parts:
                parts.append(f"font({', '.join(font_parts)})")
        
        if cell_info.get('fill'):
            fill = cell_info['fill']
            if fill.fill_type and fill.fill_type != 'none':
                fill_str = f"fill({fill.fill_type}"
                if hasattr(fill.start_color, 'rgb'):
                    fill_str += f", {fill.start_color.rgb}"
                fill_str += ")"
                parts.append(fill_str)
        
        if cell_info.get('border'):
            border = cell_info['border']
            border_parts = []
            if border.left and border.left.style: border_parts.append(f"L:{border.left.style}")
            if border.right and border.right.style: border_parts.append(f"R:{border.right.style}")
            if border.top and border.top.style: border_parts.append(f"T:{border.top.style}")
            if border.bottom and border.bottom.style: border_parts.append(f"B:{border.bottom.style}")
            if border_parts:
                parts.append(f"border({', '.join(border_parts)})")
        
        if cell_info.get('alignment'):
            align = cell_info['alignment']
            align_parts = []
            if align.horizontal: align_parts.append(f"h={align.horizontal}")
            if align.vertical: align_parts.append(f"v={align.vertical}")
            if align.wrap_text: align_parts.append("wrap")
            if align_parts:
                parts.append(f"align({', '.join(align_parts)})")
        
        return f"{cell_coord}: {', '.join(parts)}" if parts else None

    def _get_cell_info(self, worksheet, row, col) -> Dict[str, Any]:
        cell = worksheet.cell(row=row, column=col)
        top_left_cell = cell
        for merged_cell_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_cell_range:
                top_left_cell = worksheet.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col)
                break

        return {
            'value': cell.value,
            'font': copy.copy(top_left_cell.font) if top_left_cell.font and not self._is_default_style(top_left_cell.font, self.default_font) else None,
            'fill': copy.copy(top_left_cell.fill) if top_left_cell.fill and not self._is_default_style(top_left_cell.fill, self.default_fill) else None,
            'border': copy.copy(top_left_cell.border) if top_left_cell.border and not self._is_default_style(top_left_cell.border, self.default_border) else None,
            'alignment': copy.copy(top_left_cell.alignment) if top_left_cell.alignment and not self._is_default_style(top_left_cell.alignment, self.default_alignment) else None,
            'number_format': top_left_cell.number_format,
        }

    def _capture_header(self, end_row: int):
        """
        Captures the state of the header section.
        """
        logger.debug(f"=== CAPTURING HEADER (rows 1 to {end_row}) ===")
        
        # Determine the actual start row of the header by finding the first row with content
        header_start_row = 1
        for r_idx in range(1, end_row + 1):
            if any(self._has_content_or_style(self.worksheet.cell(row=r_idx, column=c_idx))
                   for c_idx in range(1, self.max_col + 1)):
                header_start_row = r_idx
                break

        logger.debug(f"  Header starts at row {header_start_row}, ends at row {end_row}")
        logger.debug(f"  Max columns: {self.max_col}")
        
        rows_captured = 0  # Track actual rows captured

        for r_idx in range(header_start_row, end_row + 1):
            rows_captured += 1
            row_data = []
            row_has_content = False
            styled_cells = []  # Track cells with interesting styling
            
            for c_idx in range(1, self.max_col + 1):
                cell_info = self._get_cell_info(self.worksheet, r_idx, c_idx)
                row_data.append(cell_info)
                
                # Debug: Log specific metadata cells (K7:K9 = column 11, rows 7-9)
                if self.debug and c_idx == 11 and r_idx in [7, 8, 9]:
                    col_letter = get_column_letter(c_idx)
                    logger.debug(f"  METADATA CELL {col_letter}{r_idx}: value={cell_info.get('value')}")
                
                # Check if this cell has content
                if cell_info['value'] is not None:
                    row_has_content = True
                
                # Track cells with styling
                if any([cell_info.get('font'), cell_info.get('fill'), cell_info.get('border')]):
                    col_letter = get_column_letter(c_idx)
                    style_str = self._format_cell_style_info(cell_info, f"{col_letter}{r_idx}")
                    if style_str:
                        styled_cells.append(style_str)
            
            self.header_state.append(row_data)
            self.row_heights[r_idx] = self.worksheet.row_dimensions[r_idx].height
            
            # Log row details
            if row_has_content:
                # Show non-empty cells in this row
                non_empty_cells = []
                for c_idx in range(1, self.max_col + 1):
                    cell_val = row_data[c_idx - 1]['value']
                    if cell_val is not None and cell_val != '':
                        col_letter = get_column_letter(c_idx)
                        # Sanitize cell value for logging to avoid encoding errors
                        safe_val = str(cell_val).encode('ascii', 'replace').decode('ascii')[:50]
                        non_empty_cells.append(f"{col_letter}{r_idx}='{safe_val}'")
                
                if non_empty_cells:
                    logger.debug(f"  Row {r_idx}: {', '.join(non_empty_cells[:5])}" + 
                               (f" ... ({len(non_empty_cells)-5} more)" if len(non_empty_cells) > 5 else ""))
                
                # Show styled cells (limit to first 2 to avoid log spam)
                if styled_cells[:2]:
                    for styled_cell in styled_cells[:2]:
                        # Sanitize styled cell string for logging
                        safe_styled = str(styled_cell).encode('ascii', 'replace').decode('ascii')[:200]
                        logger.debug(f"    Style: {safe_styled}")

        # Capture merged cells within the header range
        header_merges = []
        for merged_cell_range in self.worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_cell_range.bounds
            if header_start_row <= min_row <= end_row and header_start_row <= max_row <= end_row:
                merge_str = str(merged_cell_range)
                self.header_merged_cells.append(merge_str)
                header_merges.append(merge_str)

        if header_merges:
            logger.debug(f"  Captured {len(header_merges)} merged cells: {', '.join(header_merges[:3])}" + 
                       (f" ... ({len(header_merges)-3} more)" if len(header_merges) > 3 else ""))

        # Capture column widths
        for c_idx in range(1, self.max_col + 1):
            self.column_widths[c_idx] = self.worksheet.column_dimensions[get_column_letter(c_idx)].width
        
        logger.debug(f"  [OK] Header capture complete: {rows_captured} rows captured (rows {header_start_row}-{end_row}), {len(self.header_merged_cells)} merges")

    def _capture_footer(self, footer_start_row: int, max_possible_footer_row: int):
        """
        Captures the state of the footer section from the original template.
        
        Args:
            footer_start_row: First row of footer in the template
            max_possible_footer_row: Last row to check for footer content
        """
        logger.debug(f"=== CAPTURING FOOTER (starting from row {footer_start_row}) ===")
        logger.debug(f"  Max columns: {self.max_col}, max search row: {max_possible_footer_row}")
        
        # Footer start is already known from parameter
        self.template_footer_start_row = footer_start_row

        # Find the true max row with content by looking for contiguous ACTUAL content (values or merges)
        # Stop after finding N consecutive empty rows (indicates end of footer)
        # Only check for VALUES or MERGES, not just styling (to avoid capturing 180 styled-but-empty rows)
        MAX_EMPTY_ROWS_BEFORE_STOP = 10
        consecutive_empty_rows = 0
        footer_end_row = footer_start_row
        
        for r_idx in range(footer_start_row, min(footer_start_row + 50, max_possible_footer_row + 1)):  # Limit search to 50 rows
            # Check if row has actual content (values) or is part of a merge
            row_has_value = any(self.worksheet.cell(row=r_idx, column=c_idx).value is not None and 
                               self.worksheet.cell(row=r_idx, column=c_idx).value != ''
                               for c_idx in range(1, self.max_col + 1))
            
            row_has_merge = any(r_idx >= merged_range.min_row and r_idx <= merged_range.max_row
                               for merged_range in self.worksheet.merged_cells.ranges)
            
            if row_has_value or row_has_merge:
                footer_end_row = r_idx
                consecutive_empty_rows = 0
            else:
                consecutive_empty_rows += 1
                if consecutive_empty_rows >= MAX_EMPTY_ROWS_BEFORE_STOP:
                    # Found enough empty rows, footer ends here
                    break

        self.template_footer_end_row = footer_end_row
        logger.debug(f"  Footer ends at row {footer_end_row} ({footer_end_row - footer_start_row + 1} footer rows)")

        for r_idx in range(footer_start_row, footer_end_row + 1):
            row_data = []
            row_has_content = False
            styled_cells = []  # Track cells with interesting styling
            
            for c_idx in range(1, self.max_col + 1):
                cell_info = self._get_cell_info(self.worksheet, r_idx, c_idx)
                row_data.append(cell_info)
                
                # Check if this cell has content
                if cell_info['value'] is not None:
                    row_has_content = True
                
                # Track cells with styling
                if any([cell_info.get('font'), cell_info.get('fill'), cell_info.get('border')]):
                    col_letter = get_column_letter(c_idx)
                    style_str = self._format_cell_style_info(cell_info, f"{col_letter}{r_idx}")
                    if style_str:
                        styled_cells.append(style_str)
            
            self.footer_state.append(row_data)
            self.row_heights[r_idx] = self.worksheet.row_dimensions[r_idx].height
            
            # Log row details
            if row_has_content:
                # Show non-empty cells in this row
                non_empty_cells = []
                for c_idx in range(1, self.max_col + 1):
                    cell_val = row_data[c_idx - 1]['value']
                    if cell_val is not None and cell_val != '':
                        col_letter = get_column_letter(c_idx)
                        # Sanitize cell value for logging to avoid encoding errors
                        safe_val = str(cell_val).encode('ascii', 'replace').decode('ascii')[:50]
                        non_empty_cells.append(f"{col_letter}{r_idx}='{safe_val}'")
                
                if non_empty_cells:
                    logger.debug(f"  Row {r_idx}: {', '.join(non_empty_cells[:5])}" + 
                               (f" ... ({len(non_empty_cells)-5} more)" if len(non_empty_cells) > 5 else ""))
                
                # Show styled cells (limit to first 2 to avoid log spam)
                if styled_cells[:2]:
                    for styled_cell in styled_cells[:2]:
                        # Sanitize styled cell string for logging
                        safe_styled = str(styled_cell).encode('ascii', 'replace').decode('ascii')[:200]
                        logger.debug(f"    Style: {safe_styled}")

        # Capture merged cells within the footer range
        footer_merges = []
        for merged_cell_range in self.worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_cell_range.bounds
            if footer_start_row <= min_row <= footer_end_row and footer_start_row <= max_row <= footer_end_row:
                merge_str = str(merged_cell_range)
                self.footer_merged_cells.append(merge_str)
                footer_merges.append(merge_str)

        if footer_merges:
            logger.debug(f"  Captured {len(footer_merges)} merged cells: {', '.join(footer_merges[:3])}" + 
                       (f" ... ({len(footer_merges)-3} more)" if len(footer_merges) > 3 else ""))

        # Capture column widths
        for c_idx in range(1, self.max_col + 1):
            self.column_widths[c_idx] = self.worksheet.column_dimensions[get_column_letter(c_idx)].width
        
        # Validate footer capture - warn if all rows are empty
        total_non_empty_cells = sum(
            1 for row_data in self.footer_state 
            for cell_info in row_data 
            if cell_info['value'] is not None and cell_info['value'] != ''
        )
        
        if total_non_empty_cells == 0:
            logger.debug(f"Template footer capture: {len(self.footer_state)} rows (all blank/empty)")
            logger.debug(f"   This is OK - blank footer rows will be preserved and restored")
        
        logger.debug(f"  [OK] Footer capture complete: {len(self.footer_state)} rows, {len(self.footer_merged_cells)} merges, template footer start: {self.template_footer_start_row}, non-empty cells: {total_non_empty_cells}")

    def restore_header_only(self, target_worksheet: Worksheet, actual_num_cols: int = None):
        """
        Restores ONLY the header (structure, values, merges, formatting) to a new clean worksheet.
        This is used when creating a fresh worksheet to avoid template footer conflicts.
        
        Uses column mapping if set to shift content when columns are filtered/removed.
        
        Args:
            target_worksheet: The worksheet to restore header to
            actual_num_cols: If provided, stretch header to this many columns (for dynamic column scenarios)
        """
        if self.debug:
            logger.debug(f"Restoring header to new worksheet")
            logger.debug(f"Header rows: {len(self.header_state)}, Header merges: {len(self.header_merged_cells)}")
            if actual_num_cols:
                template_cols = self.max_col - self.min_col + 1
                logger.debug(f"Dynamic column stretch: Template={template_cols} cols, Actual={actual_num_cols} cols")
            if self.column_mapping:
                logger.debug(f"Using column mapping to shift template content")
        
        # Calculate template column count and actual target
        template_num_cols = self.max_col - self.min_col + 1
        target_num_cols = actual_num_cols if actual_num_cols else template_num_cols
        
        # Restore header cell values and formatting
        for row_idx, row_data in enumerate(self.header_state):
            actual_row = row_idx + self.min_row
            
            # Restore template columns with column mapping
            for col_idx, cell_info in enumerate(row_data):
                template_col = col_idx + self.min_col
                
                # Get mapped column (may be shifted or None if removed)
                output_col = self._get_mapped_column(template_col)
                
                # Handle removed columns with content - try to shift left first, then right
                if output_col is None:
                    if cell_info['value'] is not None:
                        # Column removed but has content - try to save it
                        shifted_col = None
                        
                        # Priority 1: Try shift LEFT (to previous column)
                        if col_idx > 0:  # Has a left neighbor
                            left_cell = row_data[col_idx - 1]
                            if left_cell['value'] is None:  # Left is empty
                                shifted_col = self._get_mapped_column(template_col - 1)
                                if self.debug:
                                    logger.debug(f"  Shifting removed column {template_col} content LEFT to {template_col-1} (row {actual_row}, value: '{cell_info['value']}')")
                        
                        # Priority 2: Try shift RIGHT (to next column) if left failed
                        if shifted_col is None and col_idx < len(row_data) - 1:  # Has a right neighbor
                            right_cell = row_data[col_idx + 1]
                            if right_cell['value'] is None:  # Right is empty
                                shifted_col = self._get_mapped_column(template_col + 1)
                                if self.debug:
                                    logger.debug(f"  Shifting removed column {template_col} content RIGHT to {template_col+1} (row {actual_row}, value: '{cell_info['value']}')")
                        
                        # HALT if cannot shift anywhere
                        if shifted_col is None:
                            left_status = "has content" if col_idx > 0 and row_data[col_idx - 1]['value'] is not None else "N/A"
                            right_status = "has content" if col_idx < len(row_data) - 1 and row_data[col_idx + 1]['value'] is not None else "N/A"
                            raise ValueError(
                                f"CRITICAL: Cannot remove column {template_col} at row {actual_row}!\n"
                                f"  Column has content: '{cell_info['value']}'\n"
                                f"  Cannot shift LEFT (col {template_col-1}): {left_status}\n"
                                f"  Cannot shift RIGHT (col {template_col+1}): {right_status}\n"
                                f"  Solution: Either remove content from neighboring columns or don't use skip_in_daf flag on this column."
                            )
                        
                        output_col = shifted_col
                    else:
                        # Column removed and empty - safe to skip
                        if self.debug:
                            logger.debug(f"  Skipping removed column {template_col} at row {actual_row} (empty)")
                        continue
                
                target_cell = target_worksheet.cell(row=actual_row, column=output_col)
                
                # Restore value
                if cell_info['value'] is not None:
                    target_cell.value = cell_info['value']
                    if self.debug and template_col != output_col:
                        logger.debug(f"  Shifted column {template_col} -> {output_col} at row {actual_row} (value: '{cell_info['value']}')")
                
                # Restore formatting
                if cell_info['font']:
                    target_cell.font = copy.copy(cell_info['font'])
                if cell_info['fill']:
                    target_cell.fill = copy.copy(cell_info['fill'])
                if cell_info['border']:
                    target_cell.border = copy.copy(cell_info['border'])
                if cell_info['alignment']:
                    target_cell.alignment = copy.copy(cell_info['alignment'])
                if cell_info['number_format']:
                    target_cell.number_format = cell_info['number_format']
            
            # If we need more columns than template had, extend the last column's styling
            if target_num_cols > template_num_cols:
                last_template_col_idx = len(row_data) - 1
                last_template_cell_info = row_data[last_template_col_idx]
                
                # Extend from template edge to new edge
                for extra_col_idx in range(template_num_cols, target_num_cols):
                    actual_col = extra_col_idx + self.min_col
                    target_cell = target_worksheet.cell(row=actual_row, column=actual_col)
                    
                    # Copy styling from template's last column (but not value)
                    if last_template_cell_info['font']:
                        target_cell.font = copy.copy(last_template_cell_info['font'])
                    if last_template_cell_info['fill']:
                        target_cell.fill = copy.copy(last_template_cell_info['fill'])
                    if last_template_cell_info['border']:
                        target_cell.border = copy.copy(last_template_cell_info['border'])
                    if last_template_cell_info['alignment']:
                        target_cell.alignment = copy.copy(last_template_cell_info['alignment'])
                    if last_template_cell_info['number_format']:
                        target_cell.number_format = last_template_cell_info['number_format']
        
        # Restore header merged cells with column mapping
        for merged_cell_range_str in self.header_merged_cells:
            try:
                # Apply column mapping to merged cells if mapping is set
                if self.column_mapping:
                    from openpyxl.utils.cell import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(merged_cell_range_str)
                    
                    # Map the columns
                    mapped_min_col = self._get_mapped_column(min_col)
                    mapped_max_col = self._get_mapped_column(max_col)
                    
                    # Skip if either column was removed
                    if mapped_min_col is None or mapped_max_col is None:
                        if self.debug:
                            logger.debug(f"Skipping merge {merged_cell_range_str} (columns removed)")
                        continue
                    
                    # Create adjusted merge range
                    adjusted_range_str = f"{get_column_letter(mapped_min_col)}{min_row}:{get_column_letter(mapped_max_col)}{max_row}"
                    target_worksheet.merge_cells(adjusted_range_str)
                    if self.debug and merged_cell_range_str != adjusted_range_str:
                        logger.debug(f"Merged (shifted): {merged_cell_range_str} -> {adjusted_range_str}")
                    elif self.debug:
                        logger.debug(f"Merged: {adjusted_range_str}")
                else:
                    # No mapping, use original
                    target_worksheet.merge_cells(merged_cell_range_str)
                    if self.debug:
                        logger.debug(f"Merged: {merged_cell_range_str}")
            except Exception as e:
                if self.debug:
                    logger.warning(f"Could not merge {merged_cell_range_str}: {e}")
        
        # Restore row heights
        for row_num, height in self.row_heights.items():
            if row_num <= self.header_end_row and height:
                target_worksheet.row_dimensions[row_num].height = height
        
        # Restore column widths
        for col_num, width in self.column_widths.items():
            if width:
                target_worksheet.column_dimensions[get_column_letter(col_num)].width = width
        
        if self.debug:
            logger.debug(f"Header restoration complete")

    def restore_footer_only(self, target_worksheet: Worksheet, footer_start_row: int, actual_num_cols: int = None, restore_footer_merges: bool = True):
        """
        Restores ONLY the footer (structure, values, merges, formatting) to the new worksheet.
        This places the template footer (static content) AFTER the dynamically created data footer.
        
        Uses column mapping if set to shift content when columns are filtered/removed.
        
        Args:
            target_worksheet: The worksheet to restore footer to
            footer_start_row: The row where the template footer should start (after data footer)
            actual_num_cols: If provided, stretch footer to this many columns (for dynamic column scenarios)
            restore_footer_merges: Whether to restore footer merged cells (default True)
        """
        logger.debug(f"restore_footer_only called with:")
        logger.debug(f"footer_start_row parameter: {footer_start_row}")
        logger.debug(f"self.template_footer_start_row: {self.template_footer_start_row}")
        logger.debug(f"self.template_footer_end_row: {self.template_footer_end_row}")
        logger.debug(f"len(self.footer_state): {len(self.footer_state)}")
        
        if self.debug:
            logger.debug(f"Restoring template footer starting at row {footer_start_row}")
            logger.debug(f"Template footer rows: {len(self.footer_state)}, Footer merges: {len(self.footer_merged_cells)}")
            if actual_num_cols:
                template_cols = self.max_col - self.min_col + 1
                logger.debug(f"Dynamic column stretch: Template={template_cols} cols, Actual={actual_num_cols} cols")
            if self.column_mapping:
                logger.debug(f"Using column mapping to shift template content")
        
        # Calculate offset: template footer was at self.template_footer_start_row, now goes to footer_start_row
        offset = footer_start_row - self.template_footer_start_row if self.template_footer_start_row > 0 else 0
        logger.debug(f"Calculated offset: {offset} (footer_start_row={footer_start_row} - template_footer_start_row={self.template_footer_start_row})")
        logger.debug(f"offset: {offset}")
        
        # Calculate template column count and actual target
        template_num_cols = self.max_col - self.min_col + 1
        target_num_cols = actual_num_cols if actual_num_cols else template_num_cols
        
        # Restore footer cell values and formatting with offset and column mapping
        for row_idx, row_data in enumerate(self.footer_state):
            actual_row = self.template_footer_start_row + row_idx + offset
            
            # Restore template columns with column mapping
            for col_idx, cell_info in enumerate(row_data):
                template_col = col_idx + self.min_col
                
                # Get mapped column (may be shifted or None if removed)
                output_col = self._get_mapped_column(template_col)
                
                # Handle removed columns with content - try to shift left first, then right
                if output_col is None:
                    if cell_info['value'] is not None:
                        # Column removed but has content - try to save it
                        shifted_col = None
                        
                        # Priority 1: Try shift LEFT (to previous column)
                        if col_idx > 0:  # Has a left neighbor
                            left_cell = row_data[col_idx - 1]
                            if left_cell['value'] is None:  # Left is empty
                                shifted_col = self._get_mapped_column(template_col - 1)
                                if self.debug:
                                    logger.debug(f"  Shifting removed column {template_col} content LEFT to {template_col-1} (row {actual_row}, value: '{cell_info['value']}')")
                        
                        # Priority 2: Try shift RIGHT (to next column) if left failed
                        if shifted_col is None and col_idx < len(row_data) - 1:  # Has a right neighbor
                            right_cell = row_data[col_idx + 1]
                            if right_cell['value'] is None:  # Right is empty
                                shifted_col = self._get_mapped_column(template_col + 1)
                                if self.debug:
                                    logger.debug(f"  Shifting removed column {template_col} content RIGHT to {template_col+1} (row {actual_row}, value: '{cell_info['value']}')")
                        
                        # HALT if cannot shift anywhere
                        if shifted_col is None:
                            left_status = "has content" if col_idx > 0 and row_data[col_idx - 1]['value'] is not None else "N/A"
                            right_status = "has content" if col_idx < len(row_data) - 1 and row_data[col_idx + 1]['value'] is not None else "N/A"
                            raise ValueError(
                                f"CRITICAL: Cannot remove column {template_col} at row {actual_row}!\n"
                                f"  Column has content: '{cell_info['value']}'\n"
                                f"  Cannot shift LEFT (col {template_col-1}): {left_status}\n"
                                f"  Cannot shift RIGHT (col {template_col+1}): {right_status}\n"
                                f"  Solution: Either remove content from neighboring columns or don't use skip_in_daf flag on this column."
                            )
                        
                        output_col = shifted_col
                    else:
                        # Column removed and empty - safe to skip
                        if self.debug:
                            logger.debug(f"  Skipping removed column {template_col} at row {actual_row} (empty)")
                        continue
                
                logger.debug(f"actual_row: {actual_row}, template_col: {template_col}, output_col: {output_col}")
                target_cell = target_worksheet.cell(row=actual_row, column=output_col)
                
                # Restore value
                if cell_info['value'] is not None:
                    target_cell.value = cell_info['value']
                    if self.debug and template_col != output_col:
                        logger.debug(f"  Shifted column {template_col} -> {output_col} at row {actual_row} (value: '{cell_info['value']}')")
                
                # Restore formatting
                if cell_info['font']:
                    target_cell.font = copy.copy(cell_info['font'])
                if cell_info['fill']:
                    target_cell.fill = copy.copy(cell_info['fill'])
                if cell_info['border']:
                    target_cell.border = copy.copy(cell_info['border'])
                if cell_info['alignment']:
                    target_cell.alignment = copy.copy(cell_info['alignment'])
                if cell_info['number_format']:
                    target_cell.number_format = cell_info['number_format']
            
            # If we need more columns than template had, extend the last column's styling
            if target_num_cols > template_num_cols:
                last_template_col_idx = len(row_data) - 1
                last_template_cell_info = row_data[last_template_col_idx]
                
                # Extend from template edge to new edge
                for extra_col_idx in range(template_num_cols, target_num_cols):
                    actual_col = extra_col_idx + self.min_col
                    target_cell = target_worksheet.cell(row=actual_row, column=actual_col)
                    
                    # Copy styling from template's last column (but not value)
                    if last_template_cell_info['font']:
                        target_cell.font = copy.copy(last_template_cell_info['font'])
                    if last_template_cell_info['fill']:
                        target_cell.fill = copy.copy(last_template_cell_info['fill'])
                    if last_template_cell_info['border']:
                        target_cell.border = copy.copy(last_template_cell_info['border'])
                    if last_template_cell_info['alignment']:
                        target_cell.alignment = copy.copy(last_template_cell_info['alignment'])
                    if last_template_cell_info['number_format']:
                        target_cell.number_format = last_template_cell_info['number_format']
        
        # Restore footer merged cells with offset and column mapping
        for merged_cell_range_str in self.footer_merged_cells:
            try:
                from openpyxl.utils.cell import range_boundaries
                min_col, min_row, max_col, max_row = range_boundaries(merged_cell_range_str)
                original_span = max_col - min_col + 1  # Calculate original column span
                
                # Apply column mapping if set
                if self.column_mapping:
                    mapped_min_col = self._get_mapped_column(min_col)
                    mapped_max_col = self._get_mapped_column(max_col)
                    
                    # If either column was removed, find the nearest valid columns
                    if mapped_min_col is None or mapped_max_col is None:
                        # Find first valid column at or after min_col
                        for col in range(min_col, max(self.column_mapping.keys()) + 1):
                            mapped_col = self._get_mapped_column(col)
                            if mapped_col is not None:
                                mapped_min_col = mapped_col
                                break
                        
                        # Find last valid column at or before max_col
                        for col in range(max_col, min_col - 1, -1):
                            mapped_col = self._get_mapped_column(col)
                            if mapped_col is not None:
                                mapped_max_col = mapped_col
                                break
                        
                        # If still no valid range, skip this merge
                        if mapped_min_col is None or mapped_max_col is None or mapped_min_col > mapped_max_col:
                            if self.debug:
                                logger.debug(f"Skipping footer merge {merged_cell_range_str} (no valid columns after mapping)")
                            continue
                        
                        if self.debug:
                            logger.debug(f"Adjusted footer merge {merged_cell_range_str}: cols {min_col}-{max_col} -> {mapped_min_col}-{mapped_max_col}")
                    
                    # Preserve the original span - extend max_col to maintain visual width
                    min_col = mapped_min_col
                    max_col = mapped_min_col + original_span - 1  # Maintain original span
                else:
                    # No mapping, keep original positions
                    pass
                
                # Adjust row numbers with offset
                min_row += offset
                max_row += offset
                adjusted_range_str = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                target_worksheet.merge_cells(adjusted_range_str)
                if self.debug:
                    if merged_cell_range_str != adjusted_range_str:
                        logger.debug(f"Merged (shifted): {merged_cell_range_str} -> {adjusted_range_str}")
                    else:
                        logger.debug(f"Merged: {adjusted_range_str}")
            except Exception as e:
                if self.debug:
                    logger.warning(f"Could not merge {merged_cell_range_str}: {e}")
        
        # Restore row heights for footer rows
        for row_num, height in self.row_heights.items():
            if self.template_footer_start_row <= row_num <= self.template_footer_end_row and height:
                target_worksheet.row_dimensions[row_num + offset].height = height
        
        if self.debug:
            logger.debug(f"Template footer restoration complete ({len(self.footer_state)} rows restored)")

    def restore_state(self, target_worksheet: Worksheet, data_start_row: int, data_table_end_row: int, restore_footer_merges: bool = True):
        """
        Restores the captured FORMATTING (not values) to preserve template structure.
        Only restores merges, heights, widths - does NOT overwrite cell values.
        
        Args:
            target_worksheet: The worksheet to restore state to
            data_start_row: Starting row of data
            data_table_end_row: Ending row of data table
            restore_footer_merges: Whether to restore footer merges (False when FooterBuilder creates its own merges)
        """
        if self.debug:
            logger.debug(f"Restoring formatting (merges, heights, widths):")
            logger.debug(f"Header merges: {len(self.header_merged_cells)}")
            logger.debug(f"Footer merges: {len(self.footer_merged_cells)} (restore: {restore_footer_merges})")
            logger.debug(f"Template footer start row: {self.template_footer_start_row}")
            logger.debug(f"Data table end row: {data_table_end_row}")
        
        # Restore header merged cells without offset
        if self.debug:
            logger.debug(f"Restoring {len(self.header_merged_cells)} header merges...")
        for merged_cell_range_str in self.header_merged_cells:
            try:
                target_worksheet.merge_cells(merged_cell_range_str)
                if self.debug:
                    logger.debug(f"Merged: {merged_cell_range_str}")
            except Exception as e:
                if self.debug:
                    logger.warning(f"Could not merge {merged_cell_range_str}: {e}")

        # Calculate the offset for footer rows and merged cells
        footer_start_row_in_new_sheet = data_table_end_row + 1
        offset = footer_start_row_in_new_sheet - self.template_footer_start_row if self.template_footer_start_row != -1 else 0
        
        if self.debug:
            logger.debug(f"Footer offset: {offset} (template row {self.template_footer_start_row} -> new row {footer_start_row_in_new_sheet})")

        # Restore footer merged cells with offset (only if requested)
        if restore_footer_merges:
            if self.debug:
                logger.debug(f"Restoring {len(self.footer_merged_cells)} footer merges with offset {offset}...")
            for merged_cell_range_str in self.footer_merged_cells:
                try:
                    from openpyxl.utils.cell import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(merged_cell_range_str)
                    original_span = max_col - min_col + 1  # Calculate original column span
                    
                    # Apply column mapping if set
                    if self.column_mapping:
                        mapped_min_col = self._get_mapped_column(min_col)
                        mapped_max_col = self._get_mapped_column(max_col)
                        
                        # If either column was removed, find the nearest valid columns
                        if mapped_min_col is None or mapped_max_col is None:
                            # Find first valid column at or after min_col
                            for col in range(min_col, max(self.column_mapping.keys()) + 1):
                                mapped_col = self._get_mapped_column(col)
                                if mapped_col is not None:
                                    mapped_min_col = mapped_col
                                    break
                            
                            # Find last valid column at or before max_col
                            for col in range(max_col, min_col - 1, -1):
                                mapped_col = self._get_mapped_column(col)
                                if mapped_col is not None:
                                    mapped_max_col = mapped_col
                                    break
                            
                            # If still no valid range, skip this merge
                            if mapped_min_col is None or mapped_max_col is None or mapped_min_col > mapped_max_col:
                                if self.debug:
                                    logger.debug(f"Skipping footer merge {merged_cell_range_str} (no valid columns after mapping)")
                                continue
                            
                            if self.debug:
                                logger.debug(f"Adjusted footer merge {merged_cell_range_str}: cols {min_col}-{max_col} -> {mapped_min_col}-{mapped_max_col}")
                        
                        # Preserve the original span - extend max_col to maintain visual width
                        min_col = mapped_min_col
                        max_col = mapped_min_col + original_span - 1  # Maintain original span
                    else:
                        # No mapping, keep original positions
                        pass
                    
                    # Adjust row numbers for all footer merged cells
                    min_row += offset
                    max_row += offset
                    adjusted_range_str = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                    target_worksheet.merge_cells(adjusted_range_str)
                    if self.debug:
                        if merged_cell_range_str != adjusted_range_str:
                            logger.debug(f"Merged (shifted): {merged_cell_range_str} -> {adjusted_range_str}")
                        else:
                            logger.debug(f"Merged: {adjusted_range_str}")
                except Exception as e:
                    if self.debug:
                        logger.warning(f"Could not merge {merged_cell_range_str}: {e}")
        else:
            if self.debug:
                logger.debug(f"Skipping footer merge restoration (FooterBuilder creates its own merges)")

        # Restore row heights for header
        if self.debug:
            logger.debug(f"Restoring row heights...")
        current_row = 1
        for row_data in self.header_state:
            target_worksheet.row_dimensions[current_row].height = self.row_heights.get(current_row, None)
            current_row += 1

        # Restore row heights for footer (with offset)
        for r_offset, row_data in enumerate(self.footer_state):
            r_idx = footer_start_row_in_new_sheet + r_offset
            original_footer_row_idx = self.template_footer_start_row + r_offset
            target_worksheet.row_dimensions[r_idx].height = self.row_heights.get(original_footer_row_idx, None)

        # Restore column widths
        if self.debug:
            logger.debug(f"Restoring column widths...")
        for c_idx, width in self.column_widths.items():
            target_worksheet.column_dimensions[get_column_letter(c_idx)].width = width
        
        if self.debug:
            logger.debug(f"Formatting restoration complete!")

    def apply_text_replacements(self, replacement_rules: list, invoice_data: dict = None) -> int:
        """
        Apply text replacements to the stored template state (header and footer).
        This modifies the stored 'value' and 'number_format' before restoration.
        
        Args:
            replacement_rules: List of replacement rule dicts with keys:
                - find: Text to find
                - replace: Text to replace with (for hardcoded replacements)
                - data_path: Path to data in invoice_data (for data-driven replacements)
                - is_date: If True, format as date (optional)
                - match_mode: 'exact' or 'substring' or 'contains' (optional, default='exact')
            invoice_data: Invoice data dict for data-driven replacements
        
        Returns:
            Number of replacements made
        """
        if self.debug:
            logger.debug(f"=== apply_text_replacements START ===")
            # Count formulas in header state
            formula_count = 0
            for row_data in self.header_state:
                for cell_info in row_data:
                    val = cell_info.get('value')
                    if val and isinstance(val, str) and val.startswith('='):
                        formula_count += 1
                        if formula_count <= 5:  # Log first 5
                            logger.debug(f"  Formula in header: {val}")
            logger.debug(f"  Total formulas in header: {formula_count}")
        
        changes_made = 0
        
        # Apply to header state
        for row_data in self.header_state:
            for cell_info in row_data:
                if cell_info.get('value') and isinstance(cell_info['value'], str):
                    original_value = cell_info['value']
                    
                    # Debug: Log every cell being processed
                    if self.debug and ('=' in original_value or 'Packing list' in original_value):
                        logger.debug(f"[HEADER] Processing cell with value: '{original_value[:80]}'")
                    
                    # Check if it's a formula (starts with =)
                    if original_value.startswith('='):
                        # Try to extract placeholder from formula
                        # E.g., "='Packing list'!J7" might refer to a cell with "JFINV"
                        # For now, we'll handle simple cases - can be enhanced later
                        if self.debug:
                            logger.debug(f"Found formula in header: {original_value}")
                    
                    new_value, format_changed, matched_term = self._apply_rules_to_cell(
                        cell_info['value'], 
                        replacement_rules, 
                        invoice_data
                    )
                    if new_value != original_value:
                        cell_info['value'] = new_value
                        if format_changed:
                            # Reset number_format to General for text replacements
                            cell_info['number_format'] = 'General'
                        changes_made += 1
                        
                        # Log the replacement
                        if matched_term:
                            self.replacements_log.append({
                                "original": original_value,
                                "new": new_value,
                                "term": matched_term,
                                "location": "header"
                            })
        
        # Apply to footer state
        for row_data in self.footer_state:
            for cell_info in row_data:
                if cell_info.get('value') and isinstance(cell_info['value'], str):
                    original_value = cell_info['value']
                    
                    # Check if it's a formula (starts with =)
                    if original_value.startswith('='):
                        if self.debug:
                            logger.debug(f"Found formula in footer: {original_value}")
                    
                    new_value, format_changed, matched_term = self._apply_rules_to_cell(
                        cell_info['value'], 
                        replacement_rules, 
                        invoice_data
                    )
                    if new_value != original_value:
                        cell_info['value'] = new_value
                        if format_changed:
                            # Reset number_format to General for text replacements
                            cell_info['number_format'] = 'General'
                        changes_made += 1
                        
                        # Log the replacement
                        if matched_term:
                            self.replacements_log.append({
                                "original": original_value,
                                "new": new_value,
                                "term": matched_term,
                                "location": "footer"
                            })
        
        logger.info(f"[TemplateStateBuilder] Text replacements complete: {changes_made} changes made")
        return changes_made
    
    def _apply_rules_to_cell(self, text: str, rules: list, invoice_data: dict = None) -> tuple:
        """
        Apply replacement rules to a single text value.
        
        Handles both direct text and formulas that reference cells with placeholders.
        For formulas like "='Packing list'!J7", if J7 contains "JFINV", 
        we replace the formula with the actual data value.
        
        Returns:
            Tuple of (new_value, format_changed, matched_term)
            - new_value: The replaced text
            - format_changed: True if number_format should be reset to General
            - matched_term: The term that triggered the replacement (e.g., "FCA")
        """
        if not text:
            return text, False, None
        
        # Original text replacement logic
        for rule in rules:
            find_text = rule.get('find', '')
            if not find_text:
                continue
            
            match_mode = rule.get('match_mode', 'exact')
            is_match = False
            
            if match_mode == 'exact':
                is_match = (text.strip() == find_text)
            elif match_mode in ['substring', 'contains']:
                is_match = (find_text in text)
            
            if is_match:
                # Get replacement value
                replacement_value = None
                is_date = rule.get('is_date', False)
                
                if 'data_path' in rule and invoice_data:
                    replacement_value = self._resolve_data_path(invoice_data, rule['data_path'])
                    if self.debug:
                        logger.debug(f"Data path {rule['data_path']} resolved to: {replacement_value}")
                    if replacement_value is not None and is_date:
                        # Format date value
                        replacement_value = self._format_date_value(replacement_value)
                elif 'replace' in rule:
                    replacement_value = rule['replace']
                    if self.debug:
                        logger.debug(f"Using hardcoded replacement: '{replacement_value}' for '{find_text}'")
                
                if replacement_value is not None:
                    # Perform replacement
                    # Perform replacement
                    if match_mode == 'exact':
                        if self.debug:
                            logger.debug(f"Replacing '{find_text}' with '{replacement_value}'")
                        return str(replacement_value), not is_date, find_text  # format_changed = True unless it's a date
                    else:  # substring/contains
                        if self.debug:
                            logger.debug(f"Replacing substring '{find_text}' with '{replacement_value}' in '{text}'")
                        return text.replace(find_text, str(replacement_value)), True, find_text
        
        return text, False, None
    
    def _resolve_data_path(self, data: dict, path: list) -> Any:
        """Resolve nested data path like ["processed_tables_data", "1", "col_inv_no", 0]"""
        current = data
        try:
            for key in path:
                if isinstance(current, dict):
                    current = current.get(key)
                elif isinstance(current, list):
                    current = current[int(key)]
                else:
                    return None
                if current is None:
                    return None
            return current
        except (KeyError, IndexError, ValueError, TypeError):
            return None
    
    def _format_date_value(self, value: Any) -> str:
        """Format date value for display"""
        import datetime
        if isinstance(value, (datetime.datetime, datetime.date)):
            return value.strftime('%d/%m/%Y')
        elif isinstance(value, str):
            # Try to parse and format
            try:
                from dateutil.parser import parse
                parsed = parse(value, dayfirst=True)
                return parsed.strftime('%d/%m/%Y')
            except:
                return str(value)
        return str(value)
