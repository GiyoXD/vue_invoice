import logging
from typing import Any, Dict, List, Optional, Tuple, Union
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import traceback

logger = logging.getLogger(__name__)

from ..data.data_preparer import prepare_data_rows, parse_mapping_rules
from ..utils.layout import apply_column_widths, merge_contiguous_cells_by_id
from ..styling.style_applier import apply_row_heights
from ..utils.merge_utils import merge_vertical_cells_in_range, apply_horizontal_merge_by_id
# Legacy apply_cell_style removed - using only StyleRegistry + CellStyler
from ..styling.style_registry import StyleRegistry
from ..styling.cell_styler import CellStyler
# FooterBuilder is now called by LayoutBuilder (proper Director pattern)
from ..styling.style_config import THIN_BORDER, NO_BORDER, CENTER_ALIGNMENT, LEFT_ALIGNMENT, BOLD_FONT, FORMAT_GENERAL, FORMAT_TEXT, FORMAT_NUMBER_COMMA_SEPARATED1, FORMAT_NUMBER_COMMA_SEPARATED2



from ..styling.models import StylingConfigModel
from .bundle_accessor import BundleAccessor

class DataTableBuilderStyler:
    """
    Builds and styles data table sections based on pre-resolved data.
    
    This class is a "dumb" builder. Its only job is to take prepared data
    and write it to the worksheet. It does not contain any data-sourcing
    or mapping logic.
    """
    
    def __init__(
        self,
        worksheet: Worksheet,
        header_info: Dict[str, Any],
        resolved_data: Dict[str, Any],
        sheet_styling_config: Optional[StylingConfigModel] = None,
        vertical_merge_columns: Optional[List[str]] = None
    ):
        """
        Initialize the builder with resolved data.
        
        Args:
            worksheet: The worksheet to write to.
            header_info: Header information with column maps.
            resolved_data: The data prepared by TableDataAdapter.
            sheet_styling_config: The styling configuration for the sheet.
        """
        self.worksheet = worksheet
        self.header_info = header_info
        self.resolved_data = resolved_data
        self.sheet_styling_config = sheet_styling_config
        self.vertical_merge_columns = vertical_merge_columns or []

        # Extract commonly used values
        self.data_rows = resolved_data.get('data_rows', [])
        self.static_info = resolved_data.get('static_info', {})
        self.formula_rules = resolved_data.get('formula_rules', {})
        self.pallet_counts = resolved_data.get('pallet_counts', [])
        self.dynamic_desc_used = resolved_data.get('dynamic_desc_used', False)
        
        self.col_id_map = header_info.get('column_id_map', {})
        self.idx_to_id_map = {v: k for k, v in self.col_id_map.items()}
        self.column_colspan = header_info.get('column_colspan', {})  # Colspan for automatic merging
        
        # Initialize StyleRegistry and CellStyler for ID-driven styling
        self.style_registry = None
        self.cell_styler = CellStyler()
        if sheet_styling_config:
            try:
                styling_dict = sheet_styling_config.model_dump() if hasattr(sheet_styling_config, 'model_dump') else sheet_styling_config
                if isinstance(styling_dict, dict) and 'columns' in styling_dict and 'row_contexts' in styling_dict:
                    self.style_registry = StyleRegistry(styling_dict)
                    logger.info("StyleRegistry initialized successfully for DataTableBuilder")
                else:
                    logger.error(f"DataTableBuilder: Invalid styling config format. Expected 'columns' and 'row_contexts'.")
                    raise ValueError("Invalid styling config format")
            except Exception as e:
                logger.error(f"Could not initialize StyleRegistry: {e}")
                raise
        else:
            logger.error("DataTableBuilder: No styling config provided!")
            raise ValueError("No styling config provided")
        
        # Static content is now injected into data_rows by TableDataResolver
        # No need to handle it separately here
        logger.debug(f"DataTableBuilder initialized with {len(self.data_rows)} total rows (including any static rows)")
        
        # Track rows that have had height applied to avoid redundant operations
        self._rows_with_height_applied = set()




    def build(self) -> bool:
        if not self.header_info or 'second_row_index' not in self.header_info:
            logger.error("Invalid header_info provided to DataTableBuilderStyler")
            return False

        num_columns = self.header_info.get('num_columns', 0)
        data_writing_start_row = self.header_info.get('second_row_index', 0) + 1
        
        actual_rows_to_process = len(self.data_rows)
        
        data_start_row = data_writing_start_row
        data_end_row = data_start_row + actual_rows_to_process - 1 if actual_rows_to_process > 0 else data_start_row - 1
        
        # --- Fill Data Rows Loop ---
        try:
            data_row_indices_written = []
            for i in range(actual_rows_to_process):
                current_row_idx = data_start_row + i
                data_row_indices_written.append(current_row_idx)
                
                row_data = self.data_rows[i]
                
                # Filter row_data to only include columns in the filtered column_id_map
                # This removes columns that were filtered by skip_in_daf or skip_in_custom
                valid_col_indices = set(self.col_id_map.values())
                row_data = {col_idx: value for col_idx, value in row_data.items() if col_idx in valid_col_indices}
                
                # First, write columns that have data
                columns_with_data = set(row_data.keys())

                # Write all columns for this row (including static if present in row_data)
                for col_idx, value in row_data.items():
                    cell = self.worksheet.cell(row=current_row_idx, column=col_idx)
                    if not isinstance(cell, MergedCell):
                        # Check if value is a formula dict
                        if isinstance(value, dict) and value.get('type') == 'formula':
                            # Convert formula dict to Excel formula string
                            formula_str = self._build_formula_string(value, current_row_idx)
                            cell.value = formula_str
                        else:
                            # Try to convert string numbers to actual numbers for Excel
                            if isinstance(value, str):
                                # Convert empty strings to None to avoid ' in Excel
                                if not value.strip():
                                    cell.value = None
                                else:
                                    try:
                                        # Try converting to float first
                                        float_val = float(value)
                                        # If it's an integer (e.g. 10.0), convert to int
                                        if float_val.is_integer():
                                            cell.value = int(float_val)
                                        else:
                                            cell.value = float_val
                                    except (ValueError, TypeError):
                                        # Keep as string if conversion fails
                                        cell.value = value
                            else:
                                cell.value = value
                        
                        # Apply styling using StyleRegistry if available
                        col_id = self.idx_to_id_map.get(col_idx)
                        if not col_id:
                            logger.error(f"❌ CRITICAL: Column index {col_idx} has NO column ID mapping!")
                            logger.error(f"   Available mappings: {self.col_id_map}")
                            logger.error(f"   This cell will have NO styling applied!")
                            continue
                        
                        if not self.style_registry:
                            logger.error(f"❌ CRITICAL: StyleRegistry not initialized! Cannot apply styling to cell {cell.coordinate}")
                            logger.error(f"   → Ensure config uses bundled format with 'columns' and 'row_contexts'")
                            continue
                        
                        # Check if column is defined
                        if not self.style_registry.has_column(col_id):
                            logger.warning(f"❌ Column '{col_id}' not found in StyleRegistry! Available: {list(self.style_registry.columns.keys())}")
                            logger.warning(f"   Add to config: styling_bundle.{self.worksheet.title}.columns.{col_id}")
                        
                        # Use 'data' context for regular data rows
                        style = self.style_registry.get_style(col_id, context='data')
                        
                        # For col_static column, apply side borders only (no top/bottom)
                        if col_id == 'col_static':
                            from copy import deepcopy
                            style = deepcopy(style)
                            # Apply side borders only
                            style['border_style'] = 'sides_only'
                        
                        self.cell_styler.apply(cell, style)
                        
                        # Apply row height ONCE per row (only on first column processed)
                        if current_row_idx not in self._rows_with_height_applied:
                            row_height = self.style_registry.get_row_height('data')
                            if row_height:
                                self.cell_styler.apply_row_height(self.worksheet, current_row_idx, row_height)
                                logger.debug(f"Applied row height {row_height} to row {current_row_idx}")
                            self._rows_with_height_applied.add(current_row_idx)
                
                # Handle columns defined in header but missing from row_data (auto-number columns)
                all_column_indices = set(self.col_id_map.values())
                missing_columns = all_column_indices - columns_with_data
                
                for col_idx in missing_columns:
                    col_id = self.idx_to_id_map.get(col_idx)
                    if col_id and 'no' in col_id.lower():  # Auto-number columns like 'col_no'
                        cell = self.worksheet.cell(row=current_row_idx, column=col_idx)
                        if not isinstance(cell, MergedCell):
                            # Auto-number: row number starting from 1
                            cell.value = i + 1
                            
                            # Apply styling
                            if not self.style_registry:
                                logger.error(f"❌ CRITICAL: StyleRegistry not initialized for auto-number column {col_id}")
                                continue
                            
                            style = self.style_registry.get_style(col_id, context='data')
                            self.cell_styler.apply(cell, style)

            # --- Apply Horizontal Merges (based on colspan from header structure) ---
            if self.column_colspan:
                for row_idx in range(data_start_row, data_end_row + 1):
                    for col_id, colspan in self.column_colspan.items():
                        if colspan > 1:  # Only merge if colspan > 1
                            col_idx = self.col_id_map.get(col_id)
                            if col_idx:
                                # Merge from col_idx to col_idx + colspan - 1
                                start_col = col_idx
                                end_col = col_idx + colspan - 1
                                self.worksheet.merge_cells(
                                    start_row=row_idx,
                                    start_column=start_col,
                                    end_row=row_idx,
                                    end_column=end_col
                                )
                                logger.debug(f"Merged data row {row_idx}, columns {start_col}-{end_col} for {col_id} (colspan={colspan})")

            # --- Apply Vertical Merges ---
            if self.vertical_merge_columns and actual_rows_to_process > 0:
                logger.debug(f"Applying vertical merges to columns: {self.vertical_merge_columns}")
                for col_id in self.vertical_merge_columns:
                    col_idx = self.col_id_map.get(col_id)
                    if col_idx:
                        logger.debug(f"  Merging contiguous cells in column '{col_id}' (index {col_idx}) from row {data_start_row} to {data_end_row}")
                        merge_vertical_cells_in_range(
                            worksheet=self.worksheet,
                            scan_col=col_idx,
                            start_row=data_start_row,
                            end_row=data_end_row
                        )
                    else:
                        logger.warning(f"warning!!  Vertical merge requested for column '{col_id}' but column not found in column_id_map")

        except Exception as fill_data_err:
            logger.error(f"Error during data filling loop: {fill_data_err}\n{traceback.format_exc()}")
            return False

        # Log completion summary
        logger.info(f"DataTableBuilder completed: {actual_rows_to_process} data rows written (rows {data_start_row}-{data_end_row})")

        return True
    
    def _build_formula_string(self, formula_dict: Dict[str, Any], row_num: int) -> str:
        """
        Convert a formula dict to an Excel formula string.
        
        Args:
            formula_dict: Dict with 'template' and 'inputs' keys
            row_num: Current row number
        
        Returns:
            Excel formula string (e.g., "=B5*C5")
        """
        template = formula_dict.get('template', '')
        inputs = formula_dict.get('inputs', [])
        
        # Replace placeholders like {col_ref_0}, {col_ref_1}, etc.
        formula = template
        for i, input_id in enumerate(inputs):
            col_idx = self.col_id_map.get(input_id)
            if col_idx:
                col_letter = get_column_letter(col_idx)
                formula = formula.replace(f'{{col_ref_{i}}}', col_letter)
        
        # Replace {row} with actual row number
        formula = formula.replace('{row}', str(row_num))
        
        # Ensure formula starts with =
        if not formula.startswith('='):
            formula = '=' + formula
        
        return formula

