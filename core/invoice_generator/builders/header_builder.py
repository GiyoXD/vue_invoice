import logging
from typing import Any, Dict, List, Optional
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

from ..styling.models import StylingConfigModel
from ..styling.style_applier import apply_header_style  # apply_cell_style removed - using StyleRegistry only
from ..styling.style_registry import StyleRegistry
from ..styling.cell_styler import CellStyler
from ..utils.layout import calculate_header_dimensions
from openpyxl.utils import get_column_letter

class HeaderBuilderStyler:
    def __init__(
        self,
        worksheet: Worksheet,
        start_row: int,
        bundled_columns: List[Dict[str, Any]],
        sheet_styling_config: Optional[StylingConfigModel] = None,
    ):
        """
        Initialize HeaderBuilder with bundled config.
        
        Args:
            worksheet: The worksheet to write to
            start_row: Starting row for header
            bundled_columns: Bundled format (list with id/header/format/rowspan/colspan/children)
            sheet_styling_config: Styling configuration
        """
        self.worksheet = worksheet
        self.start_row = start_row
        self.sheet_styling_config = sheet_styling_config
        self.bundled_columns_original = bundled_columns  # Store for later reference
        
        # Initialize StyleRegistry and CellStyler for ID-driven styling
        self.style_registry = None
        self.cell_styler = CellStyler()
        
        if sheet_styling_config:
            try:
                # Try to create registry from styling_config (if it has columns/row_contexts)
                styling_dict = sheet_styling_config.model_dump() if hasattr(sheet_styling_config, 'model_dump') else sheet_styling_config
                
                if isinstance(styling_dict, dict) and 'columns' in styling_dict and 'row_contexts' in styling_dict:
                    self.style_registry = StyleRegistry(styling_dict)
                    logger.info("StyleRegistry initialized successfully for HeaderBuilder")
                else:
                    logger.error(f"HeaderBuilder: Invalid styling config format. Expected 'columns' and 'row_contexts'.")
                    raise ValueError("Invalid styling config format")
            except Exception as e:
                logger.error(f"Could not initialize StyleRegistry: {e}")
                raise
        else:
             logger.error("HeaderBuilder: No styling config provided!")
             raise ValueError("No styling config provided")
        
        # Convert bundled columns to internal format
        if bundled_columns:
            logger.info(f"Using BUNDLED config (columns={len(bundled_columns)})")
            self.header_layout_config = self._convert_bundled_columns(bundled_columns)
            logger.debug(f"Converted to {len(self.header_layout_config)} header cells")
        else:
            logger.error("HeaderBuilder: No bundled columns provided!")
            raise ValueError("No bundled columns provided")
        
        # Track rows that have had height applied to avoid redundant operations
        self._rows_with_height_applied = set()

    def build(self) -> Optional[Dict[str, Any]]:
        if not self.header_layout_config or self.start_row <= 0:
            return None

        num_header_rows, num_header_cols = calculate_header_dimensions(self.header_layout_config)

        first_row_index = self.start_row
        last_row_index = self.start_row
        max_col = 0
        column_map = {}
        column_id_map = {}
        column_colspan = {}  # Track colspan for each column ID (excluding parents with children)
        
        # Identify parent columns (those with children) - they should NOT be in column_colspan
        parent_column_ids = set()
        if self.bundled_columns_original:
            for col in self.bundled_columns_original:
                if 'children' in col and col['children']:
                    parent_column_ids.add(col.get('id'))

        for cell_config in self.header_layout_config:
            row_offset = cell_config.get('row', 0)
            col_offset = cell_config.get('col', 0)
            text = cell_config.get('text', '')
            cell_id = cell_config.get('id')
            rowspan = cell_config.get('rowspan', 1)
            colspan = cell_config.get('colspan', 1)

            cell_row = self.start_row + row_offset
            cell_col = 1 + col_offset

            last_row_index = max(last_row_index, cell_row + rowspan - 1)
            max_col = max(max_col, cell_col + colspan - 1)

            # Get cell (don't write value yet if it's going to be merged)
            cell = self.worksheet.cell(row=cell_row, column=cell_col)
            
            # Only write value if cell is not already a MergedCell
            from openpyxl.cell.cell import MergedCell
            if not isinstance(cell, MergedCell):
                cell.value = text
            else:
                logger.debug(f"Skipping value write to {cell.coordinate} - already a MergedCell")
            
            # Use StyleRegistry (strict - no legacy fallback)
            if not self.style_registry or not cell_id:
                logger.error(f"❌ CRITICAL: StyleRegistry not initialized or no cell_id for header cell {cell.coordinate}")
                logger.error(f"   → Ensure config uses bundled format with 'columns' and 'row_contexts'")
                continue
            
            # Check if column is defined
            if not self.style_registry.has_column(cell_id):
                logger.warning(f"❌ Column '{cell_id}' not found in StyleRegistry! Available columns: {list(self.style_registry.columns.keys())}")
                logger.warning(f"   Add to config: styling_bundle.{self.worksheet.title}.columns.{cell_id}")
            
            # Get column-specific header style (column base + header context)
            style = self.style_registry.get_style(cell_id, context='header')
            self.cell_styler.apply(cell, style)
            logger.debug(f"Applied StyleRegistry style to header cell {cell_id}")
            
            # Apply row height ONCE per row (only on first column processed for each row)
            if cell_row not in self._rows_with_height_applied:
                row_height = self.style_registry.get_row_height('header')
                if row_height:
                    self.cell_styler.apply_row_height(self.worksheet, cell_row, row_height)
                    logger.debug(f"Applied header row height {row_height} to row {cell_row}")
                self._rows_with_height_applied.add(cell_row)

            if cell_id:
                column_map[text] = get_column_letter(cell_col)
                column_id_map[cell_id] = cell_col
                # Only store colspan for NON-PARENT columns (parents with children shouldn't merge data/footer)
                if cell_id not in parent_column_ids:
                    column_colspan[cell_id] = colspan

            if rowspan > 1 or colspan > 1:
                self.worksheet.merge_cells(start_row=cell_row, start_column=cell_col,
                                      end_row=cell_row + rowspan - 1, end_column=cell_col + colspan - 1)

        return {
            'first_row_index': first_row_index,
            'second_row_index': last_row_index,
            'column_map': column_map,
            'column_id_map': column_id_map,
            'num_columns': max_col,
            'column_colspan': column_colspan  # Add colspan info for automatic merging
        }
    
    def _convert_bundled_columns(self, columns: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Convert bundled columns format to internal header_layout_config format.
        
        Bundled format:
            {"id": "col_po", "header": "P.O. №", "format": "@", "rowspan": 2}
        
        Internal format:
            {"row": 0, "col": 1, "text": "P.O. №", "id": "col_po", "rowspan": 2, "colspan": 1}
        """
        headers = []
        col_index = 0
        
        for col in columns:
            col_id = col.get('id', '')
            header_text = col.get('header', '')
            rowspan = col.get('rowspan', 1)
            colspan = col.get('colspan', 1)
            
            # Handle parent column with children (e.g., Quantity with PCS/SF)
            if 'children' in col:
                # Add parent header
                headers.append({
                    'row': 0,
                    'col': col_index,
                    'text': header_text,
                    'id': col_id,
                    'rowspan': 1,
                    'colspan': len(col['children'])
                })
                
                # Add children headers
                for child in col['children']:
                    headers.append({
                        'row': 1,
                        'col': col_index,
                        'text': child.get('header', ''),
                        'id': child.get('id', ''),
                        'rowspan': 1,
                        'colspan': 1
                    })
                    col_index += 1
            else:
                headers.append({
                    'row': 0,
                    'col': col_index,
                    'text': header_text,
                    'id': col_id,
                    'rowspan': rowspan,
                    'colspan': colspan
                })
                # Increment by colspan to skip physical columns occupied by merge
                col_index += colspan
        
        return headers
