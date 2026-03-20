import logging
import openpyxl
from datetime import datetime
from typing import Dict, Any

logger = logging.getLogger(__name__)

class DeepSheetBuilder:
    """
    Builder responsible for creating and populating the hidden 'DeepSheet' 
    which stores metadata for stable referencing in Excel formulas.
    """

    @staticmethod
    def build(workbook: openpyxl.Workbook, invoice_data: Dict[str, Any]):
        """
        Injects a hidden 'DeepSheet' containing metadata.
        
        Row-based layout for easy VLOOKUP:
        - Column A: label (invoice_no, ref_no, date, net, gross)
        - Column B: value
        
        Usage: =VLOOKUP("net", DeepSheet!A:B, 2, FALSE)
        """
        SHEET_NAME = "DeepSheet"
        
        # Create or get sheet - ENSURE CLEAN OVERRIDE
        if SHEET_NAME in workbook.sheetnames:
            del workbook[SHEET_NAME]
        ws = workbook.create_sheet(SHEET_NAME)
        
        # Set to very hidden
        ws.sheet_state = 'veryHidden'
        
        # --- Extract values from invoice_data ---
        inv_no = ""
        ref_no = ""
        inv_date = ""
        
        # source 1: invoice_info (Created by some parsers or UI overrides)
        if 'invoice_info' in invoice_data:
            info = invoice_data['invoice_info']
            inv_no = info.get('col_inv_no', "") or info.get('inv_no', "")
            ref_no = info.get('col_inv_ref', "") or info.get('inv_ref', "")
            inv_date = info.get('col_inv_date', "") or info.get('inv_date', "")
        
        # source 2: processed_tables_multi['1'] (Standard parser output)
        if not (inv_no and ref_no and inv_date):
            tables = invoice_data.get('processed_tables_multi', {})
            table_1 = tables.get('1', {})
            
            def get_first_val(key):
                """Helper to get first non-empty value from a column list."""
                vals = table_1.get(key, [])
                if isinstance(vals, list):
                    for v in vals:
                        if v: return str(v)
                return ""

            if not inv_no: inv_no = get_first_val('col_inv_no')
            if not inv_date: inv_date = get_first_val('col_inv_date')
            if not ref_no: ref_no = get_first_val('col_inv_ref')

        # --- NEW: Ensure inv_date is a native Date/Datetime object for Excel ---
        if inv_date and isinstance(inv_date, str):
            try:
                # Handle ISO format (e.g. "2026-03-20" or "2026-03-20T00:00:00")
                if 'T' in inv_date:
                    inv_date = datetime.fromisoformat(inv_date)
                else:
                    # Try YYYY-MM-DD
                    inv_date = datetime.strptime(inv_date, "%Y-%m-%d")
            except Exception as e:
                logger.warning(f"Failed to parse date string '{inv_date}': {e}")
                # Fallback: leave as string if parsing fails

        # source 3: footer_data.grand_total for net/gross
        net_val = ""
        gross_val = ""
        footer_data = invoice_data.get('footer_data', {})
        grand_total = footer_data.get('grand_total', {})
        if grand_total:
            net_val = grand_total.get('col_net', "")
            gross_val = grand_total.get('col_gross', "")
            try:
                if net_val: net_val = float(net_val)
            except (ValueError, TypeError):
                net_val = ""
            try:
                if gross_val: gross_val = float(gross_val)
            except (ValueError, TypeError):
                gross_val = ""

        # --- Write row-based layout (Col A = label, Col B = value) ---
        rows = [
            ("ref_no",     ref_no),
            ("invoice_no", inv_no),
            ("date",       inv_date),
            ("net",        net_val),
            ("gross",      gross_val),
        ]
        for row_idx, (label, value) in enumerate(rows, 1):
            ws.cell(row=row_idx, column=1, value=label)
            cell = ws.cell(row=row_idx, column=2, value=value)
            if label == "date" and isinstance(value, datetime):
                cell.number_format = 'yyyy-mm-dd'
        
        logger.info(f"Injected veryHidden '{SHEET_NAME}' with metadata: Inv={inv_no}, Ref={ref_no}, Date={inv_date}, Net={net_val}, Gross={gross_val}")
