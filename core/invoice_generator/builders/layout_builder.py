import logging
from typing import Any, Dict, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook

from ..styling.models import StylingConfigModel, FooterData
from ..data.table_calculator import TableCalculator
from .header_builder import HeaderBuilderStyler as HeaderBuilder
from .data_table_builder import DataTableBuilderStyler as DataTableBuilder
from .footer_builder import TableFooterBuilder
from .json_template_builder import JsonTemplateStateBuilder
from openpyxl.drawing.image import Image
from ...system_config import sys_config

# Initialize logger for this module
logger = logging.getLogger(__name__)

class LayoutBuilder:
    """
    The Director in the Builder pattern.
    Coordinates all builders to construct the complete document layout.
    
    RECOMMENDED USAGE (Modern Bundled Config Approach):
        Use BuilderConfigResolver to prepare configuration bundles, then pass them
        via style_config, context_config, and layout_config parameters. This approach
        centralizes config resolution logic and eliminates duplication.
        
        Example:
            from invoice_generator.config.builder_config_resolver import BuilderConfigResolver
            
            resolver = BuilderConfigResolver(
                config_loader=config_loader,
                sheet_name='Invoice',
                worksheet=worksheet,
                args=args,
                invoice_data=invoice_data,
                pallets=31
            )
            
            # Get bundles - resolver handles all data extraction
            style_config, context_config, layout_config, data_config = resolver.get_datatable_bundles()
            
            layout_builder = LayoutBuilder(
                workbook=workbook,
                worksheet=worksheet,
                template_worksheet=template,
                style_config=style_config,
                context_config=context_config,
        """
    def __init__(
        self,
        workbook: Workbook,
        worksheet: Worksheet,
        template_worksheet: Worksheet,
        style_config: Dict[str, Any],
        context_config: Dict[str, Any],
        layout_config: Dict[str, Any],
        template_state_builder: Optional[JsonTemplateStateBuilder] = None,
        template_json_config: Optional[Dict[str, Any]] = None
    ):
        """
        Initialize LayoutBuilder with strict bundle architecture.
        
        Args:
            workbook: Output workbook (writable)
            worksheet: Output worksheet (writable)
            template_worksheet: Template worksheet (read-only)
            style_config: Bundle containing styling configuration
            context_config: Bundle containing context (sheet_name, invoice_data, args, etc.)
            layout_config: Bundle containing layout rules, structure, and resolved data
            template_state_builder: Optional pre-captured template state (optimization)
        """
        self.workbook = workbook
        self.worksheet = worksheet
        self.template_worksheet = template_worksheet
        
        # Unpack Style Bundle
        self.styling_config = style_config.get('styling_config')
        
        # Unpack Context Bundle
        self.sheet_name = context_config.get('sheet_name')
        self.invoice_data = context_config.get('invoice_data')
        self.all_sheet_configs = context_config.get('all_sheet_configs')
        self.args = context_config.get('args')
        self.final_grand_total_pallets = context_config.get('final_grand_total_pallets', 0)
        self.total_net_weight = context_config.get('total_net_weight')
        self.total_gross_weight = context_config.get('total_gross_weight')
        self.is_last_table = context_config.get('is_last_table', False)
        self.show_grand_total_addons = context_config.get('show_grand_total_addons', False)
        
        # Unpack Layout Bundle
        self.sheet_config = layout_config.get('sheet_config', {})
        
        # Skip flags
        self.skip_template_header_restoration = layout_config.get('skip_template_header_restoration', False)
        self.skip_header_builder = layout_config.get('skip_header_builder', False)
        self.skip_data_table_builder = layout_config.get('skip_data_table_builder', False)
        self.style_config = style_config or {}
        self.context_config = context_config or {}
        self.layout_config = layout_config or {}
        self.template_state_builder = template_state_builder
        self.skip_footer_builder = self.layout_config.get('skip_footer_builder', False)
        
        # We need this to apply padding/dimensions after build
        self.header_info = self.layout_config.get('header_info', {})
        self.skip_template_footer_restoration = layout_config.get('skip_template_footer_restoration', False)
        
        # Data Source (Must be provided via resolved_data in layout_config)
        self.provided_resolved_data = layout_config.get('resolved_data')
        self.provided_header_info = layout_config.get('header_info')
        self.provided_mapping_rules = layout_config.get('mapping_rules')
        
        # Pre-captured template state
        self.pre_captured_template_state = template_state_builder
        self.template_json_config = template_json_config
        
        logger.debug(f"LayoutBuilder initialized for '{self.sheet_name}' with pure bundle config")
        
        # Store results after build
        self.header_info = None
        self.next_row_after_footer = -1
        self.data_start_row = -1
        self.data_end_row = -1
        self.template_state_builder: Optional[JsonTemplateStateBuilder] = None
        self.footer_data: Optional[FooterData] = None
        self.leather_summary: Optional[Dict[str, Any]] = None

    def build(self) -> bool:
        """
        Orchestrates all builders in the correct sequence.
        Reads template state from template_worksheet, writes to self.worksheet (output).
        This completely avoids merge conflicts since template and output are separate.
        """
        logger.info(f"Building layout for sheet '{self.sheet_name}'")
        logger.debug(f"Reading from template, writing to output worksheet")
        
        # 1. Text Replacement (if enabled) - Pre-processing
        # Removed per user request
        
        # 2. Calculate header boundaries for template state capture
        header_row = self.sheet_config.get('header_row', 1)

        # header_to_write removed - using bundled columns only
        num_header_cols = 0
        
        # IMPORTANT: Clarify terminology - there are TWO types of headers:
        # 1. TEMPLATE HEADER: Decorative header section (company name, logo, etc.) - rows 1 to (table_header_row - 1)
        # 2. TABLE HEADER: Column headers for data table (e.g., "Item", "Quantity", "Price") - at table_header_row
        
        # Get table_header_row from config (where the data table column headers are)
        # For multi-table sheets, multi_table_processor dynamically injects the correct
        # expected header_row into self.sheet_config['structure']['header_row'].
        # We MUST respect this injected value over the static global sheet_layout original value.
        sheet_layout = self.all_sheet_configs.get(self.sheet_name, {}) if self.all_sheet_configs else {}
        
        # Priority 1: Injected structure.header_row from multi_table_processor
        if self.sheet_config and 'structure' in self.sheet_config and 'header_row' in self.sheet_config['structure']:
            table_header_row = self.sheet_config['structure']['header_row']
        # Priority 2: Original static template header_row
        else:
            table_header_row = sheet_layout.get('structure', {}).get('header_row', header_row)
            
        header_row_for_builder = table_header_row
        logger.debug(f"[LayoutBuilder DEBUG] sheet_name={self.sheet_name}, header_row={header_row}, table_header_row={table_header_row}")
        logger.debug(f"[LayoutBuilder DEBUG] all_sheet_configs keys: {list(self.all_sheet_configs.keys()) if self.all_sheet_configs else 'None'}")
        
        # Template decorative header spans from row 1 to the row BEFORE the table header
        template_header_start_row = 1
        template_header_end_row = table_header_row - 1  # Decorative header ends BEFORE table header
        
        # Calculate footer_start_row from template (estimate: table_header_row + 2-row table header + minimal data rows)
        # Table header is at table_header_row, second header row at table_header_row + 1
        # Data starts at table_header_row + 2, footer would be around data_start + 2 rows
        # Calculate footer_start_row dynamically from template
        # 3. Template State Capture
        if self.pre_captured_template_state:
            logger.info(f"Using pre-captured template state (multi-table optimization)")
            self.template_state_builder = self.pre_captured_template_state
            logger.debug(f"Reusing template state")
        elif self.template_json_config and self.sheet_name in self.template_json_config:
            # === NEW JSON-BASED PATH ===
            logger.info(f"Using JSON-based template state for sheet '{self.sheet_name}'")
            try:
                sheet_layout_json = self.template_json_config[self.sheet_name]
                self.template_state_builder = JsonTemplateStateBuilder(
                    sheet_layout_data=sheet_layout_json,
                    debug=getattr(self.args, 'debug', False) if self.args else False
                )
                
                # Setup critical boundaries from the loaded builder
                template_header_end_row = self.template_state_builder.header_end_row
                template_footer_start_row = self.template_state_builder.template_footer_start_row
                
                logger.info(f"JSON Template loaded: Header ends {template_header_end_row}, Footer starts {template_footer_start_row}")
                
            except Exception as e:
                logger.critical(f"CRITICAL: JsonTemplateStateBuilder failed for '{self.sheet_name}': {e}", exc_info=True)
                return False
        else:
            # JSON template required - XLSX scanning has been removed
            logger.critical(f"CRITICAL: No JSON template found for sheet '{self.sheet_name}'. XLSX scanning has been removed.")
            return False
            
        # Common: Text replacements removed per user request
        
        # 3b. Template header restoration DEFERRED - will be done AFTER table building
        # This ensures template content aligns with actual column count after filtering
        logger.debug(f"Deferring template header restoration until after table building")
        
        # 4. Header Builder - writes header data to NEW worksheet (unless skipped)
        if not self.skip_header_builder:
            # Convert styling_config dict to StylingConfigModel if needed
            # BUT: If it's already in NEW format (has 'columns' and 'row_contexts'), keep it as-is!
            styling_model = self.styling_config
            if styling_model and not isinstance(styling_model, StylingConfigModel):
                # Check if it's already in NEW format (columns + row_contexts)
                if isinstance(styling_model, dict) and 'columns' in styling_model and 'row_contexts' in styling_model:
                    # NEW format: keep as dict, don't convert to StylingConfigModel
                    logger.debug("Keeping NEW format styling (columns + row_contexts) as dict")
                else:
                    logger.error("LayoutBuilder: Invalid styling config format. Expected 'columns' and 'row_contexts'.")
                    # Don't fallback, let it fail or be None if critical
                    styling_model = None

            # Get bundled columns from sheet_config (bundled config v2.1 format)
            # These are in layout_config -> sheet_config -> 'structure' -> 'columns'
            bundled_columns = None
            column_mapping: Dict[int, Optional[int]] = {}  # For template state column shifting
            
            if self.sheet_config:
                structure = self.sheet_config.get('structure', {})
                original_columns = structure.get('columns', [])
                bundled_columns = original_columns
                
                # Filter columns based on DAF/custom mode flags
                if bundled_columns:
                    DAF_mode = self.args.DAF if self.args and hasattr(self.args, 'DAF') else False
                    custom_mode = self.args.custom if self.args and hasattr(self.args, 'custom') else False
                    
                    # Build column mapping BEFORE filtering
                    # Map each template Excel column position to its output position (or None if removed)
                    if DAF_mode or custom_mode:
                        template_col = 1  # current template col
                        output_col = 1    # current output col
                        
                        for col_def in original_columns:
                            # Use full descriptive names
                            column_id = str(col_def.get('id', ''))
                            skip_daf = bool(col_def.get('skip_in_daf', False))
                            skip_custom = bool(col_def.get('skip_in_custom', False))
                            colspan_val = int(col_def.get('colspan', 1))
                            children_list = col_def.get('children', [])
                            
                            num_columns = len(children_list) if children_list else colspan_val
                            should_skip = (DAF_mode and skip_daf) or (custom_mode and skip_custom)
                            
                            if should_skip:
                                for i in range(num_columns):
                                    column_mapping[template_col + i] = None
                                logger.debug(f"Column '{column_id}' removed: template cols {template_col}-{template_col + num_columns - 1} → None")
                            else:
                                for i in range(num_columns):
                                    column_mapping[template_col + i] = output_col + i
                                logger.debug(f"Column '{column_id}': template cols {template_col}-{template_col + num_columns - 1} → output cols {output_col}-{output_col + num_columns - 1}")
                                output_col += num_columns
                            
                            template_col += num_columns
                        
                        mapping_vals = [v for v in column_mapping.values() if v is not None]
                        logger.info(f"Built column mapping for template shifting: {len(mapping_vals)} active columns")
                    
                    # Now filter the columns list
                    original_count = len(bundled_columns)
                    bundled_columns = [
                        col for col in bundled_columns
                        if not (DAF_mode and col.get('skip_in_daf', False))
                        and not (custom_mode and col.get('skip_in_custom', False))
                    ]
                    if len(bundled_columns) < original_count:
                        logger.info(f"Filtered bundled_columns: {original_count} → {len(bundled_columns)} (DAF={DAF_mode}, custom={custom_mode})")
                
                if not bundled_columns:
                    logger.warning(f"No columns found in sheet_config.structure for sheet '{self.sheet_name}'")

            try:
                logger.debug(f"Creating HeaderBuilder at row {header_row_for_builder}")
                logger.debug(f"Creating HeaderBuilder at row {header_row_for_builder}")
                logger.debug(f"HeaderBuilder input - bundled_columns: {len(bundled_columns) if bundled_columns else 0}")
                header_builder = HeaderBuilder(
                    worksheet=self.worksheet,
                    start_row=header_row_for_builder,  # Use table_header_row (row 21), NOT header_row (row 1)
                    bundled_columns=bundled_columns,  # Bundled format (preferred)
                    sheet_styling_config=styling_model,
                )
                logger.debug(f"Calling HeaderBuilder.build() starting at row {header_row_for_builder}")
                self.header_info = header_builder.build()
                
                if not self.header_info or not self.header_info.get('column_map'):
                    logger.error(f"HeaderBuilder failed for sheet '{self.sheet_name}'")
                    logger.error(f"header_info or column_map is missing - HALTING EXECUTION")
                    logger.error(f"start_row: {header_row_for_builder}, bundled_columns: {len(bundled_columns) if bundled_columns else 0}")
                    return False
                
                header_end_row = self.header_info.get('second_row_index', header_row_for_builder)
                logger.debug(f"HeaderBuilder completed - rows {header_row_for_builder}-{header_end_row}, {len(self.header_info.get('column_map', {}))} columns")
                
                # DEBUG: Check if font is still set after HeaderBuilder
                if self.worksheet:
                    debug_cell = self.worksheet.cell(row=header_row_for_builder, column=1)
                    logger.debug(f"POST-HeaderBuilder - Cell({header_row_for_builder},1) font: name={debug_cell.font.name}, size={debug_cell.font.size}, bold={debug_cell.font.bold}")
            except Exception as e:
                logger.error(f"HeaderBuilder crashed for sheet '{self.sheet_name}'")
                logger.error(f"Error: {e}", exc_info=True)
                logger.error(f"header_row_for_builder={header_row_for_builder}, header_to_write={header_to_write}, bundled_columns={len(bundled_columns) if bundled_columns else 0}")
                return False
        else:
            logger.info(f"Skipping header builder (skip_header_builder=True)")
            # Check if header_info was pre-provided in layout_config (bundled config pattern)
            if self.sheet_config and 'header_info' in self.sheet_config:
                self.header_info = self.sheet_config['header_info']
                logger.debug(f"Using pre-provided header_info from layout_config")
            else:
                # Must provide dummy header_info for downstream builders
                self.header_info = {'column_map': {}, 'first_row_index': header_row, 'second_row_index': header_row + 1}
            styling_model = self.styling_config

        # 5. Data Table Builder (writes data rows, returns footer position) (unless skipped)
        logger.debug(f"skip_data_table_builder = {self.skip_data_table_builder}")
        if not self.skip_data_table_builder:
            logger.info(f"Entering data table builder block")
            sheet_inner_mapping_rules_dict = self.sheet_config.get('mappings', {})
            add_blank_after_hdr_flag = self.sheet_config.get("add_blank_after_header", False)
            static_content_after_hdr_dict = self.sheet_config.get("static_content_after_header", {})
            add_blank_before_ftr_flag = self.sheet_config.get("add_blank_before_footer", False)
            static_content_before_ftr_dict = self.sheet_config.get("static_content_before_footer", {})
            merge_rules_after_hdr = self.sheet_config.get("merge_rules_after_header", {})
            merge_rules_before_ftr = self.sheet_config.get("merge_rules_before_footer", {})
            merge_rules_footer = self.sheet_config.get("merge_rules_footer", {})
            data_cell_merging_rules = self.sheet_config.get("data_cell_merging_rule", None)
            
            # ========== Data Source Resolution ==========
            
            # Primary path: Use TableDataAdapter-provided resolved_data (modern approach)
            if self.provided_resolved_data:
                logger.info(f"Using resolver-provided resolved_data (modern approach)")
                # DataTableBuilder expects resolved_data directly, not wrapped in dtb_data_config
                dtb_data_config = self.provided_resolved_data
            else:
                # If no resolved_data is provided, we cannot proceed in strict bundle mode
                logger.error(f"LayoutBuilder: No resolved_data provided in layout_config. Strict bundle mode requires TableDataAdapter.")
                logger.error(f"Sheet: {self.sheet_name}")
                return False

            # ========== End Data Source Resolution ==========

            # DataTableBuilder uses the new simplified interface
            try:
                expected_row_start = self.header_info.get('second_row_index', 0) + 1
                logger.debug(f"Creating DataTableBuilder - Expected to start at row {expected_row_start}")
                
                # --- 4. Calculate Data (TableCalculator) ---
                # Extract business logic: Calculate sums, pallets, etc. BEFORE rendering
                logger.info("LayoutBuilder: Calculating table data...")
                table_calculator = TableCalculator(self.header_info)
                footer_data = table_calculator.calculate(dtb_data_config)
                self.footer_data = footer_data
                
                if not footer_data:
                    logger.error("LayoutBuilder: TableCalculator failed to return data.")
                    return False



                # --- 5. Build Data Table (DataTableBuilder) ---
                if not self.skip_data_table_builder:
                    logger.info("LayoutBuilder: Building data table...")
                    
                    # NEW: Determine global uniqueness of descriptions across ALL tables
                    is_global_unique_desc = False
                    if self.invoice_data:
                        all_descriptions = set()
                        
                        # 1. Check multi_table (List of Lists of Dicts)
                        multi_table = self.invoice_data.get('multi_table', [])
                        if isinstance(multi_table, list):
                            for table in multi_table:
                                if isinstance(table, list):
                                    for row in table:
                                        desc = str(row.get('col_desc', "")).strip()
                                        if desc: all_descriptions.add(desc)
                        
                        # 2. Check single_table (Dict of Aggregations)
                        single_table = self.invoice_data.get('single_table', {})
                        if isinstance(single_table, dict):
                            # Check standard aggregation
                            agg = single_table.get('aggregation', [])
                            if isinstance(agg, list):
                                for row in agg:
                                    desc = str(row.get('col_desc', "")).strip()
                                    if desc: all_descriptions.add(desc)
                            
                            # Check custom aggregation
                            agg_cust = single_table.get('aggregation_custom', [])
                            if isinstance(agg_cust, list):
                                for row in agg_cust:
                                    desc = str(row.get('col_desc', "")).strip()
                                    if desc: all_descriptions.add(desc)

                        # Final Check: If only ONE unique non-empty description exists globally
                        if len(all_descriptions) == 1:
                            is_global_unique_desc = True
                            logger.info(f"LayoutBuilder: Globally unique description detected: {list(all_descriptions)[0]}")
                        elif len(all_descriptions) > 1:
                            logger.info(f"LayoutBuilder: Globally mixed descriptions detected ({len(all_descriptions)} types).")
                    
                    merge_cols = ['col_pallet_count']
                    if self.layout_config.get('allow_col_desc_merge', True):
                        merge_cols.append('col_desc')
                        
                    data_builder = DataTableBuilder(
                        worksheet=self.worksheet,
                        header_info=self.header_info,
                        resolved_data=dtb_data_config,
                        sheet_styling_config=styling_model,
                        vertical_merge_columns=merge_cols,
                        is_global_unique_desc=is_global_unique_desc
                    )
                    result = data_builder.build()
                    if not result:
                        logger.error("LayoutBuilder: DataTableBuilder failed.")
                        return False
                else:
                    logger.info("LayoutBuilder: Skipping data table build as requested.")

                # --- 6. Build Footer (TableFooterBuilder) ---
                # MOVED: Footer building is now handled explicitly after the data table block
                # to ensure strict separation of concerns and avoid duplication.
                
                # Extract legacy values for logging/compatibility if needed
                if footer_data:
                    self.data_start_row = footer_data.data_start_row
                    self.data_end_row = footer_data.data_end_row
                    data_start_row = footer_data.data_start_row
                    data_end_row = footer_data.data_end_row
                    footer_row_position = footer_data.footer_row_start_idx
                    local_chunk_pallets = footer_data.total_pallets
                    self.leather_summary = footer_data.leather_summary
                else:
                    data_start_row = 0
                    data_end_row = 0
                    footer_row_position = (header_row_for_builder or 0) + 2
                    local_chunk_pallets = 0.0
                    self.leather_summary = None
                
                rows_written = data_end_row - data_start_row + 1 if data_end_row >= data_start_row else 0
                logger.debug(f"DataTableBuilder completed - rows {data_start_row}-{data_end_row} ({rows_written} rows), footer at row {footer_row_position}")
                
                # 5b. NOW restore template header - AFTER table is built
                # This ensures template content aligns with actual number of columns used
                # CRITICAL: This should only restore decorative header (rows 1 to table_header_row-1)
                # It must NOT overwrite the table header row that HeaderBuilder styled
                if not self.skip_template_header_restoration:
                    logger.info(f"Restoring template header AFTER table build (correct column alignment)")
                    try:
                        # Get actual column count from header_info (this reflects filtered columns)
                        actual_num_cols = self.header_info.get('num_columns', None)
                        table_header_row_num = self.header_info.get('second_row_index', 0)
                        logger.debug(f"Template header will use actual column count: {actual_num_cols}")
                        if self.template_state_builder:
                            logger.debug(f"Template header ends at row {self.template_state_builder.header_end_row}")
                        logger.debug(f"Table header row is at: {table_header_row_num}")
                        logger.debug(f"These should NOT overlap! (template_end < table_header)")
                        # DO NOT apply column mapping to the template header!
                        # The user specifically requested that we do not skip anything
                        # when capturing/restoring the template wrapper.
                        
                        # Resolve generation mode for mode-dependent header values
                        header_mode = "daf" if (self.args and getattr(self.args, 'DAF', False)) else "standard"
                        
                        if self.template_state_builder:
                            self.template_state_builder.restore_header_only(
                                target_worksheet=self.worksheet,
                                actual_num_cols=actual_num_cols,
                                mode=header_mode
                            )
                            logger.info(f"Template header restored successfully with {actual_num_cols} columns (rows 1-{self.template_state_builder.header_end_row})")
                    except Exception as e:
                        logger.error(f"Failed to restore template header after table build")
                        logger.error(f"Error: {e}", exc_info=True)
                        return False
                else:
                    logger.debug(f"Skipping template header restoration (skip_template_header_restoration=True)")
                
            except Exception as e:
                logger.error(f"DataTableBuilder crashed for sheet '{self.sheet_name}'")
                logger.error(f"Error: {e}", exc_info=True)
                logger.error(f"header_info={self.header_info}")
                if dtb_data_config and hasattr(dtb_data_config, 'keys'):
                    logger.error(f"data_config keys: {list(dtb_data_config.keys())}")
                return False
        else:
            logger.info(f"Skipping data table builder (skip_data_table_builder=True)")
            # Provide dummy values for downstream builders
            footer_row_position = header_row + 2  # After header
            data_start_row = 0
            data_end_row = 0
            local_chunk_pallets = 0
            data_source_type = None
        
        # 6. Footer Builder (proper Director pattern - called explicitly by LayoutBuilder) (unless skipped)
        logger.debug(f"Checking TableFooterBuilder - skip_footer_builder={self.skip_footer_builder}")
        if not self.skip_footer_builder:
            # Prepare footer parameters
            # Use local_chunk_pallets from data if available, otherwise use grand total
            # For multi-table sheets, local_chunk_pallets will be specific to this table
            # For single-table sheets, use the final_grand_total_pallets
            if local_chunk_pallets > 0:
                pallet_count = local_chunk_pallets
            else:
                pallet_count = self.final_grand_total_pallets

            # Get footer config and sum ranges
            # Support both bundled config format ('footer') and legacy format ('footer_configurations')
            footer_config = self.sheet_config.get('footer', {})
            # Support both bundled config format ('data_flow.mappings') and legacy format ('mappings')
            data_flow = self.sheet_config.get('data_flow', {})
            sheet_inner_mapping_rules_dict = data_flow.get('mappings', self.sheet_config.get('mappings', {}))
            data_range_to_sum = []
            if data_start_row > 0 and data_end_row >= data_start_row:
                data_range_to_sum = [(data_start_row, data_end_row)]

            # Bundle configs for TableFooterBuilder
            footer_builder_style_config = {
                'styling_config': styling_model
            }
            
            footer_builder_context_config = {
                'header_info': self.header_info,
                'pallet_count': pallet_count,
                'sheet_name': self.sheet_name,
                # Pass through weight totals from processor context (if available)
                'total_net_weight': self.total_net_weight,
                'total_gross_weight': self.total_gross_weight,
                'is_last_table': self.is_last_table,
                'show_grand_total_addons': self.show_grand_total_addons
            }
            
            footer_builder_data_config = {
                'sum_ranges': data_range_to_sum,
                'footer_config': footer_config,
                'mapping_rules': sheet_inner_mapping_rules_dict,
                'DAF_mode': bool(getattr(self.args, 'DAF', False)) if self.args else False,
                'override_total_text': None,
                'leather_summary': self.leather_summary
            }

            logger.debug(f"Creating TableFooterBuilder at row {footer_row_position}")
            logger.debug(f"TableFooterBuilder input - footer_type: {footer_config.get('type', 'regular')}, add_blank_before: {footer_config.get('add_blank_before', False)}, pallet_count: {pallet_count}")
            try:
                # 4. Build Footer
                # Use TableFooterBuilder (builds table data footer - TOTAL: row)
                footer_builder = TableFooterBuilder(
                    worksheet=self.worksheet,
                    footer_data=self.footer_data,
                    style_config=footer_builder_style_config,
                    context_config=footer_builder_context_config,
                    data_config=footer_builder_data_config
                )
                
                logger.debug(f"Calling TableFooterBuilder.build() with footer_row_position={footer_row_position}")
                footer_start = footer_row_position
                self.next_row_after_footer = footer_builder.build()
                
                # Validate footer builder result
                if self.next_row_after_footer is None or self.next_row_after_footer <= 0:
                    logger.error(f"TableFooterBuilder failed for sheet '{self.sheet_name}'")
                    logger.error(f"Invalid next_row_after_footer={self.next_row_after_footer} - HALTING EXECUTION")
                    logger.error(f"footer_row_position={footer_row_position}, sum_ranges={data_range_to_sum}")
                    logger.error(f"footer_config: {footer_config}")
                    return False
                
                footer_rows_written = self.next_row_after_footer - footer_start
                logger.debug(f"TableFooterBuilder completed - rows {footer_start}-{self.next_row_after_footer - 1} ({footer_rows_written} rows), next available: {self.next_row_after_footer}")
            except Exception as e:
                logger.error(f"TableFooterBuilder crashed for sheet '{self.sheet_name}'")
                logger.error(f"Error: {e}", exc_info=True)
                logger.error(f"footer_row_position={footer_row_position}, pallet_count={pallet_count}")
                logger.error(f"footer_config: {footer_config}")
                return False
            
            # Apply footer height to all footer rows (including add-ons like grand total)
            if self.next_row_after_footer > footer_row_position:
                # Multiple footer rows were created (e.g., regular footer + grand total)
                for footer_row in range(footer_row_position, self.next_row_after_footer):
                    self._apply_footer_row_height(footer_row, styling_model)
            else:
                # Single footer row
                self._apply_footer_row_height(footer_row_position, styling_model)
        else:
            logger.info(f"Skipping footer builder (skip_footer_builder=True)")
            # No footer, so next row is right after data (or header if no data)
            self.next_row_after_footer = footer_row_position
        
        # 6b. Auto-fit column widths and row heights based on actual cell content
        from ..utils.layout import auto_fit_dimensions
        _af_header_start = self.header_info.get('second_row_index', 1) + 1
        _af_data_end = self.next_row_after_footer - 1
        _af_num_cols = self.header_info.get('num_columns', 0)
        _af_header_row_start = self.header_info.get('first_row_index', None)
        _af_header_row_end = self.header_info.get('second_row_index', None)
        
        # Calculate template boundaries for last-column scanning
        _af_template_top_end = _af_header_row_start - 1 if _af_header_row_start else None
        _af_template_bottom_start = self.next_row_after_footer
        _af_max_row = self.worksheet.max_row
        
        logger.info(f"auto_fit_dimensions CALL: header_start={_af_header_start}, data_end={_af_data_end}, num_columns={_af_num_cols}, header_rows={_af_header_row_start}-{_af_header_row_end}, template_top_end={_af_template_top_end}, template_bottom_start={_af_template_bottom_start}, max_row={_af_max_row}")
        try:
            auto_fit_dimensions(
                worksheet=self.worksheet,
                header_start_row=_af_header_start,
                data_end_row=_af_data_end,
                num_columns=_af_num_cols,
                padding=7,
                line_height=20.0,
                header_row_start=_af_header_row_start,
                header_row_end=_af_header_row_end,
                template_top_end_row=_af_template_top_end,
                template_bottom_start_row=_af_template_bottom_start,
                max_row=_af_max_row
            )
        except Exception as e:
            logger.error(f"auto_fit_dimensions FAILED: {e}", exc_info=True)

        # 7. Template Footer Restoration
        # This restores the static content (signatures, etc.) from the JSON template
        # that appears AFTER the dynamic table footer.
        skip_template_footer = self.layout_config.get('skip_template_footer_restoration', False)
        
        if self.template_state_builder and not skip_template_footer:
            try:
                # Get actual column count if not already set
                actual_num_cols = self.header_info.get('num_columns', None)
                
                # CRITICAL FIX: Only restore template footer if this is the LAST table on the sheet.
                # Otherwise, the static footer content (signatures, etc.) will be printed in the middle
                # of the sheet, distorting subsequent tables.
                if self.is_last_table:
                    logger.info(f"--- RESTORING TEMPLATE FOOTER (Last Table) ---")
                    logger.info(f"next_row_after_footer: {self.next_row_after_footer}")
                    
                    self.template_state_builder.restore_template_footer(
                        target_worksheet=self.worksheet,
                        footer_start_row=self.next_row_after_footer,
                        actual_num_cols=actual_num_cols
                    )
                else:
                    logger.info(f"Skipping template footer restoration (Not last table)")
                logger.info(f"Template footer restored successfully")
            except Exception as e:
                logger.error(f"Failed to restore template footer: {e}", exc_info=True)
        else:
            logger.debug("Skipping template footer restoration (no template_state_builder)")

        # 8. Inject Template Images (New Feature)
        self._inject_images()
        
        logger.info(f"Layout built successfully for sheet '{self.sheet_name}'")
        
        return True

    def _inject_images(self):
        """
        Injects images from the configured directory into the worksheet.
        """
        try:
            img_dir = sys_config.template_image_dir
            if not img_dir.exists():
                logger.debug(f"Template image directory not found: {img_dir}")
                return

            images = list(img_dir.glob("*"))
            if not images:
                logger.debug(f"No images found in {img_dir}")
                return

            logger.info(f"Injecting {len(images)} images from {img_dir}")
            
            for i, img_path in enumerate(images):
                if img_path.suffix.lower() not in ['.png', '.jpg', '.jpeg', '.bmp', '.gif']:
                    continue
                    
                try:
                    img = Image(str(img_path))
                    
                    # Resize to 70px height (maintaining aspect ratio)
                    # This only affects display size; original image data is preserved
                    target_height = 140
                    if img.height > 0:
                        aspect_ratio = img.width / img.height
                        new_width = target_height * aspect_ratio
                        
                        img.height = target_height
                        img.width = new_width
                        
                    # Default placement at N1 (as requested)
                    self.worksheet.add_image(img, 'N1')
                    logger.debug(f"Injected image: {img_path.name} (resized to 70px height) at N1")
                except Exception as e:
                    logger.warning(f"Failed to inject image {img_path.name}: {e}")
        except Exception as e:
            logger.error(f"Image injection failed: {e}", exc_info=True)
    
    def _apply_footer_row_height(self, footer_row: int, styling_config):
        """Helper method to apply footer height to a single footer row."""
        if not styling_config:
            return
        
        # Handle NEW format (dict with 'row_contexts')
        if isinstance(styling_config, dict):
            # NEW format: row heights are in row_contexts.footer.row_height
            if 'row_contexts' in styling_config:
                footer_context = styling_config['row_contexts'].get('footer', {})
                if 'row_height' in footer_context:
                    # NEW format stores height directly in context
                    height = footer_context['row_height']
                    if height:
                        self.worksheet.row_dimensions[footer_row].height = height
                        logger.debug(f"Applied footer height {height} to row {footer_row} (NEW format)")
                return
        else:
             logger.warning("LayoutBuilder: Legacy styling config format detected (not a dict). Row heights NOT applied.")

    def _detect_template_footer_start(self, worksheet: Worksheet, table_header_row: int) -> int:
        """
        Dynamically detect the start of the footer in the template.
        Scans downwards from the table header to find where the data area likely ends
        and the footer (static content) begins.
        
        Args:
            worksheet: The template worksheet
            table_header_row: The row number of the table header
            
        Returns:
            The detected footer start row number
        """
        # Start scanning after the header. 
        # We assume at least 1 row for the header itself, and maybe 1 row for a sub-header or first data row.
        scan_start_row = table_header_row + 2
        
        # Validation: If header is below the last row of the sheet, something is wrong with the config
        if scan_start_row > worksheet.max_row + 1:
             error_msg = f"Configuration Mismatch: Config expects header at row {table_header_row}, but template only has {worksheet.max_row} rows. Please check your 'header_row' setting in the config."
             logger.error(error_msg)
             raise ValueError(error_msg)

        max_scan_row = min(scan_start_row + 50, worksheet.max_row + 1) # Limit scan to 50 rows
        
        logger.debug(f"Scanning for footer start from row {scan_start_row} to {max_scan_row}")
        
        # Scan for strict footer markers:
        # 1. "total" AND ":" (case-insensitive)
        # 2. Starts with "=SUM" (case-insensitive)
        
        for r_idx in range(scan_start_row, max_scan_row):
            # Already capped at 20, which is safe.
            for c_idx in range(1, min(20, worksheet.max_column + 1)):
                cell = worksheet.cell(row=r_idx, column=c_idx)
                if cell.value:
                    val_str = str(cell.value).strip().lower()
                    
                    # Check 1: "total" AND ":"
                    if "total" in val_str and ":" in val_str:
                         logger.info(f"Found strict footer marker 'total...:' at row {r_idx}. Using this as footer start.")
                         return r_idx
                         
                    if val_str.startswith("=sum"):
                        logger.info(f"Found strict footer marker '=SUM' at row {r_idx}. Using this as footer start.")
                        return r_idx

                    # Check 3: Signature/Bank markers (Fallback if Total is missing/deleted)
                    # User requested deleting entire table including footer, so we must detect what comes AFTER the table.
                    strict_static = ["the buyer", "the seller", "beneficiary", "authorized signature"]
                    for kw in strict_static:
                        if kw in val_str:
                             logger.info(f"Found strict footer marker '{kw}' at row {r_idx}. Using this as footer start.")
                             return r_idx

        # If loop finishes without finding anything:
        error_msg = f"Footer not detected! Could not find 'total...:', '=SUM', or signature keywords (scanned rows {scan_start_row}-{max_scan_row}). Please ensure your template has a Total row, SUM formula, or 'The Buyer/Seller' signature."
        logger.error(error_msg)
        raise ValueError(error_msg)
