# invoice_generator/builders/workbook_builder.py
import logging
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List

logger = logging.getLogger(__name__)

class WorkbookBuilder:
    """
    Builder responsible for creating a new clean workbook with specified sheet names.
    This allows us to start with a completely clean workbook without any template conflicts.
    """
    
    def __init__(self, sheet_names: List[str]):
        """
        Initialize the WorkbookBuilder.
        
        Args:
            sheet_names: List of sheet names to create in the new workbook
        """
        self.sheet_names = sheet_names
        self.workbook = None
    
    def build(self) -> Workbook:
        """
        Creates a new workbook with sheets matching the template sheet names.
        All sheets are empty and ready for restoration/building.
        
        Returns:
            A new Workbook instance with empty sheets
        """
        logger.info(f"Creating new workbook with {len(self.sheet_names)} sheets")
        
        # Create new workbook
        self.workbook = Workbook()
        
        # Remove the default 'Sheet' created by openpyxl
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Create all required sheets with correct names
        for sheet_name in self.sheet_names:
            self.workbook.create_sheet(title=sheet_name)
            logger.debug(f"Created sheet: '{sheet_name}'")
        
        logger.info(f"New workbook created successfully")
        return self.workbook
    
    def get_worksheet(self, sheet_name: str) -> Worksheet:
        """
        Get a specific worksheet from the created workbook.
        
        Args:
            sheet_name: Name of the sheet to retrieve
            
        Returns:
            The requested worksheet
        """
        if self.workbook is None:
            raise RuntimeError("Workbook not created yet. Call build() first.")
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        return self.workbook[sheet_name]
