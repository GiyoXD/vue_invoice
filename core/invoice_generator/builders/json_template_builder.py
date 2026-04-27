import logging
import copy
from typing import List, Dict, Any, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# Utils

logger = logging.getLogger(__name__)

class JsonTemplateStateBuilder:
    """
    JsonTemplateStateBuilder: Reconstructs Excel template state from JSON configuration.

    This builder is responsible for "hydrating" a template state (headers, footers, styles, merges, dimensions)
    directly from a JSON dictionary, bypassing the need to open and scan a physical .xlsx template file.
    It is a drop-in replacement for the scanning logic found in `TemplateStateBuilder`, designed to work
    with the `layout_template.json` structure.

    Key Responsibilities:
    1.  **Parsing**: Converts coordinate-based JSON data (e.g., "A1": {...}) into row-based grid structures.
    2.  **Style Reconstruction**: Re-creates OpenPyXL style objects (Font, Border, Fill, Alignment) from their JSON representations.
    3.  **State Management**: Maintains separate states for the Header (top of sheet) and Footer (bottom of sheet).
    4.  **Restoration**: Provides methods (`restore_header_only`, `restore_template_footer`) to write this state onto a new worksheet.

    Usage:
        layout_data = loaded_json['template_layout']['Invoice']
        builder = JsonTemplateStateBuilder(layout_data)
        
        # Later, apply to a target worksheet:
        builder.restore_header_only(target_ws)
        builder.restore_template_footer(target_ws, footer_start_row=50)
    """
    
    DEBUG = False

    def __init__(self, sheet_layout_data: Dict[str, Any], debug: bool = False):
        """
        Initialize and populate state from JSON data.
        
        Args:
            sheet_layout_data: The dictionary for a specific sheet from the layout_template.json
                               (e.g., loaded_json['template_layout']['Invoice'])
            debug: Enable debug printing
        """
        self.layout_data = sheet_layout_data
        self.debug = debug or self.DEBUG
        
        # DEBUG INPUT
        logger.debug(f"[JsonTemplateStateBuilder] __init__ INPUT: sheet_layout_data keys={list(sheet_layout_data.keys()) if sheet_layout_data else 'None'}")
        
        # State structures (same as TemplateStateBuilder)
        self.header_state: List[List[Dict[str, Any]]] = []
        self.footer_state: List[List[Dict[str, Any]]] = []
        self.header_merged_cells: List[str] = []
        self.footer_merged_cells: List[str] = []
        self.row_heights: Dict[int, float] = {}
        self.column_widths: Dict[int, float] = {}
        
        # Relative Footer State (0-indexed)
        # These store the footer structure decoupled from absolute template coordinates.
        # Row 0 = The first row of the footer block.
        self.relative_footer_row_heights: Dict[int, float] = {}
        self.relative_footer_merges: List[tuple] = [] # List of (min_col, min_row, max_col, max_row)
        
        # Row tracking
        self.template_footer_start_row: int = -1
        self.template_footer_end_row: int = -1
        self.header_end_row: int = -1
        
        # Dimensions
        self.min_row = 1
        self.max_row = 1
        self.min_col = 1
        self.max_col = 1
        
        # Column mapping for shifting
        self.column_mapping: Dict[int, int] = {}

        # Parse the JSON data immediately
        self._parse_layout_data()

    def set_column_mapping(self, mapping: Dict[int, int]):
        """Set the column mapping for restoration (same as TemplateStateBuilder)."""
        self.column_mapping = mapping

    def _get_mapped_column(self, template_col: int) -> int:
        """Get output column index (same as TemplateStateBuilder)."""
        if not self.column_mapping:
            return template_col
        return self.column_mapping.get(template_col, template_col)

    def _parse_layout_data(self):
        """
        Parses the raw JSON layout data into internal state structures.
        
        This method iterates through the JSON dictionary provided to __init__ and:
        1.  Extracts column widths.
        2.  Builds the `header_state` grid and detects header boundaries.
        3.  Builds the `footer_state` grid and detects footer boundaries.
        4.  Identifies merged cells for both header and footer.
        5.  Populates row heights.
        6.  Calculates `template_footer_start_row` to determine where static footer content begins.
            
            WHY THIS IS NEEDED:
            The template's footer is defined at absolute coordinates (e.g., Row 50). When we generate a dynamic
            data table that might end at Row 100, we need to know the original "reference point" (Row 50)
            to calculate the vertical shift needed (Offset = 100 - 50 = +50 rows).
            Without this starting row, we cannot correctly reposition the footer below the variable-length table.
        
        It handles missing keys gracefully and attempts to infer structure where possible (e.g., fallback logic).
        """
        logger.info("[JsonTemplateStateBuilder] Parsing Layout Data")
        
        # 1. Parse Dimensions and Basic Props
        # Note: JSON 'col_widths' keys are letters 'A', 'B'...
        col_widths = self.layout_data.get('col_widths', {})
        for col_letter, width in col_widths.items():
            self.column_widths[column_index_from_string(col_letter)] = width
            
        def _flatten_grouped_styles(style_dict_in: Dict[str, Any]) -> Dict[str, Any]:
            """
            Option B Support: Check if the style map is using grouped IDs (Style -> [Coord1, Coord2]).
            If yes, flatten it back to Coord -> StyleID for the legacy grid builder.
            """
            if not style_dict_in: return {}
            # Quick check if first value is a list (which means it's grouped)
            first_val = next(iter(style_dict_in.values()))
            if isinstance(first_val, list):
                flattened = {}
                for style_id, coord_list in style_dict_in.items():
                    for coord in coord_list:
                        flattened[coord] = style_id
                return flattened
            return style_dict_in

        # 2. Parse Header State
        # 'header_content' is {"A1": {...}, "B1": {...}}
        header_content = self.layout_data.get('header_content', {})
        header_styles_raw = self.layout_data.get('header_styles', {})
        header_styles = _flatten_grouped_styles(header_styles_raw)
        
        self.style_palette = self.layout_data.get('style_palette', {}) # Option B Support
        header_merges_raw = self.layout_data.get('header_merges', [])
        self.header_merged_cells = list(header_merges_raw.keys()) if isinstance(header_merges_raw, dict) else header_merges_raw
        
        self.header_state, self.header_end_row = self._build_state_grid(header_content, header_styles, is_header=True)
        
        # Load header row heights
        header_row_heights = self.layout_data.get('header_row_heights', {})
        for r_str, h in header_row_heights.items():
            self.row_heights[int(r_str)] = h

        # 3. Parse Footer State
        self.footer_rows = self.layout_data.get('footer_rows')
        
        if self.footer_rows is not None:
            # --- NEW GRID-ROW FORMAT ---
            logger.info(f"[JsonTemplateStateBuilder] Using new footer_rows grid format ({len(self.footer_rows)} rows)")
            
            self.template_footer_start_row = (self.header_end_row + 1) if self.header_end_row > 0 else 1
            max_rel_idx = max((r.get('relative_index', 0) for r in self.footer_rows), default=-1)
            self.template_footer_end_row = (self.template_footer_start_row + max_rel_idx) if max_rel_idx >= 0 else -1
            
            # Update max_col based on new footer cells
            for r_dict in self.footer_rows:
                for c_dict in r_dict.get('cells', []):
                    c_idx = c_dict.get('col_index', 1)
                    if not hasattr(self, 'max_col') or c_idx > self.max_col:
                        self.max_col = c_idx
                for m_dict in r_dict.get('merges', []):
                    c_idx = m_dict.get('max_col', 1)
                    if not hasattr(self, 'max_col') or c_idx > self.max_col:
                        self.max_col = c_idx
        else:
            # --- OLD COORDINATE FORMAT (Fallback) ---
            logger.info(f"[JsonTemplateStateBuilder] Using old coordinate-based footer format")
            footer_content = self.layout_data.get('footer_content', {})
            footer_styles_raw = self.layout_data.get('footer_styles', {})
            footer_styles = _flatten_grouped_styles(footer_styles_raw)
            footer_merges_raw = self.layout_data.get('footer_merges', [])
            self.footer_merged_cells = list(footer_merges_raw.keys()) if isinstance(footer_merges_raw, dict) else footer_merges_raw
            
            self.footer_state, self.template_footer_end_row = self._build_state_grid(footer_content, footer_styles, is_header=False)
            
            # Determine footer start row
            footer_row_heights = self.layout_data.get('footer_row_heights', {})
            
            if footer_content or footer_styles or self.footer_merged_cells or footer_row_heights:
                all_keys = list(footer_content.keys()) + list(footer_styles.keys())
                min_r = float('inf')
                
                # Check content and styles coordinates
                for k in all_keys:
                    try:
                        _, r = coordinate_from_string(k)
                        if r < min_r: min_r = r
                    except: pass
                    
                # Check merged cells
                from openpyxl.utils.cell import range_boundaries
                for merge in self.footer_merged_cells:
                    try:
                        _, min_row, _, _ = range_boundaries(merge)
                        if min_row < min_r: min_r = min_row
                    except: pass
                    
                # Prevent overlap
                minimum_safe_footer_row = (self.header_end_row + 1) if self.header_end_row > 0 else 1
                
                if min_r != float('inf'):
                    self.template_footer_start_row = max(min_r, minimum_safe_footer_row)
                    if min_r < minimum_safe_footer_row:
                        logger.warning(
                            f"[JsonTemplateStateBuilder] Detected footer marker at row {min_r}, "
                            f"but this overlaps with header/data area. Forcing footer start to {minimum_safe_footer_row}."
                        )
                else:
                    self.template_footer_start_row = -1
                    
                if self.template_footer_start_row == -1:
                    fallback_row = (self.header_end_row + 1) if self.header_end_row > 0 else -1
                    if fallback_row > 0:
                        self.template_footer_start_row = fallback_row
                    else:
                        logger.error("[JsonTemplateStateBuilder] Failed to resolve old footer logic.")
                        
            # Normalize Footer State to Relative Coordinates
            if self.template_footer_start_row > 0:
                for r_str, h in footer_row_heights.items():
                    try:
                        r = int(r_str)
                        if r >= self.template_footer_start_row:
                            rel_r = r - self.template_footer_start_row
                            self.relative_footer_row_heights[rel_r] = h
                    except ValueError: pass
                    
                from openpyxl.utils.cell import range_boundaries
                for merge in self.footer_merged_cells:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(merge)
                        if min_row >= self.template_footer_start_row:
                            rel_min = min_row - self.template_footer_start_row
                            rel_max = max_row - self.template_footer_start_row
                            self.relative_footer_merges.append((min_col, rel_min, max_col, rel_max))
                    except ValueError: pass
            
        # Update max_col
        if self.column_widths:
            self.max_col = max(self.column_widths.keys())
        
        # CRITICAL FIX: Update max_row to reflect the actual last row in the template
        # This is used by layout_builder.py line 608 to calculate footer row count:
        #   template_footer_rows = self.template_state_builder.max_row - self.template_state_builder.template_footer_start_row + 1
        # Without this, max_row stays at default value of 1, causing wrong footer placement
        if self.template_footer_end_row > 0:
            self.max_row = self.template_footer_end_row
        elif self.header_end_row > 0:
            self.max_row = self.header_end_row

    def _build_state_grid(self, content_map: Dict, style_map: Dict, is_header: bool) -> tuple:
        """
        Converts coordinate-based maps (e.g., "A1": val) into a row-major list-of-lists grid.

        Args:
            content_map: Dictionary mapping coordinates to cell values (e.g., {"A1": "Invoice"}).
            style_map: Dictionary mapping coordinates to style dictionaries.
            is_header: Boolean flag used for debug logging context.

        Returns:
            A tuple (state_grid, max_row_index):
                - state_grid: List[List[Dict]] where each inner Dict represents a cell's properties
                  (value, font, border, fill, alignment, number_format).
                - max_row_index: The highest 1-based row index found in the input maps.
        """
        # DEBUG INPUT
        logger.debug(f"[JsonTemplateStateBuilder] _build_state_grid INPUT: is_header={is_header}, content_keys={len(content_map) if content_map else 0}")

        if not content_map and not style_map:
            return [], 0

        # Find bounds
        all_coords = set(content_map.keys()) | set(style_map.keys())
        if not all_coords:
            return [], 0
            
        rows = set()
        cols = set()
        for coord in all_coords:
            try:
                c, r = coordinate_from_string(coord)
                rows.add(r)
                cols.add(column_index_from_string(c))
            except:
                continue
                
        if not rows: return [], 0
        
        min_r = min(rows)
        max_r = max(rows)
        max_c = max(cols) if cols else 1
        
        # Ensure we cover at least the columns defined in widths or arbitrary max
        final_max_c = max(max_c, max(self.column_widths.keys()) if self.column_widths else 0)
        
        grid = []
        
        # Iterate row by row
        for r in range(min_r, max_r + 1):
            row_data = []
            for c in range(1, final_max_c + 1):
                col_letter = get_column_letter(c)
                coord = f"{col_letter}{r}"
                
                # Extract value
                raw_val = content_map.get(coord, None)
                
                # Extract style dict
                style_entry = style_map.get(coord, {})
                
                # Option B Support: If the style is a string (hash ID), locate it in the palette 
                if isinstance(style_entry, str):
                    style_dict = self.style_palette.get(style_entry, {})
                else:
                    style_dict = style_entry
                
                # Convert style dict to OpenPyXL objects
                cell_info = {
                    'value': raw_val,
                    'font': self._create_font(style_dict.get('font')),
                    'fill': self._create_fill(style_dict.get('fill')),
                    'border': self._create_border(style_dict.get('border')),
                    'alignment': self._create_alignment(style_dict.get('alignment')),
                    'number_format': style_dict.get('number_format', 'General')
                }
                row_data.append(cell_info)
            grid.append(row_data)
            
            
        # DEBUG OUTPUT
        logger.debug(f"[JsonTemplateStateBuilder] _build_state_grid OUTPUT: grid_rows={len(grid)}, max_r={max_r}")
        return grid, max_r

    # --- Style Factory Methods ---
    def _create_font(self, d: Dict) -> Optional[Font]:
        if not d: return None
        # Handle color dict/str
        color = d.get('color')
        # If color is dict (RGB), extract rgb
        if isinstance(color, dict) and 'rgb' in color:
             color = color['rgb']
        elif isinstance(color, dict) and 'theme' in color:
             # Simplify theme colors to None or black for now unless we look up theme
             color = None
             
        return Font(
            name=d.get('name'),
            size=d.get('size'),
            bold=d.get('bold'),
            italic=d.get('italic'),
            strike=d.get('strike'),
            underline=d.get('underline'),
            color=color
        )
        
    def _create_fill(self, d: Dict) -> Optional[PatternFill]:
        if not d: return None
        if not d.get('type'): return None
        # Simplification: mostly dealing with solid fills usually
        # The serializer saves 'color' as '00000000' usually for transparent
        # We need check how sanitizer saves it. 
        # For now, instantiate basic PatternFill
        fgColor = d.get('color')
        if fgColor == '00000000': fgColor = None # Transparent
        
        return PatternFill(
            fill_type=d.get('type'),
            start_color=fgColor,
            end_color=fgColor # Simple solid fill assumption
        )
        
    def _create_border(self, d: Dict) -> Optional[Border]:
        if not d: return None
        def _side(s_data):
            if not s_data: return None
            # s_data might be simple style string or dict? 
            # Review sanitizer: "left": cell.border.left.style
            # It saves just the style string (e.g. 'thin', 'medium')
            return Side(style=s_data) if s_data else None

        return Border(
            left=_side(d.get('left')),
            right=_side(d.get('right')),
            top=_side(d.get('top')),
            bottom=_side(d.get('bottom'))
        )

    def _create_alignment(self, d: Dict) -> Optional[Alignment]:
        if not d: return None
        return Alignment(
            horizontal=d.get('horizontal'),
            vertical=d.get('vertical'),
            text_rotation=d.get('text_rotation', 0),
            wrap_text=d.get('wrap_text'),
            shrink_to_fit=d.get('shrink_to_fit'),
            indent=d.get('indent', 0)
        )

    # --- Restoration Logic (Mirrors TemplateStateBuilder) ---
    # We copy this verbatim from TemplateStateBuilder to allow safe refactor later.
    
    def restore_header_only(self, target_worksheet: Worksheet, actual_num_cols: int = None, mode: str = "standard"):
        """
        Restores ONLY the header to a new clean worksheet.
        
        Args:
            target_worksheet: The worksheet to write header content onto.
            actual_num_cols: Optional column count to limit restoration.
            mode: Generation mode ('standard', 'daf', 'custom'). Used to resolve
                  mode-dependent cell values in header_content.
        """
        logger.info(f"[JsonTemplateStateBuilder] Restoring Header to '{target_worksheet.title}' (mode={mode})")
        logger.debug(f"[JsonTemplateStateBuilder] restore_header_only INPUT: target_worksheet={target_worksheet.title}, actual_num_cols={actual_num_cols}, mode={mode}")

        template_num_cols = self.max_col
        target_num_cols = actual_num_cols if actual_num_cols else template_num_cols
        
        # Restore header cell values and formatting
        for row_idx, row_data in enumerate(self.header_state):
            # For header, we start at min_row (usually 1)
            actual_row = row_idx + self.min_row
            
            for col_idx, cell_info in enumerate(row_data):
                template_col = col_idx + self.min_col
                output_col = self._get_mapped_column(template_col)
                
                if output_col is None:
                    continue # Skip removed columns (simple version of logic)
                
                target_cell = target_worksheet.cell(row=actual_row, column=output_col)
                self._write_cell(target_cell, cell_info, mode=mode)
                
        # Restore header merges
        for merge_str in self.header_merged_cells:
            self._apply_merge(target_worksheet, merge_str)
            
        # Restore dimensions
        for r_idx in range(self.min_row, self.header_end_row + 1):
             if r_idx in self.row_heights:
                 target_worksheet.row_dimensions[r_idx].height = self.row_heights[r_idx]
                 
        for c_idx, w in self.column_widths.items():
            target_worksheet.column_dimensions[get_column_letter(c_idx)].width = w

    def restore_template_footer(self, target_worksheet: Worksheet, footer_start_row: int, actual_num_cols: int = None, mode: str = "standard"):
        """
        Restores the template footer content onto the target worksheet at a specific starting row.
        """
        logger.info(f"[JsonTemplateStateBuilder] Restoring Footer to '{target_worksheet.title}' at row {footer_start_row} (mode={mode})")

        # --- NEW GRID-ROW FORMAT ---
        if hasattr(self, 'footer_rows') and self.footer_rows is not None:
            if not self.footer_rows:
                logger.warning(f"[JsonTemplateStateBuilder] Footer rows is empty for '{target_worksheet.title}'.")
                return
                
            for row_dict in self.footer_rows:
                rel_idx = row_dict.get('relative_index', 0)
                actual_row = footer_start_row + rel_idx
                
                # 1. Restore Row Height
                h = row_dict.get('height')
                if h is not None:
                    target_worksheet.row_dimensions[actual_row].height = h
                    
                # 2. Restore Cells (Values & Styles)
                for cell_dict in row_dict.get('cells', []):
                    template_col = cell_dict.get('col_index')
                    output_col = self._get_mapped_column(template_col)
                    
                    if output_col is None: continue
                    
                    target_cell = target_worksheet.cell(row=actual_row, column=output_col)
                    
                    val = cell_dict.get('value')
                    if val is not None:
                        resolved = self._resolve_mode_value(val, mode)
                        if resolved is not None:
                            target_cell.value = resolved
                        
                    style_id = cell_dict.get('style_id')
                    if style_id:
                        style_dict = self.style_palette.get(style_id, {})
                        
                        font = self._create_font(style_dict.get('font'))
                        fill = self._create_fill(style_dict.get('fill'))
                        border = self._create_border(style_dict.get('border'))
                        align = self._create_alignment(style_dict.get('alignment'))
                        num_fmt = style_dict.get('number_format', 'General')
                        
                        if font: target_cell.font = copy.copy(font)
                        if fill: target_cell.fill = copy.copy(fill)
                        if border: target_cell.border = copy.copy(border)
                        if align: target_cell.alignment = copy.copy(align)
                        if num_fmt: target_cell.number_format = num_fmt
                        
                # 3. Restore Merges
                for m_dict in row_dict.get('merges', []):
                    min_col = m_dict.get('min_col')
                    max_col = m_dict.get('max_col')
                    row_span = m_dict.get('row_span', 1)
                    
                    mapped_min_col = self._get_mapped_column(min_col)
                    mapped_max_col = self._get_mapped_column(max_col)
                    
                    if mapped_min_col and mapped_max_col:
                        new_range = f"{get_column_letter(mapped_min_col)}{actual_row}:{get_column_letter(mapped_max_col)}{actual_row + row_span - 1}"
                        try:
                            target_worksheet.merge_cells(new_range)
                        except ValueError:
                            pass
            return
            
        # --- OLD COORDINATE FORMAT (Fallback) ---

        # GUARD: Refuse to restore if footer parsing failed.
        if self.template_footer_start_row <= 0:
            logger.error(
                f"[JsonTemplateStateBuilder] Cannot restore footer: template_footer_start_row "
                f"is {self.template_footer_start_row}. Footer parsing likely failed or no footer data found."
            )
            return

        try:
            # Check for empty state
            if not getattr(self, 'footer_state', []) and not getattr(self, 'relative_footer_merges', []) and not getattr(self, 'relative_footer_row_heights', {}):
                logger.warning(f"[JsonTemplateStateBuilder] Footer state is empty for '{target_worksheet.title}'. Nothing to restore.")

            # 1. Restore Cell Values & Styles
            for row_idx, row_data in enumerate(getattr(self, 'footer_state', [])):
                # row_idx is already 0-indexed relative to start
                actual_row = footer_start_row + row_idx
                
                for col_idx, cell_info in enumerate(row_data):
                    template_col = col_idx + self.min_col
                    output_col = self._get_mapped_column(template_col)
                    
                    if output_col is None: continue
                    
                    target_cell = target_worksheet.cell(row=actual_row, column=output_col)
                    self._write_cell(target_cell, cell_info, mode=mode)

            # 2. Restore Merged Cells (from relative tuples)
            for merge_tuple in getattr(self, 'relative_footer_merges', []):
                 self._apply_merge(target_worksheet, merge_tuple, start_row_offset=footer_start_row)
                 
            # 3. Restore Row Heights (from relative dict)
            for rel_r, h in getattr(self, 'relative_footer_row_heights', {}).items():
                actual_r = footer_start_row + rel_r
                target_worksheet.row_dimensions[actual_r].height = h
                
        except Exception as e:
            logger.error(
                f"[JsonTemplateStateBuilder] Failed to restore footer on '{target_worksheet.title}': {e}",
                exc_info=True
            )

    @staticmethod
    def _resolve_mode_value(raw_value, mode: str = "standard"):
        """
        Resolves a cell value that may be mode-dependent.
        
        Resolution priority (highest to lowest):
            1. Exact mode-specific override (e.g., 'daf' key when mode='daf')
            2. 'standard' — the UNIVERSAL base value that applies to ALL modes
            3. 'default' — the original/fallback value
        
        This means 'standard' is NOT mode-specific; it acts as a universal
        override that applies to standard, custom, DAF, and any future mode.
        Only a mode-specific key (e.g., 'daf') can take priority over it.
        
        Args:
            raw_value: The raw value from header_content (str, number, or dict).
            mode: The active generation mode ('standard', 'daf', 'custom').
        
        Returns:
            The resolved scalar value.
        """
        if isinstance(raw_value, dict):
            # 1. Exact mode-specific override (e.g., 'daf' key when mode='daf')
            #    This does NOT match 'standard' key when mode='standard' — that's
            #    handled by step 2 as the universal base.
            if mode != "standard" and mode in raw_value:
                return raw_value[mode]
            
            # 2. 'standard' = universal base value (applies to ALL modes)
            if "standard" in raw_value:
                return raw_value["standard"]
                
            # 3. 'default' fallback (original template value)
            return raw_value.get('default', next(iter(raw_value.values()), None))
            
        return raw_value

    def _write_cell(self, cell, info, mode: str = "standard"):
        """
        Writes a single cell's state (value and styles) to an OpenPyXL cell object.
        
        Args:
            cell: The target OpenPyXL Cell object.
            info: A dictionary containing 'value', 'font', 'fill', 'border', 'alignment', 'number_format'.
            mode: Generation mode for resolving mode-dependent values.
        """
        if info['value'] is not None:
            resolved = self._resolve_mode_value(info['value'], mode)
            if resolved is not None:
                cell.value = resolved
        if info['font']: cell.font = copy.copy(info['font'])
        if info['fill']: cell.fill = copy.copy(info['fill'])
        if info['border']: cell.border = copy.copy(info['border'])
        if info['alignment']: cell.alignment = copy.copy(info['alignment'])
        if info['number_format']: cell.number_format = info['number_format']

    def _apply_merge(self, ws, merge_data, start_row_offset=0):
        """
        Applies a merge range to the worksheet.
        
        Args:
            ws: The target worksheet.
            merge_data: Either a string "A1:B2" (absolute) OR a tuple (min_col, min_r, max_col, max_r) (relative).
            start_row_offset: Offset to add to row indices (typically the footer start row).
        """
        min_col, min_row, max_col, max_row = 0, 0, 0, 0
        
        # Determine input type
        if isinstance(merge_data, str):
            # Classic string parsing (used by Header) - Absolute
            from openpyxl.utils.cell import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(merge_data)
            # No offset usually needed for absolute strings, unless shifted?
            # Existing logic was confusing. For Header, we use it as-is.
        elif isinstance(merge_data, tuple) or isinstance(merge_data, list):
            # Relative tuple (used by Footer) - (col, rel_row, col, rel_row)
            min_col, rel_min, max_col, rel_max = merge_data
            min_row = rel_min + start_row_offset
            max_row = rel_max + start_row_offset
            
        # Apply column mapping
        mapped_min_col = self._get_mapped_column(min_col)
        mapped_max_col = self._get_mapped_column(max_col)
        
        if mapped_min_col and mapped_max_col:
            new_range = f"{get_column_letter(mapped_min_col)}{min_row}:{get_column_letter(mapped_max_col)}{max_row}"
            try:
                ws.merge_cells(new_range)
            except ValueError:
                # Overlapping merges can cause this
                pass


