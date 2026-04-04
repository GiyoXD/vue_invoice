import sys
import traceback
import logging
from numpy import rint
import openpyxl
import traceback
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
# from openpyxl.worksheet.dimensions import RowDimension # Not strictly needed for access
from typing import Dict, List, Optional, Tuple, Any

logger = logging.getLogger(__name__)

center_alignment = Alignment(horizontal='center', vertical='center')# --- store_original_merges FILTERED to ignore merges ABOVE row 16 ---
def store_original_merges(workbook: openpyxl.Workbook, sheet_names: List[str]) -> Dict[str, List[Tuple[int, Any, Optional[float]]]]:
    """
    Stores the HORIZONTAL span (colspan), the value of the top-left cell,
    and the height of the starting row for merged ranges in specified sheets,
    ASSUMING all merges are only 1 row high AND **start at row 16 or below**.
    Merges starting above row 16 (row < 16) are ignored.
    WARNING: Does NOT store starting coordinates... (rest of docstring unchanged)

    Args: (args unchanged)

    Returns:
        A dictionary where keys are sheet names and values are lists of
        tuples: (col_span, top_left_cell_value, row_height).
        row_height will be None if the original row had default height.
    """
    original_merges = {}
    logger.info("Storing original merge horizontal spans, top-left values, and row heights (NO coordinates)...")
    logger.debug("(Ignoring merges that start above row 16)")
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            worksheet: Worksheet = workbook[sheet_name] # Type hint for clarity
            merges_data = []
            merged_ranges_copy = list(worksheet.merged_cells.ranges)
            skipped_above_16_count = 0 # Counter for this filter

            for merged_range in merged_ranges_copy:
                min_col, min_row, max_col, max_row = merged_range.bounds

                # --- Check 1: Skip if multi-row ---
                if max_row != min_row:
                    
                    logger.info(f"  Skipping merge {merged_range.coord} on sheet '{sheet_name}' - it spans multiple rows ({min_row} to {max_row}).")
                    continue

                # ***** NEW CHECK 2: Skip if merge starts ABOVE row 16 *****
                if min_row < 16:
                    logger.info(f"  Skipping merge {merged_range.coord} on sheet '{sheet_name}' - starts at row {min_row} (above row 16).") # Keep commented unless needed
                    skipped_above_16_count += 1
                    continue
                # ***** END NEW CHECK *****

                # --- If not skipped, proceed to get span, height, value ---
                col_span = max_col - min_col + 1
                row_height = None # Default to None
                try:
                    # Get Row Height
                    row_dim = worksheet.row_dimensions[min_row]
                    row_height = row_dim.height
                    logger.debug(f"    DEBUG Store: Sheet='{sheet_name}', MergeCoord='{merged_range.coord}', StartRow={min_row}, Storing Height={row_height} (Type: {type(row_height)})")

                    # Get Value
                    top_left_value = worksheet.cell(row=min_row, column=min_col).value

                    # Store Data (span, value, height)
                    merges_data.append((col_span, top_left_value, row_height))

                except KeyError:
                     logger.warning(f"Could not find row dimension for row {min_row} on sheet '{sheet_name}' while getting height. Storing height as None")
                     try:
                         top_left_value = worksheet.cell(row=min_row, column=min_col).value
                     except Exception as val_e:
                         logger.warning(f"Also failed to get value for merge at ({min_row},{min_col}) on sheet '{sheet_name}'. Storing value as None. Error: {val_e}")
                         top_left_value = None
                     merges_data.append((col_span, top_left_value, None))

                except Exception as e:
                    logger.warning(f"Could not get value/height for merge starting at ({min_row},{min_col}) on sheet '{sheet_name}'. Storing value/height as None. Error: {e}")
                    merges_data.append((col_span, None, None))

            original_merges[sheet_name] = merges_data
            logger.info(f"Stored {len(original_merges[sheet_name])} horizontal merge span/value/height entries for sheet '{sheet_name}'")
            # Report skipped count for this filter
            if skipped_above_16_count > 0:
                logger.debug(f"(Skipped {skipped_above_16_count} merges starting above row 16)")
        else:
             logger.warning(f"Sheet '{sheet_name}' specified but not found during merge storage")
             original_merges[sheet_name] = []
    return original_merges

# --- find_and_restore_merges_heuristic remains unchanged (still searches bottom-up, applies stored value/height) ---
def find_and_restore_merges_heuristic(workbook: openpyxl.Workbook,
                                      stored_merges: Dict[str, List[Tuple[int, Any, Optional[float]]]],
                                      processed_sheet_names: List[str],
                                      search_range_str: str = "A16:H200"):
    """
    Attempts to restore merges based on stored HORIZONTAL spans, values, and row heights
    by searching for the value within a specified range (default A16:H200).
    This version is silent, with no detailed logging.

    WARNING: This is a HEURISTIC approach... (rest of docstring unchanged)

    Args: (args unchanged)
    """
    logger.info("Starting merge restoration process...")

    # These counters are still used by the logic but are no longer printed.
    restored_count = 0
    failed_count = 0
    skipped_count = 0
    skipped_duplicate_value_count = 0

    # --- Define search boundaries (critical errors are still reported) ---
    try:
        search_min_col, search_min_row, search_max_col, search_max_row = range_boundaries(search_range_str)
    except TypeError as te:
        logger.error(f"Error processing search range '{search_range_str}'. Check openpyxl version compatibility or range format. Internal error: {te}")
        traceback.print_exc()
        return
    except Exception as e:
        logger.error(f"Invalid search range string '{search_range_str}'. Cannot proceed with restoration. Error: {e}")
        return

    # --- Loop through sheets ---
    for sheet_name in processed_sheet_names:
        if sheet_name in workbook.sheetnames and sheet_name in stored_merges:
            worksheet: Worksheet = workbook[sheet_name]
            original_merges_data = stored_merges[sheet_name]
            successfully_restored_values_on_sheet = set()

            # --- Loop through stored merge info ---
            for col_span, stored_value, stored_height in original_merges_data:

                if col_span <= 1:
                    skipped_count += 1
                    continue

                if stored_value in successfully_restored_values_on_sheet:
                    skipped_duplicate_value_count += 1
                    continue

                found = False
                # --- Search range loop - ROW SEARCH REVERSED ---
                for r in range(search_max_row, search_min_row - 1, -1):
                    for c in range(search_min_col, search_max_col + 1):
                        current_cell = worksheet.cell(row=r, column=c)
                        current_val = current_cell.value

                        if current_val == stored_value:
                            start_row, start_col = r, c
                            end_row = start_row
                            end_col = start_col + col_span - 1

                            # --- Proactively unmerge any conflicting ranges ---
                            merged_ranges_copy = list(worksheet.merged_cells.ranges)
                            for existing_merge in merged_ranges_copy:
                                rows_overlap = (existing_merge.min_row <= end_row) and (existing_merge.max_row >= start_row)
                                cols_overlap = (existing_merge.min_col <= end_col) and (existing_merge.max_col >= start_col)

                                if rows_overlap and cols_overlap:
                                    try:
                                        worksheet.unmerge_cells(str(existing_merge))
                                    except Exception:
                                        # Fails silently as requested
                                        pass

                            # --- Apply the new merge, Row Height, AND Value ---
                            try:
                                worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

                                if stored_height is not None:
                                    try:
                                        worksheet.row_dimensions[start_row].height = stored_height
                                    except Exception:
                                        # Fails silently
                                        pass

                                top_left_cell_to_set = worksheet.cell(row=start_row, column=start_col)
                                top_left_cell_to_set.value = stored_value

                                successfully_restored_values_on_sheet.add(stored_value)
                                restored_count += 1
                                found = True
                                break

                            except Exception:
                                failed_count += 1
                                found = True
                                break

                    if found:
                        break

                if not found:
                    if stored_value not in successfully_restored_values_on_sheet:
                        failed_count += 1

    logger.info("Merge restoration process finished.")

def apply_horizontal_merge(worksheet: Worksheet, row_num: int, num_cols: int, merge_rules: Optional[Dict[str, int]]):
    """
    Applies horizontal merges to a specific row based on a dictionary of rules.
    This is the only function needed for your request.

    Args:
        worksheet: The openpyxl Worksheet object.
        row_num: The 1-based row index to apply merges to.
        num_cols: The total number of columns in the table for validation.
        merge_rules: Dictionary where keys are the starting column index (as a string)
                     and values are the number of columns to span (colspan).
    """
    # Exit if there are no rules to apply
    if not merge_rules:
        return

    logger.info(f"Applying custom merge rules for row {row_num}...")
    for start_col_str, colspan_val in merge_rules.items():
        try:
            start_col = int(start_col_str)
            colspan = int(colspan_val)

            # Skip if the rule is invalid (e.g., merging 1 or fewer columns)
            if start_col < 1 or colspan <= 1:
                continue

            # Calculate the end column and ensure it doesn't exceed the table's width
            end_col = start_col + colspan - 1
            if end_col > num_cols:
                end_col = num_cols

            # Perform the merge and apply center alignment
            worksheet.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
            cell = worksheet.cell(row=row_num, column=start_col)
            cell.alignment = center_alignment
            logger.debug(f"Merged row {row_num} from column {start_col} to {end_col}")

        except (ValueError, TypeError):
            # Ignore if the rule is badly formatted in the JSON (e.g., "A": 5)
            continue


def merge_vertical_cells_in_range(worksheet: Worksheet, scan_col: int, start_row: int, end_row: int, col_id):
    """
    Merges contiguous groups of identical values in a column range.
    
    Walks top-to-bottom through the column, identifying runs of identical
    adjacent values and merging each group individually.
    
    Example: "1-25", "2-25", "2-25", "3-25", "3-25", "3-25"
    → Group 1: row 1 (standalone "1-25")
    → Group 2: rows 2-3 merged ("2-25")
    → Group 3: rows 4-6 merged ("3-25")

    Args:
        worksheet: The openpyxl Worksheet object.
        scan_col: The 1-based column index to scan and merge.
        start_row: The 1-based starting row index.
        end_row: The 1-based ending row index.
    """
    if not all(isinstance(i, int) and i > 0 for i in [scan_col, start_row, end_row]) or start_row >= end_row:
        return

    if col_id == "col_desc":
        # Find the first non-empty value to use as the baseline
        buffer_value = None
        for row_idx in range(start_row, end_row + 1):
            val = worksheet.cell(row=row_idx, column=scan_col).value
            if val:
                buffer_value = val
                break
                
        # Only verify if we found a description
        if buffer_value is not None:
            for row_idx in range(start_row, end_row + 1):
                val = worksheet.cell(row=row_idx, column=scan_col).value
                # If we find actual text that differs from our baseline, abort
                if val and val != buffer_value:
                    return

    # Walk through the column, tracking contiguous groups
    group_start = start_row
    group_value = worksheet.cell(row=start_row, column=scan_col).value
    
    for row_idx in range(start_row + 1, end_row + 2):  # +2 to flush the last group
        if row_idx <= end_row:
            current_value = worksheet.cell(row=row_idx, column=scan_col).value
        else:
            current_value = None  # Sentinel to flush last group
        
        if current_value == group_value and row_idx <= end_row:
            # Same value, extend the group
            continue
        else:
            # Value changed or end of range — merge the previous group if 2+ rows
            group_end = row_idx - 1
            if group_end > group_start and group_value is not None:
                # Skip merging if the value is a pure integer — merging
                # integers causes data loss (e.g. quantities collapsed into one cell).
                # Only text values like "1-25" should be merged.
                skip_int = False
                try:
                    int(group_value)
                    skip_int = True
                except (ValueError, TypeError):
                    pass

                if skip_int:
                    logger.debug(f"  Skipped merge column {scan_col} rows {group_start}-{group_end} — value '{group_value}' is an integer")
                else:
                    try:
                        worksheet.merge_cells(
                            start_row=group_start,
                            start_column=scan_col,
                            end_row=group_end,
                            end_column=scan_col
                        )
                        # Apply center alignment to the merged cell
                        anchor_cell = worksheet.cell(row=group_start, column=scan_col)
                        anchor_cell.alignment = center_alignment
                        logger.debug(f"  Merged column {scan_col} rows {group_start}-{group_end} (value = '{group_value}')")
                    except Exception as e:
                        logger.warning(f"  Failed to merge column {scan_col} rows {group_start}-{group_end}: {e}")
            
            # Start new group
            group_start = row_idx
            group_value = current_value


def apply_horizontal_merge_by_id(
    worksheet: Worksheet,
    row_num: int,
    column_id_map: Dict[str, int],
    num_total_columns: int,
    merge_rules: Dict[str, Dict[str, Any]],
    style_registry=None,
    cell_styler=None
):
    """
    Applies horizontal merges to a specific row based on column IDs.
    Modern ID-driven approach with StyleRegistry support.
    
    Args:
        worksheet: The openpyxl Worksheet object
        row_num: The 1-based row index to apply merges to
        column_id_map: Maps column ID to 1-based column index (e.g., {'col_item': 3})
        num_total_columns: Total number of columns for validation
        merge_rules: Dict where keys are column IDs and values contain merge config
                     e.g., {'col_item': {'rowspan': 2}} - 'rowspan' means horizontal colspan
        style_registry: StyleRegistry instance for ID-driven styling (optional)
        cell_styler: CellStyler instance for applying styles (optional)
    
    Note: 'rowspan' in merge_rules actually means horizontal colspan (legacy naming).
    """
    if not merge_rules or row_num <= 0:
        return

    for col_id, rule_details in merge_rules.items():
        colspan = rule_details.get("rowspan")  # Legacy naming - actually horizontal span
        
        if not isinstance(colspan, int) or colspan <= 1:
            continue
        
        # Get starting column index from ID map
        start_col_idx = column_id_map.get(col_id)
        if not start_col_idx:
            logger.warning(f"Cannot merge: column ID '{col_id}' not found in column_id_map")
            continue
        
        # Calculate end column (respect table boundaries)
        end_col_idx = min(start_col_idx + colspan - 1, num_total_columns)
        
        if start_col_idx >= end_col_idx:
            continue
        
        try:
            # Apply the merge
            worksheet.merge_cells(
                start_row=row_num,
                start_column=start_col_idx,
                end_row=row_num,
                end_column=end_col_idx
            )
            
            # Style the anchor cell
            anchor_cell = worksheet.cell(row=row_num, column=start_col_idx)
            if style_registry and cell_styler:
                style = style_registry.get_style(col_id, context='data')
                cell_styler.apply(anchor_cell, style)
            else:
                # Fallback to simple center alignment
                anchor_cell.alignment = center_alignment
            
            logger.debug(f"Merged row {row_num}, col_id '{col_id}' (cols {start_col_idx}-{end_col_idx})")
            
        except Exception as e:
            logger.error(f"Error merging col_id '{col_id}' on row {row_num}: {e}")