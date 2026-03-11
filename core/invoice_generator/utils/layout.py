from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict, Any, Optional, Tuple
from openpyxl.utils import get_column_letter
from ..styling.models import StylingConfigModel
from ..styling.style_applier import apply_cell_style, apply_header_style
from ..styling.style_config import THIN_BORDER, NO_BORDER, CENTER_ALIGNMENT, LEFT_ALIGNMENT, BOLD_FONT
from decimal import Decimal, InvalidOperation
import re
import traceback
import logging

logger = logging.getLogger(__name__)




def apply_column_widths(worksheet: Worksheet, sheet_styling_config: Optional[StylingConfigModel], header_map: Optional[Dict[str, int]]):
    """
    Sets column widths based on the configuration.

    Args:
        worksheet: The openpyxl Worksheet object.
        sheet_styling_config: Styling configuration containing the 'column_widths' dictionary.
        header_map: Dictionary mapping header text to column index (1-based).
    """
    if not sheet_styling_config or not header_map: return
    column_widths_cfg = sheet_styling_config.columnIdWidths
    if not column_widths_cfg or not isinstance(column_widths_cfg, dict): return
    for header_text, width in column_widths_cfg.items():
        col_idx = header_map.get(header_text)
        if col_idx:
            col_letter = get_column_letter(col_idx)
            try:
                width_val = float(width)
                if width_val > 0: worksheet.column_dimensions[col_letter].width = width_val
                else: pass # Ignore non-positive widths
            except (ValueError, TypeError): pass # Ignore invalid width values
            except Exception as width_err: pass # Log other errors?
        else: pass # Header text not found in map


def _estimate_display_text(cell_value, number_format: str) -> str:
    """
    Estimates the display text of a cell value as Excel would render it.

    For numbers, formats with comma grouping and 4 decimal places.
    For text values, returns str(cell_value) as-is.

    Args:
        cell_value: The raw cell value.
        number_format: The cell's number_format string (unused, kept for API).

    Returns:
        The estimated display text string.
    """
    if cell_value is None:
        return ""

    if isinstance(cell_value, str):
        return cell_value

    if isinstance(cell_value, (int, float)):
        try:
            return f"{cell_value:,.2f}"
        except (ValueError, TypeError):
            pass

    return str(cell_value)


def _header_effective_width(text: str) -> int:
    """
    Calculates the effective display width of a header cell by splitting
    at breakable boundaries like '(' so the header can wrap to a narrower width.

    Does NOT modify the actual cell text — only used for width calculation.

    Examples:
        "Unit price (USD)" → max(len("Unit price"), len("(USD)")) = 10
        "Amount (USD)"     → max(len("Amount"), len("(USD)")) = 6
        "Description"      → 11 (no split)

    Args:
        text: The header cell text.

    Returns:
        The effective character width (longest segment after splitting).
    """
    if not text or not isinstance(text, str):
        return 0

    # Split at existing line breaks first
    segments = text.split('\n')

    # Further split each segment at '(' boundary
    all_parts = []
    for segment in segments:
        if '(' in segment:
            # Split before '(' — e.g. "Unit price (USD)" → ["Unit price ", "(USD)"]
            idx = segment.index('(')
            before = segment[:idx].strip()
            after = segment[idx:].strip()
            if before:
                all_parts.append(before)
            if after:
                all_parts.append(after)
        else:
            all_parts.append(segment.strip())

    if not all_parts:
        return len(text)

    return max(len(part) for part in all_parts)


def auto_fit_dimensions(
    worksheet: Worksheet,
    header_start_row: int,
    data_end_row: int,
    num_columns: int,
    padding: int = 5,
    line_height: float = 15.0,
    min_width: float = 8.0,
    max_width: float = 60.0,
    header_row_start: int = None,
    header_row_end: int = None,
    template_top_end_row: int = None,
    template_bottom_start_row: int = None,
    max_row: int = None
):
    """
    Auto-fits column widths and row heights based on actual cell content.

    Width strategy:
        For each column, scan data rows to find the longest formatted
        display value. Set width = len(formatted_value) * 1.3 + padding.

    Height strategy:
        For each row, count line breaks (\\n) in every cell.
        Set row height = max(line_breaks_in_row) * line_height.
        Only overrides rows where a cell has more than 1 line.

    Args:
        worksheet: The openpyxl Worksheet.
        header_start_row: First data row to scan.
        data_end_row: Last row to scan (last data or footer row).
        num_columns: Total number of columns to scan.
        padding: Extra character units added to the computed width.
        line_height: Points per line for height calculation.
        min_width: Minimum column width to prevent tiny columns.
        max_width: Maximum column width to prevent absurdly wide columns.
        header_row_start: Optional first row of the table header (for header width calc).
        header_row_end: Optional last row of the table header (for header width calc).
        template_top_end_row: Optional row just before the table header starts.
        template_bottom_start_row: Optional row just after the table footer ends.
        max_row: Optional max row in the worksheet (for scanning bottom template).
    """
    if header_start_row <= 0 or data_end_row <= 0 or num_columns <= 0:
        logger.warning(f"auto_fit_dimensions: invalid bounds (header_start={header_start_row}, data_end={data_end_row}, cols={num_columns})")
        return

    logger.info(f"auto_fit_dimensions: scanning rows {header_start_row}-{data_end_row}, {num_columns} columns")

    # Build a set of cells that belong to multi-column merges.
    # These cells should be skipped during width calculation because their
    # value spans multiple columns and would inflate a single column's width.
    merged_cell_coords = set()
    for merged_range in worksheet.merged_cells.ranges:
        # Only skip if the merge spans more than 1 column
        if merged_range.max_col > merged_range.min_col:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cell_coords.add((row, col))
    if merged_cell_coords:
        logger.debug(f"auto_fit_dimensions: found {len(merged_cell_coords)} cells in multi-column merges (will skip)")

    # Track the widest display text length per column + whether it's text or number
    # Format: {col_idx: (max_len, is_text)}
    col_max_info: Dict[int, tuple] = {}

    # --- Header width pass: scan header rows with smart line-breaking ---
    if header_row_start and header_row_end and header_row_start > 0 and header_row_end >= header_row_start:
        from openpyxl.styles import Alignment
        logger.info(f"auto_fit_dimensions: scanning header rows {header_row_start}-{header_row_end} for width")
        for row_idx in range(header_row_start, header_row_end + 1):
            for col_idx in range(1, num_columns + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                # Skip cells in multi-column merges
                if (row_idx, col_idx) in merged_cell_coords:
                    continue

                header_text = str(cell.value)
                
                # Insert \n before '(' to force a clean wrap if it isn't already there
                if '(' in header_text and '\n(' not in header_text:
                    import re
                    # Replace an optional space followed by '(' with '\n('
                    new_text = re.sub(r' ?\(', '\n(', header_text, count=1)
                    cell.value = new_text
                    header_text = new_text
                
                effective_len = _header_effective_width(header_text)

                current_max, current_is_text = col_max_info.get(col_idx, (0, False))
                if effective_len > current_max:
                    # Use is_text=False so the compact formula (max_len + padding) is used.
                    # Headers wrap via wrap_text=True, so we don't need the inflated text multiplier.
                    col_max_info[col_idx] = (effective_len, False)

                # Enable wrap_text so Excel wraps at the narrower width
                existing_alignment = cell.alignment
                cell.alignment = Alignment(
                    horizontal=existing_alignment.horizontal,
                    vertical=existing_alignment.vertical,
                    text_rotation=existing_alignment.text_rotation,
                    indent=existing_alignment.indent,
                    shrink_to_fit=existing_alignment.shrink_to_fit,
                    wrap_text=True
                )

    # --- Data width pass: scan data rows normally ---
    for row_idx in range(header_start_row, data_end_row + 1):
        max_lines_in_row = 1  # At least 1 line per row

        for col_idx in range(1, num_columns + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell_value = cell.value

            if cell_value is None:
                continue

            # Skip cells in multi-column merges — their value spans
            # multiple columns and would inflate this column's width
            if (row_idx, col_idx) in merged_cell_coords:
                continue

            # Skip formula cells — their string representation (=SUM...)
            # doesn't reflect displayed width
            if isinstance(cell_value, str) and cell_value.startswith('='):
                continue

            # Track if this cell is text (not a number/formula)
            is_text = isinstance(cell_value, str)

            # Use formatted display text instead of raw str()
            text = _estimate_display_text(cell_value, cell.number_format)

            # --- Width: track longest display text per column ---
            # For multi-line cells, use the longest single line
            lines = text.split('\n')
            longest_line_len = max(len(line) for line in lines) if lines else 0

            current_max, current_is_text = col_max_info.get(col_idx, (0, False))
            if longest_line_len > current_max:
                col_max_info[col_idx] = (longest_line_len, is_text)

            # --- Height: track max line count per row ---
            line_count = len(lines)
            if line_count > max_lines_in_row:
                max_lines_in_row = line_count

        # Apply row height if there are line breaks (more than 1 line)
        if max_lines_in_row > 1:
            computed_height = max_lines_in_row * line_height
            worksheet.row_dimensions[row_idx].height = computed_height
            logger.debug(f"auto_fit row {row_idx}: height={computed_height} ({max_lines_in_row} lines)")

    # --- Template Last Column Width Pass ---
    # Scan template rows (top and bottom) for the last column only to prevent text spill
    last_col_idx = num_columns
    
    def scan_template_rows_for_last_col(start_r, end_r):
        if not (start_r > 0 and end_r >= start_r):
            return
            
        # Build a set of rows where the last column is part of a horizontal merge
        merged_last_col_rows = set()
        for merged_range in worksheet.merged_cells.ranges:
            # If the merge spans multiple columns AND includes the last column
            if merged_range.max_col > merged_range.min_col:
                if merged_range.min_col <= last_col_idx <= merged_range.max_col:
                    for row in range(merged_range.min_row, merged_range.max_row + 1):
                        merged_last_col_rows.add(row)

        logger.info(f"auto_fit_dimensions: scanning template rows {start_r}-{end_r} in column {last_col_idx} for spillover")
        for r_idx in range(start_r, end_r + 1):
            # Skip if this row's last column is part of any horizontal merge
            if r_idx in merged_last_col_rows:
                continue
                
            cell = worksheet.cell(row=r_idx, column=last_col_idx)
            if cell.value is None:
                continue
                
            cell_value = cell.value
            is_text = isinstance(cell_value, str)
            text = _estimate_display_text(cell_value, cell.number_format)
            
            lines = text.split('\n')
            longest_line_len = max(len(line) for line in lines) if lines else 0
            
            if longest_line_len > 0:
                logger.info(f"auto_fit_dimensions: LAST COL (r={r_idx}, c={last_col_idx}) found UNMERGED text of len {longest_line_len}: {repr(text)}")
            
            current_max, current_is_text = col_max_info.get(last_col_idx, (0, False))
            if longest_line_len > current_max:
                col_max_info[last_col_idx] = (longest_line_len, is_text)

    if template_top_end_row and template_top_end_row >= 1:
        scan_template_rows_for_last_col(1, template_top_end_row)
        
    if template_bottom_start_row and max_row and max_row >= template_bottom_start_row:
        scan_template_rows_for_last_col(template_bottom_start_row, max_row)

    # Apply column widths — text gets 1.2x multiplier (wider font chars), numbers stay 1:1
    for col_idx, (max_len, is_text) in col_max_info.items():
        if is_text:
            # Text chars are wider than digits in proportional fonts
            computed_width = max_len * 1.2 + padding + 5
        else:
            computed_width = max_len + padding
        # Clamp between min and max
        computed_width = max(min_width, min(computed_width, max_width))
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = computed_width
        logger.debug(f"auto_fit col {col_letter}: width={computed_width:.1f} (max_len={max_len})")

    logger.info(f"auto_fit_dimensions: applied {len(col_max_info)} column widths")


def calculate_header_dimensions(header_layout: List[Dict[str, Any]]) -> Tuple[int, int]:
    """
    Calculates the total number of rows and columns a header will occupy.
    """
    if not header_layout:
        return (0, 0)
    num_rows = max(cell.get('row', 0) + cell.get('rowspan', 1) for cell in header_layout)
    num_cols = max(cell.get('col', 0) + cell.get('colspan', 1) for cell in header_layout)
    return (num_rows, num_cols)

def merge_contiguous_cells_by_id(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    col_id_to_merge: str,
    column_id_map: Dict[str, int]
):
    """
    Finds and merges contiguous vertical cells within a column that have the same value.
    This is called AFTER all data has been written to the sheet.
    """
    col_idx = column_id_map.get(col_id_to_merge)
    if not col_idx or start_row >= end_row:
        return

    current_merge_start_row = start_row
    value_to_match = worksheet.cell(row=start_row, column=col_idx).value

    for row_idx in range(start_row + 1, end_row + 2):
        cell_value = worksheet.cell(row=row_idx, column=col_idx).value if row_idx <= end_row else object()
        if cell_value != value_to_match:
            if row_idx - 1 > current_merge_start_row:
                if value_to_match is not None and str(value_to_match).strip():
                    try:
                        worksheet.merge_cells(
                            start_row=current_merge_start_row,
                            start_column=col_idx,
                            end_row=row_idx - 1, end_column=col_idx
                        )
                    except Exception as e:
                        logger.error(f"Could not merge cells for ID {col_id_to_merge} from row {current_merge_start_row} to {row_idx - 1}. Error: {e}")
            current_merge_start_row = row_idx
            if row_idx <= end_row:
                value_to_match = cell_value


                value_to_match = cell_value