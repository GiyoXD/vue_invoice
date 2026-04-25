# --- START OF FULL REFACTORED FILE: sheet_parser.py ---

import re
import logging
from typing import Dict, List, Optional, Tuple, Any, Union
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation

# --- Loop Profiler (non-invasive measurement) ---
from core.utils.loop_profiler import loop_profiler, tick

# Import config values, now including the new pattern-matching configs
from .config import (
    TARGET_HEADERS_MAP,
    HEADER_SEARCH_ROW_RANGE,
    HEADER_SEARCH_COL_RANGE,
    HEADER_IDENTIFICATION_PATTERN,
    STOP_EXTRACTION_ON_EMPTY_COLUMN,
    MAX_DATA_ROWS_TO_SCAN,
    DISTRIBUTION_BASIS_COLUMN,
    COLUMNS_TO_DISTRIBUTE,
    # --- Dependencies for the new smart function ---
    EXPECTED_HEADER_DATA_TYPES,
    EXPECTED_HEADER_PATTERNS,
    HEADERLESS_COLUMN_PATTERNS,
    EXPECTED_HEADER_VALUES,
)


# --- PRE-COMPUTED: Reverse alias lookup (built once at import time) ---
# Maps UPPERCASED alias text → list of canonical names it could match.
# This replaces the O(N×M) per-cell scan with O(1) dict lookup.
def _build_alias_lookup() -> Dict[str, List[str]]:
    """Build {ALIAS_UPPER: [canonical_name, ...]} from TARGET_HEADERS_MAP."""
    lookup: Dict[str, List[str]] = {}
    for canonical, aliases in TARGET_HEADERS_MAP.items():
        for alias in aliases:
            key = str(alias).upper()
            if key not in lookup:
                lookup[key] = []
            lookup[key].append(canonical)
    return lookup

_ALIAS_REVERSE_LOOKUP = _build_alias_lookup()


# --- NEW: Helper functions for smart validation ---

def _is_numeric(value: Any) -> bool:
    """Helper to check if a value is a number."""
    if value is None:
        return False
    return isinstance(value, (int, float, Decimal))

def _is_string_like(value: Any) -> bool:
    """Helper to check if a value is a non-empty string or a number."""
    if value is None:
        return False
    if isinstance(value, str) and value.strip():
        return True
    if _is_numeric(value):
        return True
    return False

def _matches_any_pattern(value: Any, patterns: Union[str, List[str]]) -> bool:
    """
    Helper to check if a value's string representation matches ANY of the regex patterns in a list.
    """
    # Convert value to a stripped string for reliable matching. Handles numbers, None, etc.
    value_str = str(value or '').strip()
    if not value_str:
        return False

    # Ensure patterns is always a list for iteration
    if isinstance(patterns, str):
        patterns_list = [patterns]
    else:
        patterns_list = patterns

    # Check against each pattern in the list
    for pattern in patterns_list:
        try:
            if re.match(pattern, value_str):
                # If any pattern matches, we return True immediately
                return True
        except re.error as e:
            logging.error(f"[Pattern Check] Invalid regex pattern provided in config '{pattern}': {e}")
            continue # Try the next pattern
            
    # If no patterns matched after checking all of them
    return False


# --- SMART HEADER DETECTION (3-function decomposition) ---

def _score_cell(canonical_name: str, data_value: Any, row_num: int, col_num: int) -> Optional[int]:
    """
    Score a single cell's data value against a candidate canonical column name.
    
    Returns:
        int score (>0 if valid candidate), or None if hard-rejected.
        - 25: Exact value match (e.g. pallet_count == 1)
        - 15: Regex pattern match (e.g. CBM matches L*W*H)
        -  5: Type match (numeric/string as expected)
        -  4: Headerless pattern match (e.g. empty header but data looks like CBM)
        -  1: Mercy rule (empty data or fallback string)
        - None: Hard reject (pattern defined but data failed)
    """
    score = 0
    used_strict_value_check = False

    # Priority 1: Exact value check (e.g. col_pallet_count must be 1)
    allowed_values = EXPECTED_HEADER_VALUES.get(canonical_name)
    if allowed_values is not None:
        used_strict_value_check = True
        processed_data_value = int(data_value) if isinstance(data_value, str) and data_value.isdigit() else data_value
        if processed_data_value in allowed_values:
            score = 25

    # Priority 2: Pattern or type check
    if not used_strict_value_check:
        patterns_to_check = EXPECTED_HEADER_PATTERNS.get(canonical_name)
        if patterns_to_check:
            if _matches_any_pattern(data_value, patterns_to_check):
                score = 15
            else:
                # Pattern was defined but data FAILED — hard reject.
                logging.debug(
                    f"[score_cell] Row {row_num} Col {col_num}: '{canonical_name}' "
                    f"pattern check FAILED for data '{data_value}'. Hard reject."
                )
                return None
        else:
            allowed_types = EXPECTED_HEADER_DATA_TYPES.get(canonical_name, ['string', 'numeric'])
            if ('numeric' in allowed_types and _is_numeric(data_value)) or \
               ('string' in allowed_types and _is_string_like(data_value)):
                score = 5

    # Priority 3: Mercy rule — don't let empty cells kill valid header matches
    if score == 0:
        if data_value is None or (isinstance(data_value, str) and not data_value.strip()):
            score = 1
        else:
            allowed_types = EXPECTED_HEADER_DATA_TYPES.get(canonical_name, ['string', 'numeric'])
            if 'string' in allowed_types:
                score = 1

    return score


@loop_profiler.watch("_process_row")
def _process_row(sheet: Worksheet, row_num: int) -> Tuple[Dict[str, str], int]:
    """
    Process a single row: scan all columns, score candidates, resolve ties.
    
    Returns:
        Tuple of (column_mapping dict, total_row_score).
        column_mapping: {canonical_name: column_letter, ...}
    """
    all_column_candidates: Dict[int, List[Dict]] = {}

    # --- Phase 1: Score every column in this row ---
    for col_num in range(HEADER_SEARCH_COL_RANGE[0], HEADER_SEARCH_COL_RANGE[1] + 1):
        header_cell = sheet.cell(row=row_num, column=col_num)
        header_value = str(header_cell.value or '').strip().upper()
        data_value = sheet.cell(row=row_num + 1, column=col_num).value
        
        col_scores = []

        if header_value:
            # O(1) lookup via pre-built reverse alias dict (was O(N×M) per cell)
            candidate_canonicals = _ALIAS_REVERSE_LOOKUP.get(header_value, [])
            tick("_process_row", sub="alias_lookups")
            for canonical_name in candidate_canonicals:
                score = _score_cell(canonical_name, data_value, row_num, col_num)
                if score is not None and score > 0:
                    col_scores.append({'score': score, 'name': canonical_name})

        # ALWAYS check strong data patterns (formerly just headerless), even if there's a header.
        # This prevents generic headers (like "PO" -> col_po) from stealing TTX PO data (25xxxxxx).
        for canonical_name, patterns in HEADERLESS_COLUMN_PATTERNS.items():
            if _matches_any_pattern(data_value, patterns):
                existing = next((c for c in col_scores if c['name'] == canonical_name), None)
                if existing:
                    existing['score'] = max(existing['score'], 15)
                else:
                    col_scores.append({'score': 16, 'name': canonical_name})

        if col_scores:
            all_column_candidates[col_num] = col_scores

    # --- Phase 2: Resolve candidates into a mapping ---
    potential_mapping: Dict[str, str] = {}
    current_row_score = 0
    processed_canonicals: set = set()

    # Tie-breaking: unit_price vs amount (both numeric, same score)
    unit_amt_tie_cols = [
        col for col, candidates in all_column_candidates.items()
        if {c['name'] for c in candidates} == {'col_unit_price', 'col_amount'}
        and all(c['score'] == 5 for c in candidates)
    ]
    if len(unit_amt_tie_cols) == 2:
        col1, col2 = sorted(unit_amt_tie_cols)
        # Heuristic: Unit Price is usually to the left of Amount
        potential_mapping['col_unit_price'] = get_column_letter(col1)
        potential_mapping['col_amount'] = get_column_letter(col2)
        processed_canonicals.update(['col_unit_price', 'col_amount'])
        current_row_score += 10  # Bonus for resolving tie
        del all_column_candidates[col1], all_column_candidates[col2]

    # Greedy selection for remaining columns
    for col_num, candidates in sorted(all_column_candidates.items()):
        valid_candidates = [c for c in candidates if c['name'] not in processed_canonicals]
        if not valid_candidates:
            continue

        best_candidate = sorted(valid_candidates, key=lambda x: x['score'], reverse=True)[0]
        potential_mapping[best_candidate['name']] = get_column_letter(col_num)
        processed_canonicals.add(best_candidate['name'])
        current_row_score += best_candidate['score']

    return potential_mapping, current_row_score


@loop_profiler.watch("find_and_map_smart_headers")
def find_and_map_smart_headers(sheet: Worksheet) -> Optional[Tuple[int, Dict[str, str]]]:
    """
    Finds and maps headers using a scoring system. Evaluates all rows in the
    search range and selects the one with the highest cumulative score.
    
    Delegates to:
        - _score_cell(): Scores a single canonical candidate against cell data.
        - _process_row(): Scans all columns in a row, resolves ties, returns mapping.
    
    Returns:
        Tuple of (header_row_number, column_mapping) or None if no valid row found.
    """
    prefix = "[find_and_map_smart_headers]"
    logging.info(f"{prefix} Starting best-fit header search...")

    best_result: Optional[Tuple[int, Dict[str, str]]] = None
    highest_row_score = 0

    for row_num in range(HEADER_SEARCH_ROW_RANGE[0], HEADER_SEARCH_ROW_RANGE[1] + 1):
        if row_num + 1 > sheet.max_row:
            continue

        tick("find_and_map_smart_headers", sub="rows_scanned")
        potential_mapping, current_row_score = _process_row(sheet, row_num)

        logging.debug(f"{prefix} Row {row_num} | Score: {current_row_score} | Mapping: {potential_mapping}")

        if len(potential_mapping) >= 3 and current_row_score > highest_row_score:
            highest_row_score = current_row_score
            best_result = (row_num, potential_mapping)
            logging.info(f"{prefix} New best candidate row at {row_num} with score {highest_row_score}.")

    if best_result:
        logging.info(f"{prefix} SUCCESS: Header row confirmed at {best_result[0]} with score {highest_row_score}.")
        return best_result

    logging.error(f"{prefix} FAILED: No row passed smart validation.")
    return None


@loop_profiler.watch("extract_multiple_tables")
def extract_multiple_tables(sheet, header_rows: List[int], column_mapping: Dict[str, str]) -> List[List[Dict[str, Any]]]:
    """
    Extracts data for multiple tables defined by header_rows using the validated column_mapping.
    Returns a list of tables. Each table is a list of row dictionaries.
    """
    if not header_rows or not column_mapping:
        logging.warning("[extract_multiple_tables] No header rows or column mapping provided.")
        return []

    all_tables_data: List[List[Dict[str, Any]]] = []
    stop_col_letter = column_mapping.get(STOP_EXTRACTION_ON_EMPTY_COLUMN)
    prefix = "[extract_multiple_tables]"

    logging.info(f"{prefix} Starting extraction for {len(header_rows)} tables: {header_rows} (Row List Format)")

    for i, header_row in enumerate(header_rows):
        table_index = i + 1
        start_data_row = header_row + 1
        
        if i + 1 < len(header_rows):
            max_possible_end_row = header_rows[i + 1]
        else:
            max_possible_end_row = sheet.max_row + 1
            
        scan_limit_row = start_data_row + MAX_DATA_ROWS_TO_SCAN
        end_data_row = min(max_possible_end_row, scan_limit_row)

        if start_data_row >= end_data_row:
            all_tables_data.append([])
            continue

        logging.info(f"{prefix} Table {table_index}: Extracting Data Rows {start_data_row} to {end_data_row - 1}")
        current_table_data: List[Dict[str, Any]] = []

        # HOISTED: These never change per-row — build once per table
        col_letter_to_canonical = {v: k for k, v in column_mapping.items()}
        _NUMERIC_COLS_TO_CLEAN = {
            'col_amount', 'col_unit_price', 'col_qty_sf',
            'col_net', 'col_gross', 'col_cbm',
        }
        
        for current_row in range(start_data_row, end_data_row):
            tick("extract_multiple_tables", sub="rows_iterated")
            if stop_col_letter:
                stop_cell_value = sheet[f"{stop_col_letter}{current_row}"].value
                if stop_cell_value is None or (isinstance(stop_cell_value, str) and not stop_cell_value.strip()):
                    logging.info(f"{prefix} Stopping extraction for Table {table_index} at row {current_row}: Empty cell in stop column '{STOP_EXTRACTION_ON_EMPTY_COLUMN}'.")
                    break

            row_dict: Dict[str, Any] = {}

            for col_letter, canonical_name in col_letter_to_canonical.items():
                cell_value = sheet[f"{col_letter}{current_row}"].value
                # Convert raw floats to Decimal immediately via str() to avoid
                # IEEE 754 imprecision (e.g. 299.2 stored as 299.20000000000005).
                if isinstance(cell_value, float):
                    processed_value = Decimal(str(cell_value))
                    # Second guard: if Excel itself computed a float with noise,
                    # str() may still carry it (e.g. 9795.400000000001).
                    # Quantize to 7 decimal places — covers the worst-case real
                    # precision while stripping IEEE noise at positions 14+.
                    # NOTE: No .normalize() — it converts round numbers to 1E+3 notation.
                    if canonical_name in _NUMERIC_COLS_TO_CLEAN:
                        processed_value = processed_value.quantize(Decimal('0.0000000001'))
                elif isinstance(cell_value, str):
                    processed_value = cell_value.strip()
                else:
                    processed_value = cell_value
                row_dict[canonical_name] = processed_value
            
            # Additional logic to skip completely empty rows (optional but good practice)
            if any(v is not None and (not isinstance(v, str) or v.strip() != "") for v in row_dict.values()):
                current_table_data.append(row_dict)

        all_tables_data.append(current_table_data)
        logging.info(f"{prefix} Successfully stored {len(current_table_data)} rows for Table Index {table_index}.")
        
    return all_tables_data


@loop_profiler.watch("find_all_header_rows")
def find_all_header_rows(sheet, search_pattern, row_range, col_range, start_after_row: int = 0) -> List[int]:
    """
    Finds all 1-indexed row numbers containing a header based on a pattern,
    optionally starting the search after a specific row.
    """
    found_rows: set[int] = set()
    try:
        regex = re.compile(search_pattern, re.IGNORECASE)
        start_row = max(row_range[0], start_after_row + 1)
        max_row_to_search = min(row_range[1], sheet.max_row)
        max_col_to_search = min(col_range[1], sheet.max_column)

        if start_row > max_row_to_search:
             return []

        logging.info(
            f"[find_all_header_rows] Searching for additional headers using pattern '{search_pattern}' "
            f"in rows {start_row}-{max_row_to_search}"
        )

        for r_idx in range(start_row, max_row_to_search + 1):
            tick("find_all_header_rows", sub="rows_scanned")
            for c_idx in range(col_range[0], max_col_to_search + 1):
                tick("find_all_header_rows", sub="cells_checked")
                cell = sheet.cell(row=r_idx, column=c_idx)
                if cell.value is not None:
                    cell_value_str = str(cell.value).strip()
                    if regex.search(cell_value_str):
                        found_rows.add(r_idx)
                        break
        
        if not found_rows:
            return []

        header_rows = sorted(list(found_rows))
        logging.info(f"[find_all_header_rows] Found {len(header_rows)} additional header rows at: {header_rows}")
        return header_rows

    except Exception as e:
        logging.error(f"[find_all_header_rows] Error finding header rows: {e}", exc_info=True)
        return []

