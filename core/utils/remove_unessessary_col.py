import openpyxl
import os
import sys
from copy import copy

def copy_cell(src, dst):
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)

def run_cli():
    if len(sys.argv) < 5:
        print("Usage: python script.py [file.xlsx] [sheet_name] [header_start_row] [header_end_row]")
        return

    file_path = sys.argv[1]
    sheet_name = sys.argv[2]
    h_start = int(sys.argv[3])
    h_end = int(sys.argv[4])

    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' not found.")
        return

    base, ext = os.path.splitext(file_path)
    output_path = f"{base}_cleaned{ext}"

    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    max_col = ws.max_column
    cols_to_delete = []

    # 1. Scan UP and DOWN. Skip header rows.
    scan_up_start = max(1, h_start - 500)
    scan_up_end = h_start - 1
    
    scan_down_start = h_end + 1
    scan_down_end = h_end + 500

    for col_idx in range(1, max_col + 1):
        has_data = False
        
        # Scan UP
        for row_idx in range(scan_up_start, scan_up_end + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None and str(cell_val).strip() != "":
                has_data = True
                break
                
        # Scan DOWN
        if not has_data:
            for row_idx in range(scan_down_start, scan_down_end + 1):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None and str(cell_val).strip() != "":
                    has_data = True
                    break
                    
        if not has_data:
            cols_to_delete.append(col_idx)

    # 2. Unmerge. Shift data to safe column. Save bounds.
    merges = list(ws.merged_cells.ranges)
    surviving_merges = []

    for m_range in merges:
        min_c = m_range.min_col
        max_c = m_range.max_col
        min_r = m_range.min_row
        max_r = m_range.max_row
        
        ws.unmerge_cells(str(m_range))
        
        first_safe_c = None
        last_safe_c = None
        for c in range(min_c, max_c + 1):
            if c not in cols_to_delete:
                if first_safe_c is None:
                    first_safe_c = c
                last_safe_c = c
                
        if first_safe_c:
            src_cell = ws.cell(row=min_r, column=min_c)
            dst_cell = ws.cell(row=min_r, column=first_safe_c)
            if src_cell != dst_cell:
                copy_cell(src_cell, dst_cell)
                
            surviving_merges.append({
                'start_r': min_r, 'end_r': max_r,
                'start_c': first_safe_c, 'end_c': last_safe_c
            })

    # 3. Delete columns backward.
    cols_to_delete.sort(reverse=True)
    for c in cols_to_delete:
        ws.delete_cols(c)

    # 4. Calculate offset. Re-merge.
    for m in surviving_merges:
        n_start = m['start_c'] - sum(1 for x in cols_to_delete if x < m['start_c'])
        n_end = m['end_c'] - sum(1 for x in cols_to_delete if x < m['end_c'])
        
        if n_end > n_start or m['end_r'] > m['start_r']:
            ws.merge_cells(start_row=m['start_r'], start_column=n_start, end_row=m['end_r'], end_column=n_end)

    wb.save(output_path)
    print(f"Success. New file: {output_path}")

if __name__ == "__main__":
    run_cli()