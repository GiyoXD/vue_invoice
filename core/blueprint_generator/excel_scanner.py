"""
Template Analyzer - Extracts structure from Excel templates for auto config generation.

This module analyzes Excel template files to extract:
- Sheet names and structure
- Header row positions
- Column layouts (merged cells, widths)
- Font/style information
- Data source hints (aggregation vs processed_tables)
"""

import logging
import re
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from .rules import BlueprintRules
from core.utils.snitch import snitch
from .utils.footer_scanner import FooterInfo, scan_footer

logger = logging.getLogger(__name__)


@dataclass
class ColumnInfo:
    """Information about a single column."""
    id: str
    header: str
    col_index: int  # 1-based
    width: float
    format: str = "@"
    alignment: str = "center"
    rowspan: int = 1
    colspan: int = 1
    children: List['ColumnInfo'] = field(default_factory=list)
    wrap_text: bool = False



@dataclass
class SheetAnalysis:
    """Complete analysis of a single sheet."""
    name: str
    header_row: int
    columns: List[ColumnInfo]
    data_source: str  # "aggregation" or "processed_tables_multi"
    header_font: Dict[str, Any]
    data_font: Dict[str, Any]
    row_heights: Dict[str, float]  # "header", "data", "footer" -> height
    has_multi_row_header: bool = False
    static_content_hints: Dict[str, List[str]] = field(default_factory=dict)
    footer_info: Optional[FooterInfo] = None

    def to_legacy_dict(self) -> Dict[str, Any]:
        """Convert to legacy JSON format for frontend compatibility."""
        # Map columns to legacy header_positions
        header_positions = []
        for col in self.columns:
            header_positions.append({
                "keyword": col.header,
                "col_id": col.id,
                "row": self.header_row,
                "column": col.col_index
            })
            # Also add children if any
            if col.children:
                for child in col.children:
                    header_positions.append({
                        "keyword": child.header,
                        "row": self.header_row + 1,
                        "column": child.col_index
                    })

        return {
            "sheet_name": self.name,
            "header_positions": header_positions,
            "start_row": self.header_row + 1 if self.header_row else 1,
            "unconfirmed_footer": self.footer_info.total_text if (self.footer_info and not self.footer_info.is_exact) else None
        }


@dataclass
class TemplateAnalysisResult:
    """Complete template analysis result."""
    file_path: str
    customer_code: str
    sheets: List[SheetAnalysis]
    warnings: List[str] = field(default_factory=list)

    def to_legacy_dict(self) -> Dict[str, Any]:
        """Convert to legacy JSON format for frontend compatibility."""
        return {
            "file_path": self.file_path,
            "sheets": [sheet.to_legacy_dict() for sheet in self.sheets],
            "warnings": self.warnings
        }
    

class ExcelLayoutScanner:
    """Analyzes Excel templates to extract structure for config generation."""
    
    def __init__(self):
        self.logger = logging.getLogger(self.__class__.__name__)

    def _get_cell_value(self, cell) -> str:
        """Safe string value from cell."""
        if cell.value is None:
            return ""
        return str(cell.value).strip()




    def _is_potential_header_row(self, cells) -> bool:
        """
        Check if a row is a potential header row (Legacy Logic).
        Rejects rows that are primarily numeric (>30%).
        """
        if not cells:
            return False
            
        numeric_count = 0
        total_count = len(cells)
        
        for cell in cells:
            val = cell.value
            if val is None:
                continue
                
            is_numeric = False
            if isinstance(val, (int, float)):
                is_numeric = True
            else:
                s_val = str(val).strip()
                # Check for number format like -123.45
                if s_val and re.match(r'^-?\d+(\.\d+)?$', s_val):
                    is_numeric = True
            
            if is_numeric:
                numeric_count += 1
        
        # If more than 30% of cells are numeric, it's likely a data row
        return (numeric_count / total_count) <= 0.3

    def _find_header_row_structural(self, worksheet: Worksheet, max_rows: int = 50) -> Optional[int]:
        """
        Legacy Structural Header Detection (Fallback).
        Finds row that is "widest" (most columns) and "text-heavy".
        """
        candidates = []
        
        for row in range(1, min(worksheet.max_row + 1, max_rows)):
            cells = []
            max_col_idx = 0
            # Cap structural scan at 50 columns
            for col in range(1, min(worksheet.max_column + 1, 50)):
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():
                    cells.append(cell)
                    max_col_idx = col
            
            if cells and self._is_potential_header_row(cells):
                candidates.append({
                    'row_num': row,
                    'max_col': max_col_idx,
                    'cell_count': len(cells)
                })
        
        if not candidates:
            self.logger.warning("Structural Fallback: No structural candidates found.")
            return None
            
        # Filter for widest rows (tolerance 2)
        max_width = max(c['max_col'] for c in candidates)
        width_tolerance = 2
        wide_candidates = [c for c in candidates if c['max_col'] >= (max_width - width_tolerance)]
        
        self.logger.info(f"Structural Fallback: Found {len(candidates)} candidates. Max Width: {max_width}. Wide Candidates: {[c['row_num'] for c in wide_candidates]}")
        
        if not wide_candidates:
            wide_candidates = candidates
            
        # Sort by cell_count desc, then row_num asc
        wide_candidates.sort(key=lambda x: (-x['cell_count'], x['row_num']))
        
        best = wide_candidates[0]
        self.logger.info(f"Structural Fallback: Selected Row {best['row_num']} (Cells={best['cell_count']}, MaxCol={best['max_col']})")
        
        return best['row_num']
 
    def _find_header_row(self, worksheet: Worksheet, mapping_config: Optional[Dict[str, Any]] = None) -> Tuple[Optional[int], List[Tuple[int, str]]]:
        """Find the header row by looking for known column keywords."""
        max_scan_rows = 50
        max_matches = 0
        # Score tuple: (matches, text_count)
        max_score = (0, 0)
        best_row = None
        best_header_cells = []
        
        for row in range(1, min(worksheet.max_row + 1, max_scan_rows)):
            matches = 0
            text_count = 0
            header_cells = []
            
            has_content = False
            # Cap header scan at 25 (Col Y)
            for col in range(1, min(worksheet.max_column + 1, 26)):
                cell = worksheet.cell(row=row, column=col)
                value = self._get_cell_value(cell)
                if value:
                    has_content = True
                    text_count += 1
                    
                    # strong, rule-based match check
                    is_match = False
                    
                    # 1. Check user mapping
                    clean_val = "".join(value.lower().split())
                    if mapping_config:
                         mappings = mapping_config.get('header_text_mappings', {}).get('mappings', {})
                         if value in mappings or any("".join(m.lower().split()) == clean_val for m in mappings):
                             is_match = True
                    
                    # 2. Check system rules
                    if not is_match and BlueprintRules.get_column_by_keyword(value):
                        is_match = True
                        
                    if is_match:
                        matches += 1
                        
                    header_cells.append((col, value))
            
            # HEADER ROW SELECTION ALGORITHM:
            # 1. Row must have AT LEAST 3 Valid Matches (keyword hits).
            # 2. Rank primarily by matches (more keyword hits = stronger header candidate).
            # 3. Tie-breaker: text_count (more occupied cells = wider row).
            # 4. Tie-breaker: Topmost row (implicit: we iterate top-down, strict >).
            
            # Validity Threshold
            if matches >= 3:
                 is_better = False
                 if best_row is None:
                     is_better = True
                 else:
                     (best_matches, best_text_count) = max_score
                     
                     if matches > best_matches:
                         is_better = True
                     elif matches == best_matches:
                         if text_count > best_text_count:
                             is_better = True
                             
                 if is_better:
                     max_score = (matches, text_count)
                     best_row = row
                     best_header_cells = header_cells
            
        if best_row:
             self.logger.info(f"Header detection: Found header at row {best_row} with Score(matches={max_score[0]}, text={max_score[1]}).")
        else:
             self.logger.warning("Header detection: No header row found meeting threshold (min 3 matches). Trying Legacy Structural Fallback...")
             # Fallback to Legacy Structural Detection
             fallback_row = self._find_header_row_structural(worksheet, max_rows=max_scan_rows)
             if fallback_row:
                 self.logger.info(f"Header detection (Fallback): Found structural header at row {fallback_row}.")
                 best_row = fallback_row
                 # Re-extract cells with CAP
                 best_header_cells = []
                 for col in range(1, min(worksheet.max_column + 1, 50)):
                     cell = worksheet.cell(row=fallback_row, column=col)
                     val = self._get_cell_value(cell)
                     if val:
                         best_header_cells.append((col, val))
             else:
                 self.logger.warning("Header detection: Fallback also failed.")
             
        return best_row, best_header_cells
    
    @snitch
    def scan_template(self, template_path: str, mapping_config: Optional[Dict[str, Any]] = None, 
                      workbook: Optional[openpyxl.Workbook] = None) -> TemplateAnalysisResult:
        """
        Analyze an Excel template file and extract structure.
        
        Args:
            template_path: Path to the Excel template file
            mapping_config: Optional configuration for header mappings
            workbook: Optional pre-loaded openpyxl Workbook object (for performance)
            
        Returns:
            TemplateAnalysisResult with complete analysis
        """
        path = Path(template_path)
        if not path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")
        
        # Extract customer code from filename (e.g., "CLW.xlsx" -> "CLW")
        customer_code = path.stem.upper()
        
        self.logger.info(f"Scanning template: {path.name} (customer: {customer_code})")
        
        if workbook is None:
             self.logger.debug("Loading workbook from disk...")
             workbook = openpyxl.load_workbook(template_path, data_only=False)
        else:
             self.logger.debug("Using pre-loaded workbook.")
        
        sheets = []
        warnings = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            analysis = self._analyze_sheet(worksheet, sheet_name, mapping_config)
            if analysis:
                sheets.append(analysis)
                # Collect proactive warnings
                if hasattr(analysis, "_temp_warning"):
                    warnings.append(getattr(analysis, "_temp_warning"))

        if not sheets:
             self.logger.warning(f"No valid sheets found in {template_path}. Ensure the file contains recognizable headers.")
             raise ValueError("No valid invoice structure detected. Please check your file content or Mapping Config.")
             
        return TemplateAnalysisResult(
            file_path=str(path.absolute()),
            customer_code=customer_code,
            sheets=sheets,
            warnings=warnings
        )


    def _analyze_sheet(self, worksheet: Worksheet, sheet_name: str, 
                       mapping_config: Optional[Dict[str, Any]] = None) -> Optional[SheetAnalysis]:
        """Analyze a single worksheet."""
        try:
            self.logger.info(f"  Analyzing sheet: {sheet_name}")
            # Filter out unsupported sheets before scanning
            normalized_name = sheet_name.lower().strip()
            
            # Fast mapping resolution using nested structure
            if mapping_config and isinstance(mapping_config, dict):
                sheet_mappings = mapping_config.get('sheet_name_mappings', {}).get('mappings', {})
                if isinstance(sheet_mappings, dict):
                    # Fast case-insensitive exact matching
                    lower_mappings = {k.lower().strip(): v for k, v in sheet_mappings.items()}
                    if normalized_name in lower_mappings:
                        normalized_name = lower_mappings[normalized_name].lower().strip()

            is_supported = False
            
            # Create a set of variants for matching (with/without underscores/spaces)
            variants_to_check = {
                normalized_name,
                normalized_name.replace(' ', '_'),
                normalized_name.replace('_', ' ')
            }
            
            # 1. Exact match check (important for mapped system names like "summary_packing_list")
            for variant in variants_to_check:
                if variant in BlueprintRules.ALLOWED_SEARCH_SHEETS:
                    is_supported = True
                    break
                    
            if not is_supported:
                self.logger.info(f"    Skipping sheet '{sheet_name}': Not in allowed search list.")
                return None
            
            # Find header row
            header_row, header_cells = self._find_header_row(worksheet, mapping_config)
            if not header_row:
                self.logger.warning(f"    No header row found in {sheet_name}")
                return None
            
            self.logger.info(f"    Header row: {header_row}")
            
            # Analyze columns
            columns = self._analyze_columns(worksheet, header_row, header_cells, mapping_config)
            self.logger.info(f"    Found {len(columns)} columns")
            # DETAILED DEBUG LOGGING (Commented out for speed)
            # self.logger.info(f"    --- Column Analysis for {sheet_name} ---")
            # for col in columns:
            #     self.logger.info(f"      [Col {col.col_index}] ID={col.id} Header='{col.header}' Width={col.width:.1f} Format='{col.format}'")
            # self.logger.info(f"    ----------------------------------------")
            
            # Determine data source type
            data_source = self._determine_data_source(sheet_name, columns, mapping_config)
            self.logger.info(f"    Data source: {data_source}")
            
            # Check for multi-row headers and determine data start row
            has_multi_row = self._check_multi_row_header(worksheet, header_row)
            data_start_row = header_row + 1
            if has_multi_row:
                 # Find the bottom-most row of the header (max span of merges starting at header_row)
                 for merged in worksheet.merged_cells.ranges:
                     if merged.min_row == header_row:
                         data_start_row = max(data_start_row, merged.max_row + 1)
            
            self.logger.info(f"    Data starts at row: {data_start_row}")
            
            # Extract font info
            header_font = self._extract_font_info(worksheet, header_row, 1)
            data_font = self._extract_font_info(worksheet, data_start_row, 1)
            
            # Extract row heights
            row_heights = self._extract_row_heights(worksheet, header_row, data_source, data_start_row)
            
            # Detect static content hints (like "Mark & Nº" column content)
            static_hints = self._detect_static_content(worksheet, header_row, columns)
            
            # Note: _extract_description_fallback was removed. 
            # Description is now only detected via label in _detect_static_content.


            # [Smart Feature] Dynamic Footer Analysis (Delegated to Utility)
            footer_info = scan_footer(worksheet, header_row, columns, self.logger, sheet_name=sheet_name, mapping_config=mapping_config)
            if footer_info:
                self.logger.info(f"    Footer detected at row {footer_info.row_num}: '{footer_info.total_text}' (colspan={footer_info.merge_curr_colspan})")
            
            sheet_analysis = SheetAnalysis(
                name=sheet_name,
                header_row=header_row,
                columns=columns,
                data_source=data_source,
                header_font=header_font,
                data_font=data_font,
                row_heights=row_heights,
                has_multi_row_header=has_multi_row,
                static_content_hints=static_hints,
                footer_info=footer_info
            )
            
            # [Proactive Warning] Check for missing footer on financial/aggregation sheets
            if not footer_info and data_source == "aggregation":
                warning_msg = (
                    f"[{sheet_name}] ⚠️ Footer (Total row) NOT detected. "
                    f"WHAT TO DO: Ensure the sheet has a row starting with 'TOTAL' or 'TOTAL AMOUNT'. "
                    f"If the label is different, update 'footer_scanner.py' or 'mapping_config.json'."
                )
                self.logger.warning(warning_msg)
                # Note: We can't easily return it here since SheetAnalysis doesn't hold it, 
                # we'll handle the collection in scan_template
                setattr(sheet_analysis, "_temp_warning", warning_msg)

            return sheet_analysis
            
        except Exception as e:
            self.logger.error(f"    Error analyzing {sheet_name}: {e}")
            return None

    def _analyze_columns(self, worksheet: Worksheet, header_row: int, 
                         header_cells: List[Tuple[int, str]],
                         mapping_config: Optional[Dict[str, Any]] = None) -> List[ColumnInfo]:
        """Analyze columns from header row."""
        columns = []
        processed_cols = set()
        
        # Get merged cell ranges that START at the header row.
        # Used for colspan/rowspan detection of actual table header columns.
        merged_ranges = []
        for merged in worksheet.merged_cells.ranges:
            if merged.min_row == header_row:
                merged_ranges.append(merged)
        
        # Also collect ALL merges that overlap the header row (including from above).
        # Used for resolving cell values within merged regions.
        all_merges_at_header = []
        for merged in worksheet.merged_cells.ranges:
            if merged.min_row <= header_row <= merged.max_row:
                all_merges_at_header.append(merged)
        
        # Cap column analysis at 25 (Col Y)
        safe_max = min(worksheet.max_column + 1, 26)
        
        for col in range(1, safe_max):
            if col in processed_cols:
                continue
            
            cell = worksheet.cell(row=header_row, column=col)
            value = self._get_cell_value(cell)
            
            if not value:
                # Check if this is part of a merged cell
                for merged in all_merges_at_header:
                    if merged.min_col <= col <= merged.max_col:
                        # Get value from top-left of merged range
                        value = self._get_cell_value(
                            worksheet.cell(row=merged.min_row, column=merged.min_col)
                        )
                        break
            
            if not value:
                continue
            
            # Determine column ID
            col_id = self._determine_column_id(value, col, mapping_config)
            
            # [Smart Feature] Leak Filter: Ignore long, unmapped template headers sitting on the same row.
            if col_id.startswith("col_unknown_") and len(value) > 35:
                self.logger.info(f"    [Leak Filter] Ignored long unmapped text as column header: '{value[:30]}...'")
                continue
                
            # Get column width (3-Step Strategy)
            col_letter = get_column_letter(col)
            width = 10.0 # Ultimate fallback
            
            # 1. Explicit
            dim = worksheet.column_dimensions.get(col_letter)
            if dim and dim.width is not None:
                width = dim.width
            # 2. Sheet Default
            elif worksheet.sheet_format and worksheet.sheet_format.defaultColWidth is not None:
                width = worksheet.sheet_format.defaultColWidth
            # 3. Failure
            else:
                 self.logger.warning(f"Could not detect width for column {col_letter} ({value}). Defaulting to 15.0")
                 width = 15.0
            
            # Check for merged cells (colspan/rowspan)
            colspan = 1
            rowspan = 1
            for merged in merged_ranges:
                if merged.min_col == col and merged.min_row == header_row:
                    colspan = merged.max_col - merged.min_col + 1
                    rowspan = merged.max_row - merged.min_row + 1
                    # Mark these columns as processed
                    for c in range(merged.min_col, merged.max_col + 1):
                        processed_cols.add(c)
                    break
            
            # [Smart Feature] Determine format: Check data first, then Rules
            format_str = self._sample_column_format(worksheet, col, header_row + 1)
            if not format_str or format_str == "General":
                # Fallback to rules if no data or general format
                format_str = self._determine_format(col_id, value)
            
            # Check alignment
            alignment = "center"
            if cell.alignment:
                alignment = cell.alignment.horizontal or "center"
            
            # Check wrap text
            wrap_text = cell.alignment.wrap_text if cell.alignment else False
            
            column = ColumnInfo(
                id=col_id,
                header=value,
                col_index=col,
                width=width,
                format=format_str,
                alignment=alignment,
                rowspan=rowspan,
                colspan=colspan,
                wrap_text=wrap_text
            )
            
            # Check for child columns (multi-row headers)
            if rowspan == 1 and colspan > 1:
                # To be a true parent, the row below MUST be split into multiple smaller columns (or cells)
                # and at least one of those cells should have text that corresponds to a mapping.
                # If the row below is merged across the EXACT SAME columns, it's just a wide data column.
                is_true_parent = False
                
                # Check how the row below is structured within this parent's colspan
                child_cells_with_data = 0
                for c in range(col, col + colspan):
                    child_cell = worksheet.cell(row=header_row + 1, column=c)
                    
                    # Look if this cell is the start of a merge covering the whole parent area
                    is_full_width_merge = False
                    for merged in worksheet.merged_cells.ranges:
                        if merged.min_row == header_row + 1 and merged.min_col == col and merged.max_col == col + colspan - 1:
                            is_full_width_merge = True
                            break
                            
                    if is_full_width_merge:
                        break # It's exactly the same width as the parent. Not a parent header.
                        
                    val = self._get_cell_value(child_cell)
                    if val:
                        # Check if it matches a known mapping to be safe
                        mapped_id = self._determine_column_id(val, c, mapping_config)
                        if mapped_id and not mapped_id.startswith("col_unknown"):
                            # Found a valid mapped child! This is a real parent header.
                            is_true_parent = True
                            break
                        child_cells_with_data += 1
                
                # If we found multiple independent cells with data under it, it's a parent
                if child_cells_with_data > 1:
                    is_true_parent = True
                    
                if is_true_parent:
                    children = self._find_child_columns(worksheet, header_row + 1, col, colspan, mapping_config)
                    column.children = children
            
            columns.append(column)
            processed_cols.add(col)
        
        return columns

    def _sample_column_format(self, worksheet: Worksheet, col: int, start_row: int, max_rows: int = 10) -> Optional[str]:
        """
        Sample data rows to find the most common number format.
        Ported from NumberFormatExtractor.
        """
        formats = []
        for row in range(start_row, min(start_row + max_rows, worksheet.max_row + 1)):
            cell = worksheet.cell(row=row, column=col)
            if cell.value is not None and cell.number_format != 'General':
                # Only care about numeric formats
                try:
                    if isinstance(cell.value, (int, float)):
                        formats.append(cell.number_format)
                except:
                    pass
        
        if not formats:
            return None
            
        # Find most common
        from collections import Counter
        most_common = Counter(formats).most_common(1)
        return most_common[0][0] if most_common else None


    
    def _find_child_columns(self, worksheet: Worksheet, row: int, 
                            start_col: int, span: int,
                            mapping_config: Optional[Dict[str, Any]] = None) -> List[ColumnInfo]:
        """Find child columns under a parent header."""
        children = []
        for col in range(start_col, start_col + span):
            cell = worksheet.cell(row=row, column=col)
            value = self._get_cell_value(cell)
            if value:
                # Value is the Child Header text (e.g. "BUFFALO LEATHER")
                col_id = self._determine_column_id(value, col, mapping_config)
                
                # Basic logging to debug mapping failures
                if not col_id or col_id.startswith("col_unknown"):
                    # Check if it was in mapping config but missed?
                    pass

                format_str = self._determine_format(col_id, value)
                col_letter = get_column_letter(col)
                width = worksheet.column_dimensions[col_letter].width or 10
                
                children.append(ColumnInfo(
                    id=col_id,
                    header=value,
                    col_index=col,
                    width=width,
                    format=format_str
                ))
        return children
    
    def _determine_column_id(self, header_text: str, col_index: int,
                             mapping_config: Optional[Dict[str, Any]] = None) -> str:
        """Determine column ID from header text using Config first, then Rules."""
        header_text_stripped = header_text.strip()
        
        # 1. Check User Mapping Config
        if mapping_config:
            mappings = mapping_config.get('header_text_mappings', {}).get('mappings', {})
            
            # Exact match (stripped)
            if header_text_stripped in mappings:
                return mappings[header_text_stripped]
                
            # Case-insensitive match (and strip keys in mapping)
            for mapped_header, mapped_id in mappings.items():
                if mapped_header.strip().lower() == header_text_stripped.lower():
                    return mapped_id

        # 2. Use simple rule-based matching (System Defaults)
        col_def = BlueprintRules.get_column_by_keyword(header_text)
        if col_def:
            return col_def.id
        
        # Fallback: Unknown Column
        # User requested NO auto-generation of IDs (like "col_n_w_kgs").
        # If not matched by Strict Rules or User Mapping, it is UNKNOWN.
        # The user must map this manually in the UI/Config.
        return f"col_unknown_{col_index}"
    
    def _determine_format(self, col_id: str, header_text: str) -> str:
        """Determine number format for column using Rules."""
        return BlueprintRules.get_format_for_id(col_id)
    
    def _determine_data_source(self, sheet_name: str, columns: List[ColumnInfo], mapping_config: Optional[Dict[str, Any]] = None) -> str:
        """
        Determine if this sheet is 'aggregation' (single table) 
        or 'processed_tables_multi' (repeating tables).
        """
        normalized_name = sheet_name.lower().strip()
        
        if mapping_config:
            sheet_mappings = mapping_config.get('sheet_name_mappings', {}).get('mappings', {})
            # Case-insensitive resolution
            lower_mappings = {k.lower().strip(): v for k, v in sheet_mappings.items()}
            if normalized_name in lower_mappings:
                normalized_name = lower_mappings[normalized_name].lower()
                
        variants_to_check = {
            normalized_name,
            normalized_name.replace(' ', '_'),
            normalized_name.replace('_', ' ')
        }
        
        # 1. Exact match check against BlueprintRules definitions
        for variant in variants_to_check:
            if variant in BlueprintRules.AGGREGATION_SHEETS:
                return "aggregation"
            elif variant in BlueprintRules.PROCESSED_TABLES_SHEETS:
                return "processed_tables_multi"
            
        # Fallback (Should not be reached if ALLOWED_SEARCH_SHEETS is strict)
        return "aggregation" # default   
    
    def _extract_font_info(self, worksheet: Worksheet, row: int, col: int) -> Dict[str, Any]:
        """Extract font information from a cell."""
        cell = worksheet.cell(row=row, column=col)
        font = cell.font
        
        return {
            "name": font.name or "Times New Roman",
            "size": font.size or 12,
            "bold": font.bold or False,
            "italic": font.italic or False
        }
    
    def _extract_row_heights(self, worksheet: Worksheet, header_row: int, data_source: str = "dataset_default", data_start_row: int = None) -> Dict[str, float]:
        """Extract row heights using 3-step fallback strategy."""
        if data_start_row is None:
            data_start_row = header_row + 1
        
        def get_height(row_idx: int, standard_key: str) -> float:
            # 1. Explicit
            if row_idx in worksheet.row_dimensions:
                h = worksheet.row_dimensions[row_idx].height
                if h is not None:
                    return float(h)
            
            # 2. Sheet Default (Strict: If 15.0 or None, fail)
            # User wants to be forced to fix the template.
            default_h = worksheet.sheet_format.defaultRowHeight
            if default_h is not None and default_h != 15.0:
                 return float(default_h)
                 
            # 3. Failure
            # Per user request: "if fail just increase it too 1000 px lol" 
            # We will use a safe default of 20.0 instead of a fatal crash, 
            # but log it so the user knows to fix the template eventually.
            self.logger.warning(f"Row {row_idx} has no explicit height. Defaulting to 20.0")
            return 20.0

        try:
            header_height = get_height(header_row, "header")
        except ValueError as e:
             self.logger.warning(f"Header Row {header_row} height detection failed: {e}")
             header_height = 20.0
            
        # Scan next 10 rows for data height (Median)
        heights = []
        for r in range(data_start_row, min(data_start_row + 10, worksheet.max_row + 1)):
            heights.append(get_height(r, "data"))
        
        if heights:
            heights.sort()
            mid = len(heights) // 2
            data_height = heights[mid]
        else:
            data_height = get_height(data_start_row, "data")
            
        return {
            "header": header_height,
            "data": data_height,
            "footer": header_height  # Usually same as header
        }
    
    def _check_multi_row_header(self, worksheet: Worksheet, header_row: int) -> bool:
        """Check if there's a multi-row header structure."""
        for merged in worksheet.merged_cells.ranges:
            if merged.min_row == header_row and merged.max_row > header_row:
                return True
        return False
    
    def _detect_static_content(self, worksheet: Worksheet, header_row: int, 
                               columns: List[ColumnInfo]) -> Dict[str, List[str]]:
        """Detect static content patterns in the data area."""
        hints = {}
        
        # Look for "Mark & Nº" type columns with description label
        for col in columns:
            if col.id == "col_static":
                # Sample first few data rows
                for row in range(header_row + 1, min(header_row + 10, worksheet.max_row + 1)):
                    cell = worksheet.cell(row=row, column=col.col_index)
                    value = self._get_cell_value(cell)
                    if not value:
                        continue
                        
                    # [Smart Feature] Label-based Description Detection
                    # If we find "Des:" or "Desc:" in col_static, capture it!
                    val_upper = value.upper()
                    if val_upper.startswith("DES:") or val_upper.startswith("DESC:"):
                        if ":" in value:
                            desc_part = value.split(":", 1)[1].strip()
                            if desc_part:
                                hints["description_fallback"] = desc_part
                                self.logger.info(f"    [Detection] Found description label in col_static: '{desc_part}'")
                                break # Stop once found
        
        return hints



if __name__ == "__main__":
    import sys
    from core.logger_config import setup_logging
    from core.system_config import sys_config
    setup_logging(log_dir=sys_config.run_log_dir)
    
    if len(sys.argv) < 2:
        print("Usage: python excel_scanner.py <template.xlsx>")
        sys.exit(1)
    
    analyzer = ExcelLayoutScanner()
    result = analyzer.scan_template(sys.argv[1])
    
    print(f"\nTemplate: {result.customer_code}")
    print(f"Sheets: {len(result.sheets)}")
    for sheet in result.sheets:
        print(f"\n  {sheet.name}:")
        print(f"    Header row: {sheet.header_row}")
        print(f"    Data source: {sheet.data_source}")
        print(f"    Columns: {len(sheet.columns)}")
        for col in sheet.columns:
            print(f"      - {col.id}: '{col.header}' (width={col.width:.1f})")
