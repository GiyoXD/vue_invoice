import openpyxl
import traceback
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import range_boundaries, get_column_letter
from typing import Dict, List, Optional, Tuple, Any

center_alignment = Alignment(horizontal='center', vertical='center')

def store_original_merges(workbook: openpyxl.Workbook, sheet_names: List[str]) -> Dict[str, List[Tuple[int, Any, Optional[float]]]]:
    """
    Stores the HORIZONTAL span (colspan), the value of the top-left cell,
    and the height of the starting row for merged ranges in specified sheets,
    ASSUMING all merges are only 1 row high AND **start at row 16 or below**.
    Merges starting above row 16 (row < 16) are ignored.
    """
    original_merges = {}
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            worksheet: Worksheet = workbook[sheet_name]
            merges_data = []
            merged_ranges_copy = list(worksheet.merged_cells.ranges)

            for merged_range in merged_ranges_copy:
                min_col, min_row, max_col, max_row = merged_range.bounds

                if max_row != min_row:
                    continue

                if min_row < 10:
                    continue

                col_span = max_col - min_col + 1
                row_height = None
                try:
                    row_dim = worksheet.row_dimensions[min_row]
                    row_height = row_dim.height
                    top_left_value = worksheet.cell(row=min_row, column=min_col).value
                    merges_data.append((col_span, top_left_value, row_height))

                except Exception:
                    merges_data.append((col_span, None, None))

            original_merges[sheet_name] = merges_data
        else:
             original_merges[sheet_name] = []
    
    return original_merges

def find_and_restore_merges_heuristic(workbook: openpyxl.Workbook,
                                      stored_merges: Dict[str, List[Tuple[int, Any, Optional[float]]]],
                                      processed_sheet_names: List[str],
                                      search_range_str: str = "A10:H200"):
    """
    Attempts to restore merges based on stored HORIZONTAL spans, values, and row heights.
    """
    try:
        search_min_col, search_min_row, search_max_col, search_max_row = range_boundaries(search_range_str)
    except Exception:
        return

    for sheet_name in processed_sheet_names:
        if sheet_name in workbook.sheetnames and sheet_name in stored_merges:
            worksheet: Worksheet = workbook[sheet_name]
            original_merges_data = stored_merges[sheet_name]
            successfully_restored_values_on_sheet = set()
            
            for merge_idx, (col_span, stored_value, stored_height) in enumerate(original_merges_data):
                if col_span <= 1:
                    continue

                if stored_value in successfully_restored_values_on_sheet:
                    continue

                found = False
                for r in range(search_max_row, search_min_row - 1, -1):
                    for c in range(search_min_col, search_max_col + 1):
                        current_cell = worksheet.cell(row=r, column=c)
                        current_val = current_cell.value

                        if current_val == stored_value:
                            start_row, start_col = r, c
                            end_row = start_row
                            end_col = start_col + col_span - 1

                            merged_ranges_copy = list(worksheet.merged_cells.ranges)
                            for existing_merge in merged_ranges_copy:
                                rows_overlap = (existing_merge.min_row <= end_row) and (existing_merge.max_row >= start_row)
                                cols_overlap = (existing_merge.min_col <= end_col) and (existing_merge.max_col >= start_col)

                                if rows_overlap and cols_overlap:
                                    try:
                                        worksheet.unmerge_cells(str(existing_merge))
                                    except Exception:
                                        pass

                            try:
                                worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

                                if stored_height is not None:
                                    try:
                                        worksheet.row_dimensions[start_row].height = stored_height
                                    except Exception:
                                        pass

                                top_left_cell_to_set = worksheet.cell(row=start_row, column=start_col)
                                top_left_cell_to_set.value = stored_value

                                successfully_restored_values_on_sheet.add(stored_value)
                                found = True
                                break

                            except Exception:
                                found = True
                                break

                    if found:
                        break

def force_unmerge_from_row_down(worksheet: Worksheet, start_row: int):
    """
    Forcefully unmerges all cells that start on or after a specific row.
    """
    all_merged_ranges = list(worksheet.merged_cells.ranges)
    
    for merged_range in all_merged_ranges:
        if merged_range.min_row >= start_row:
            try:
                worksheet.unmerge_cells(str(merged_range))
            except Exception:
                pass

class MergeOffsetTracker:
    """
    Tracks row operations to calculate position offsets for empty merge restoration.
    """
    
    def __init__(self):
        self.operations = []  # List of (operation_type, position, count, sheet_name)
    
    def log_delete_rows(self, start_row: int, count: int, sheet_name: str):
        """Log a row deletion operation."""
        self.operations.append(('delete', start_row, count, sheet_name))
    
    def log_insert_rows(self, position: int, count: int, sheet_name: str):
        """Log a row insertion operation."""
        self.operations.append(('insert', position, count, sheet_name))
    
    def calculate_new_position(self, original_row: int, sheet_name: str) -> int:
        """Calculate the new position of a row after all operations."""
        current_row = original_row
        
        for op_type, position, count, op_sheet in self.operations:
            if op_sheet != sheet_name:
                continue
            
            if op_type == 'delete':
                delete_start = position
                delete_end = position + count - 1
                
                if current_row < delete_start:
                    pass
                elif current_row <= delete_end:
                    current_row = -1
                    break
                else:
                    current_row -= count
            
            elif op_type == 'insert':
                if current_row >= position:
                    current_row += count
        
        return current_row


def store_empty_merges_with_coordinates(workbook: openpyxl.Workbook, sheet_names: List[str]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Store empty merges (merges with no value) along with their coordinates for offset-based restoration.
    """
    empty_merges = {}
    
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_empty_merges = []
            merged_ranges_copy = list(worksheet.merged_cells.ranges)
            
            for merged_range in merged_ranges_copy:
                min_col, min_row, max_col, max_row = merged_range.bounds
                
                if max_row != min_row:
                    continue
                
                if min_row < 10:
                    continue
                
                top_left_cell = worksheet.cell(row=min_row, column=min_col)
                cell_value = top_left_cell.value
                
                if cell_value is None or str(cell_value).strip() == "":
                    col_span = max_col - min_col + 1
                    
                    row_height = None
                    try:
                        row_dim = worksheet.row_dimensions[min_row]
                        row_height = row_dim.height
                    except KeyError:
                        pass
                    
                    empty_merge_data = {
                        'original_row': min_row,
                        'col': min_col,
                        'span': col_span,
                        'height': row_height,
                        'coord': merged_range.coord
                    }
                    
                    sheet_empty_merges.append(empty_merge_data)
            
            empty_merges[sheet_name] = sheet_empty_merges
        else:
            empty_merges[sheet_name] = []
    
    return empty_merges


def restore_empty_merges_with_offset(workbook: openpyxl.Workbook, 
                                   empty_merges: Dict[str, List[Dict[str, Any]]], 
                                   offset_tracker: MergeOffsetTracker,
                                   sheet_names: List[str]):
    """
    Restore empty merges using offset calculations.
    """
    for sheet_name in sheet_names:
        if sheet_name not in empty_merges or not empty_merges[sheet_name]:
            continue
        
        worksheet = workbook[sheet_name]
        sheet_merges = empty_merges[sheet_name]
        
        for merge_data in sheet_merges:
            original_row = merge_data['original_row']
            col = merge_data['col']
            span = merge_data['span']
            height = merge_data['height']
            
            new_row = offset_tracker.calculate_new_position(original_row, sheet_name)
            
            if new_row <= 0:
                continue
            
            try:
                end_col = col + span - 1
                worksheet.merge_cells(start_row=new_row, start_column=col, 
                                    end_row=new_row, end_column=end_col)
                
                if height is not None:
                    worksheet.row_dimensions[new_row].height = height
                
            except Exception:
                pass
