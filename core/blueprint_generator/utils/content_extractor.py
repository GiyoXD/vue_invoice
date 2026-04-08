import logging
import re
from typing import List, Optional, Tuple, Dict, Any, TYPE_CHECKING
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell

if TYPE_CHECKING:
    from core.blueprint_generator.excel_scanner import ColumnInfo

logger = logging.getLogger(__name__)

# --- REGEX PATTERNS ---
# Description label pattern: matches "DES:", "DESC. :", "DESCRIPTION:", etc.
DESC_LABEL_PATTERN = re.compile(r'^(?:DES|DESC|DESCRIPTION|DES\.|DESC\.)\s*[:：]\s*(.+)$', re.IGNORECASE)

# Pallet count literal pattern: matches "25 PALLETS", "1 PALLET"
PALLET_PATTERN = re.compile(r'\d+\s*PALLETS?', re.IGNORECASE)
# Pallet count formula pattern: matches "=SUM(...) & \" PALLETS\""
PALLET_FORMULA_PATTERN = re.compile(r'PALLETS?', re.IGNORECASE)

# --- HELPER ---
def _get_cell_value_safe(worksheet: Worksheet, cell) -> Optional[str]:
    """Safely get string value from a cell, handling MergedCells by checking the anchor."""
    if isinstance(cell, MergedCell) or cell.value is None:
        for merged_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                if top_left_cell.value is not None:
                    return str(top_left_cell.value)
        return None
    return str(cell.value)


# --- 1. FALLBACK DESCRIPTION EXTRACTION ---

def detect_static_description_label(worksheet: Worksheet, header_row: int, columns: List['ColumnInfo'], max_sample_rows: int = 10) -> Optional[str]:
    """
    Detects the description label (e.g. "DES: COW LEATHER") from the static column.
    Uses regex for robust discovery.
    """
    for col in columns:
        if col.id == "col_static":
            for row in range(header_row + 1, min(header_row + max_sample_rows + 1, worksheet.max_row + 1)):
                cell = worksheet.cell(row=row, column=col.col_index)
                value = _get_cell_value_safe(worksheet, cell)
                if not value:
                    continue
                
                # Check using regex instead of strict startswith
                match = DESC_LABEL_PATTERN.search(value.strip())
                if match:
                    desc_part = match.group(1).strip()
                    if desc_part:
                        logger.info(f"    [Detection] Found description label in col_static using regex: '{desc_part}'")
                        return desc_part
    return None

def extract_table_fallback_description(worksheet: Worksheet, data_start: int, data_end: int, col_desc_index: int) -> Optional[str]:
    """
    Extracts unique descriptions straight from the data table's description column.
    """
    unique_descs = []
    seen_desc = set()
    
    if data_start <= data_end:
        for r in range(data_start, data_end + 1):
            val = _get_cell_value_safe(worksheet, worksheet.cell(row=r, column=col_desc_index))
            if val:
                val_str = str(val).strip()
                if val_str and val_str not in seen_desc:
                    unique_descs.append(val_str)
                    seen_desc.add(val_str)
    
    if unique_descs:
        fallback_description = " / ".join(unique_descs)
        logger.info(f"    [Extracted] Fallback descriptions from data table: '{fallback_description}'")
        return fallback_description
    return None


# --- 2. HS CODE EXTRACTION ---

def extract_global_hs_code(worksheet: Worksheet, max_row: int = 150, max_col: int = 25) -> Optional[str]:
    """
    Performs a global search for "HS.CODE" or "HSCODE" across the worksheet.
    """
    for row_cells in worksheet.iter_rows(max_row=max_row, max_col=max_col):
        for cell in row_cells:
            val = cell.value
            if val and isinstance(val, (str, bytes)):
                val_str = str(val)
                # Remove spaces and uppercase to check
                val_upper = val_str.upper().replace(" ", "")
                if "HS.CODE" in val_upper or "HSCODE" in val_upper:
                    hs_code = val_str.strip()
                    logger.info(f"    [Global Search] HS Code cell found at {cell.coordinate}: '{hs_code}'")
                    return hs_code
    return None

def find_footer_hs_code(worksheet: Worksheet, start_row: int, end_row: int) -> Tuple[Optional[str], int]:
    """
    Scan specifically in the footer bounds for HS Code to determine if it's there and its colspan.
    """
    hs_keywords = {"HS.CODE", "HS CODE", "HS-CODE"}
    
    for row in range(start_row, end_row + 1):
        for col in range(1, min(worksheet.max_column + 1, 20)):
            cell = worksheet.cell(row=row, column=col)
            val = _get_cell_value_safe(worksheet, cell)
            if not val:
                continue
                
            upper_val = val.upper()
            if any(kw in upper_val for kw in hs_keywords):
                # Calculate colspan
                colspan = 1
                for merged in worksheet.merged_cells.ranges:
                    if merged.min_row <= cell.row <= merged.max_row and merged.min_col <= cell.column <= merged.max_col:
                        colspan = merged.max_col - merged.min_col + 1
                        break
                return val, colspan
                
    return None, 1


# --- 3. FOOTER ELEMENTS (PALLET & TOTAL LABELS) ---

def find_total_label_cell(worksheet: Worksheet, start_row: int, end_row: int, mapping_config: Optional[Dict[str, Any]] = None, logger_instance: Optional[logging.Logger] = None, sheet_name: str = "Unknown"):
    """
    Scan rows for the first cell containing a TOTAL-like label.
    """
    total_keywords = []
    if mapping_config and "footer_label_mappings" in mapping_config:
        mappings = mapping_config["footer_label_mappings"].get("keywords", [])
        total_keywords = [kw.upper() for kw in mappings]
        
    if not total_keywords:
        if logger_instance:
            logger_instance.warning(f"    ⚠ [{sheet_name}] No footer keywords configured. Cannot detect footer row!")
        return None
    
    for row in range(start_row, end_row + 1):
        for col in range(1, min(worksheet.max_column + 1, 20)):
            cell = worksheet.cell(row=row, column=col)
            val = _get_cell_value_safe(worksheet, cell)
            if not val:
                continue
                
            val_upper = val.upper()
            for kw in total_keywords:
                if kw in val_upper:
                    is_exact = (kw == val_upper)
                    return cell, is_exact
    return None

def find_pallet_count_column(worksheet: Worksheet, footer_row: int, columns: List['ColumnInfo'], find_col_id_func, logger_instance: logging.Logger, sheet_name: str = "Unknown") -> Optional[str]:
    """
    Scan the footer row for a pallet count pattern.
    """
    for col in range(1, min(worksheet.max_column + 1, 20)):
        cell = worksheet.cell(row=footer_row, column=col)
        val = _get_cell_value_safe(worksheet, cell)
        
        if not val:
            continue
        
        if PALLET_PATTERN.search(val) or (val.startswith("=") and PALLET_FORMULA_PATTERN.search(val)):
            pallet_col_id = find_col_id_func(col, columns)
            logger_instance.info(f"    [{sheet_name}] Pallet count detected at col {col} -> {pallet_col_id}")
            return pallet_col_id
    
    logger_instance.warning(f"    ⚠ [{sheet_name}] No pallet count pattern found on footer row {footer_row}")
    return None
