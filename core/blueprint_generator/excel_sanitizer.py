"""
Template Cleaner - Cleans raw Excel files to create blank templates.

This module is responsible for:
1. Stripping data rows from populated invoices/packing lists.
2. Preserving header rows and styling.
3. Injecting system placeholders (JFINV, JFTIME, etc.) into specific cells.
"""

import re
import logging
import hashlib
import json
from typing import Dict, Any, List, Optional, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter

from .excel_scanner import TemplateAnalysisResult, SheetAnalysis

logger = logging.getLogger(__name__)

class ExcelTemplateSanitizer:
    """Cleans (sanitizes) raw Excel files to create reusable templates."""

    def __init__(self):
        self.logger = logging.getLogger(self.__class__.__name__)
        # Excel default dimensions (used to filter out empty cells with default size)
        self.DEFAULT_ROW_HEIGHT = 15.0  # Excel default row height in points
        self.DEFAULT_COL_WIDTH = 8.43   # Excel default column width in characters

    def sanitize_template(self, workbook: openpyxl.Workbook, analysis: TemplateAnalysisResult) -> Tuple[openpyxl.Workbook, Dict[str, Any]]:
        """
        Clean the provided workbook based on analysis.
        
        Args:
            workbook: openpyxl Workbook object (raw file)
            analysis: TemplateAnalysisResult
            
        Returns:
            Tuple of (Cleaned Workbook, layout_metadata_dict)
        """
        self.logger.info(f"Cleaning template for {analysis.customer_code}...")
        
        layout_metadata = {}
        
        for sheet_analysis in analysis.sheets:
            if sheet_analysis.name in workbook.sheetnames:
                ws = workbook[sheet_analysis.name]
                sheet_layout = self._clean_sheet(ws, sheet_analysis)
                layout_metadata[sheet_analysis.name] = sheet_layout
        
        # === FIX FOR CORRUPTION ===
        # Clear images ONLY from cleaned sheets.
        # Images in openpyxl can cause corruption when the workbook is saved
        # after row deletions. We've already captured image data to metadata.
        # Unknown sheets (not in analysis) are untouched, so they can keep their images.
        analyzed_sheet_names = {sheet.name for sheet in analysis.sheets}
        
        for sheet in workbook.worksheets:
            if sheet.title in analyzed_sheet_names:
                if hasattr(sheet, '_images'):
                    sheet._images = []
                # Also clear drawings which can cause issues
                if hasattr(sheet, '_charts'):
                    sheet._charts = []
                
        return workbook, layout_metadata

    def _clean_sheet(self, ws: Worksheet, analysis: SheetAnalysis) -> Dict[str, Any]:
        """Clean a single sheet: strip data rows, inject placeholders."""
        self.logger.info(f"  Cleaning sheet: {analysis.name}")
        
        preserved_layout = {
            "header_merges": {},
            "header_row_heights": {},
            "header_content": {},
            "header_styles": {},
            "footer_rows": [],
            "style_palette": {},
            "col_widths": {},
            "header_images": [],
            "footer_images": []
        }
        
        # Local palette cache mapping hash string -> actual style dict
        # This will be injected into preserved_layout
        local_style_palette = {}
        
        def process_and_store_style(style_dict: Dict[str, Any]) -> str:
            """Hashes a style dict and adds it to the palette if new. Returns the hash ID."""
            # Sort keys so the hash is deterministic
            style_str = json.dumps(style_dict, sort_keys=True)
            style_hash = "style_" + hashlib.md5(style_str.encode('utf-8')).hexdigest()[:8]
            
            if style_hash not in local_style_palette:
                local_style_palette[style_hash] = style_dict
                
            return style_hash
        
        
        # Determine strict max column based on BOTH table and header content.
        # The header area may extend wider than the table (e.g. Ref No, Date cells).
        
        # Step 1: Get table column boundary
        table_cur_max = 0
        if analysis.columns:
            for col in analysis.columns:
                end_col = col.col_index + (col.colspan - 1)
                if end_col > table_cur_max:
                    table_cur_max = end_col
        
        # Step 2: Pre-scan header rows to find actual rightmost content column
        # Capped at 25 to prevent runaway from phantom-formatted cells
        header_max_col = 0
        if analysis.header_row > 1:
            for row in ws.iter_rows(min_row=1, max_row=analysis.header_row - 1,
                                    min_col=1, max_col=25):
                for cell in row:
                    if cell.value is not None and cell.column > header_max_col:
                        header_max_col = cell.column
        
        # Step 3: Combine — use the wider of table vs header, +1 buffer, hard cap at 25
        dynamic_limit = max(table_cur_max, header_max_col) + 1
        safe_max_column = min(ws.max_column, 25, dynamic_limit)
            
        self.logger.info(f"    Dynamic Column Scan Limit: {safe_max_column} (Table Max: {table_cur_max}, Header Max: {header_max_col})")
        
        
        # --- CAPTURE GLOBAL LAYOUT (Column Widths) ---

        for c in range(1, safe_max_column + 1):
            letter = get_column_letter(c)
            if letter in ws.column_dimensions:
                w = ws.column_dimensions[letter].width
                # width can be None for default
                if w is not None:
                     preserved_layout["col_widths"][letter] = w
        
        # --- [Smart Feature] Extract Fallback Description (Handles Mixed Values) ---
        # Look for col_desc in the analysis and collect ALL unique values from the data body.
        fallback_description = None
        col_desc_index = None
        for col in analysis.columns:
            if col.id == "col_desc":
                col_desc_index = col.col_index
                break
        
        if col_desc_index and analysis.header_row > 0:
            # We need the footer to know where to stop scanning data
            footer_row = self._find_footer_start(ws, analysis.header_row + 1, analysis)
            data_start = analysis.header_row + 1
            data_end = footer_row - 1 if footer_row else ws.max_row
            
            unique_vals = []
            seen = set()
            
            if data_start <= data_end:
                for r in range(data_start, data_end + 1):
                    val = ws.cell(row=r, column=col_desc_index).value
                    if val:
                        val_str = str(val).strip()
                        if val_str and val_str not in seen:
                            unique_vals.append(val_str)
                            seen.add(val_str)
            
            if unique_vals:
                fallback_description = " / ".join(unique_vals)
                self.logger.info(f"    [Extracted] Fallback descriptions from {analysis.name}: '{fallback_description}'")

        preserved_layout["fallback_description"] = fallback_description

        # --- CAPTURE HEADER LAYOUT & CONTENT ---
        # Capture strictly ABOVE the header row (Metadata area)
        
        # 1. Header Merges

        for merged_range in ws.merged_cells:
            if merged_range.max_row < analysis.header_row:
                 range_str = str(merged_range)
                 # Extract value from top-left cell to act as key
                 top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                 val = str(top_left_cell.value) if top_left_cell.value is not None else ""
                 # Clean up newlines for cleaner JSON
                 val_clean = val.replace('\n', ' ').strip()
                 preserved_layout["header_merges"][range_str] = val_clean
                 
        # 2. Header Row Heights
        for r in range(1, analysis.header_row):
            if r in ws.row_dimensions:
                h = ws.row_dimensions[r].height
                if h is not None:
                    preserved_layout["header_row_heights"][str(r)] = h
                    
        # 3. Header Content & Styles
        # Use iter_rows for speed
        if analysis.header_row > 1:
            for row in ws.iter_rows(min_row=1, max_row=analysis.header_row-1, min_col=1, max_col=safe_max_column):
                for cell in row:
                    coord = cell.coordinate
                    is_empty = (cell.value is None)
                    
                    # Content
                    if not is_empty:
                         val_str = str(cell.value)
                         # Strip external workbook refs [N] from formulas
                         # Excel converts =DeepSheet!B1 to =[1]DeepSheet!B1 when the sheet doesn't exist
                         if val_str.startswith('='):
                             val_str = re.sub(r'\[\d+\]', '', val_str)
                         preserved_layout["header_content"][coord] = val_str
                    
                    # Style - skip empty cells with default dimensions
                    if is_empty and not self._should_record_empty_cell(ws, cell.row, cell.column):
                        continue
                        
                    style_data = self._capture_cell_style(cell, is_empty=is_empty)
                    if style_data:
                        style_id = process_and_store_style(style_data)
                        if style_id not in preserved_layout["header_styles"]:
                            preserved_layout["header_styles"][style_id] = []
                        preserved_layout["header_styles"][style_id].append(coord)
        
        # 1. Find Footer (Total Row) FIRST, before messing with indices?
        # Actually indices are stable if we haven't deleted yet.
        # Scan bottom-up for "Total" + "=SUM"
        # We search from header_row + 1
        footer_start_row = self._find_footer_start(ws, analysis.header_row + 1, analysis)
        
        if not footer_start_row:
             self.logger.warning(f"    Footer not detected in {analysis.name}. Assuming this is a Form/Static sheet (no dynamic table body).")
             self.logger.warning("    Skipping row deletion to preserve content.")
             # Set end_delete to header_row - 1 so rows_to_delete becomes 0
             end_delete = analysis.header_row - 1
             footer_start_row = ws.max_row # Treat end of sheet as footer start for image capture purposes
        else:
             end_delete = footer_start_row

        # 2. Delete ENTIRE Table (Header to Footer inclusive)
        # User requested: "remove from the header to the footer entirely"
        # This removes Header, Data Body, and the Footer row found.
        # Remaining content (like Signature) will shift up.
        
        start_delete = analysis.header_row
        
        if end_delete < start_delete:
            rows_to_delete = 0
        else:
            rows_to_delete = end_delete - start_delete + 1
        
        
        # 3. SKIP Image Capture - was causing corruption
        # Images are cleared in sanitize_template() before saving
        preserved_layout["header_images"] = []
        preserved_layout["footer_images"] = []
        
        if rows_to_delete > 0:
            self.logger.info(f"    Deleting ENTIRE TABLE: {rows_to_delete} rows (Rows {start_delete}-{end_delete} | Header {start_delete} to Footer {end_delete})")
            
            # --- PRE-DELETION MERGE HANDLING ---
            # 1. Ranges IN the deletion zone: Destroy them.
            # 2. Ranges BELOW the deletion zone (Footer): Store & Destroy, then Restore after deletion (to prevent corruption).
            
            footer_merges_to_restore = []
            
            # Iterate over a copy of the list because we modify it
            for merged_range in list(ws.merged_cells):
                m_min_row, m_min_col, m_max_row, m_max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
                
                # Case A: Merge intersects deletion zone -> Destroy (Ghost Merge Prevention)
                if (m_min_row <= end_delete and m_max_row >= start_delete):
                    self.logger.info(f"    Unmerging intersecting range {merged_range} before deletion.")
                    ws.unmerge_cells(str(merged_range))
                    
                # Case B: Merge is strictly BELOW deletion zone (Footer) -> Store & Unmerge
                elif m_min_row > end_delete:
                    self.logger.info(f"    Storing & Temporarily Unmerging footer range {merged_range} to preserve format.")
                    footer_merges_to_restore.append((m_min_row, m_min_col, m_max_row, m_max_col))
                    ws.unmerge_cells(str(merged_range))

            # Case C: Capture Row Heights, Content, Styles, and Merges as Row Objects
            footer_heights_to_restore = []
            footer_rows = []
            
            current_max_row = ws.max_row
            
            for r in range(end_delete + 1, current_max_row + 1):
                rel_r = r - (end_delete + 1) # 0-indexed relative to footer block start
                row_dict = {
                    "relative_index": rel_r,
                    "height": None,
                    "merges": [],
                    "cells": []
                }
                
                # 1. Capture Height
                if r in ws.row_dimensions:
                    h = ws.row_dimensions[r].height
                    if h is not None:
                        row_dict["height"] = h
                        footer_heights_to_restore.append((r, h))
                        
                # 2. Capture Merges specifically starting on this row
                for (old_min_r, min_c, old_max_r, max_c) in footer_merges_to_restore:
                    if old_min_r == r:
                        # Find the top-left cell value if any (mimicking old logic)
                        top_left_cell = ws.cell(row=r, column=min_c)
                        val = str(top_left_cell.value) if top_left_cell.value is not None else ""
                        val_clean = val.replace('\n', ' ').strip()
                        
                        row_dict["merges"].append({
                            "min_col": min_c,
                            "max_col": max_c,
                            "row_span": old_max_r - old_min_r + 1,
                            "value": val_clean
                        })
                        
                # 3. Capture Content & Styles
                has_content_or_style = False
                for c in range(1, safe_max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    is_empty = (cell.value is None)
                    
                    if is_empty and not self._should_record_empty_cell(ws, r, c):
                        continue
                        
                    cell_dict = {"col_index": c}
                    has_content_or_style = True
                    
                    if not is_empty:
                        val_str = str(cell.value)
                        # Strip external workbook refs [N] from formulas
                        if val_str.startswith('='):
                            val_str = re.sub(r'\[\d+\]', '', val_str)
                        cell_dict["value"] = val_str
                        
                    style_data = self._capture_cell_style(cell, is_empty=is_empty)
                    if style_data:
                        style_id = process_and_store_style(style_data)
                        cell_dict["style_id"] = style_id
                        
                    row_dict["cells"].append(cell_dict)
                    
                # Store the row if it bears ANY layout information
                if row_dict["height"] is not None or row_dict["merges"] or has_content_or_style:
                    footer_rows.append(row_dict)
                    
            preserved_layout["footer_rows"] = footer_rows

            # --- DELETION ---
            ws.delete_rows(start_delete, amount=rows_to_delete)
            
            # --- RESTORE FOOTER MERGES ---
            # Shift rows up by rows_to_delete
            for (old_min_r, min_c, old_max_r, max_c) in footer_merges_to_restore:
                new_min_r = old_min_r - rows_to_delete
                new_max_r = old_max_r - rows_to_delete
                
                # Sanity check
                if new_min_r < 1: 
                    continue
                    
                self.logger.info(f"    Restoring footer merge at rows {new_min_r}-{new_max_r} (was {old_min_r}-{old_max_r})")
                ws.merge_cells(start_row=new_min_r, start_column=min_c, end_row=new_max_r, end_column=max_c)
                
                # (Value extraction now happens in the row loop above)

            # --- RESTORE FOOTER ROW HEIGHTS ---
            for (old_r, height) in footer_heights_to_restore:
                new_r = old_r - rows_to_delete
                if new_r < 1: continue
                
                # self.logger.info(f"    Restoring footer row height {height} at row {new_r} (was {old_r})")
                ws.row_dimensions[new_r].height = height

        else:
            self.logger.warning("    Row calculation result <= 0? Check header/footer detection.")
        
        # Inject the final populated palette back into the layout
        preserved_layout["style_palette"] = local_style_palette
        
        return preserved_layout



    def _find_footer_start(self, ws: Worksheet, search_start_row: int, analysis: Optional[SheetAnalysis] = None) -> Optional[int]:
        """
        Find the footer row by scanning for =SUM or =SUBTOTAL formula adjacency.

        Algorithm:
            1. Scan bottom-up from max_row to search_start_row.
            2. For each row, collect column indices where cell value starts with '=SUM' or '=SUBTOTAL'.
               Skip any cell not starting with '=' for speed.
            3. If 2+ adjacent (consecutive) column indices have =SUM/=SUBTOTAL, mark row as candidate.
            4. Return the LAST (highest row number) candidate found.
            5. Fallback 1: Use the top-down scanner detection from SheetAnalysis if available.
            6. Fallback 2: if no adjacency match, try strict 'TOTAL OF:' keyword detection bottom-up.

        Args:
            ws: The worksheet to scan.
            search_start_row: The row to start scanning from (header_row + 1).
            analysis: The SheetAnalysis from the scanner (optional).

        Returns:
            The 1-based row number of the footer, or None if not found.
        """
        end_scan = min(ws.max_row, search_start_row + 500)
        max_col = min(ws.max_column + 1, 20)
        last_candidate = None  # Highest row with 2+ adjacent formula cells

        for row in range(search_start_row, end_scan + 1):
            formula_cols = []  # Column indices with =SUM or =SUBTOTAL in this row

            for col in range(1, max_col):
                cell = ws.cell(row=row, column=col)
                value = self._get_cell_value(cell)

                if not value:
                    continue
                # Fast skip: only care about formulas
                if not value.startswith("="):
                    continue
                upper_val = value.upper()
                if upper_val.startswith("=SUM(") or upper_val.startswith("=SUBTOTAL("):
                    formula_cols.append(col)

            # Check adjacency: need 2+ consecutive column indices
            if len(formula_cols) >= 2 and self._has_adjacent_pair(formula_cols):
                last_candidate = row

        if last_candidate:
            self.logger.info(f"    Footer detected via formula adjacency at row {last_candidate}")
            return last_candidate

        # --- FALLBACK 1: Scanner's Top-Down Footer Detection ---
        if analysis and analysis.footer_info and analysis.footer_info.row_num:
            self.logger.info(f"    Formula adjacency scan found nothing. Falling back to scanner footer info at row {analysis.footer_info.row_num}.")
            return analysis.footer_info.row_num

        # --- FALLBACK 2: Original 'TOTAL' keyword scan (bottom-up), but strict ---
        self.logger.info("    Formula adjacency scan and scanner info found nothing. Falling back to strict TOTAL keyword scan.")
        fallback_candidate = None
        for row in range(ws.max_row, search_start_row, -1):
            for col in range(1, max_col):
                cell = ws.cell(row=row, column=col)
                value = self._get_cell_value(cell)
                if value:
                    val_upper = value.upper().strip()
                    # Use exact/near-exact match to prevent picking up random sentence "Total Net Weight"
                    if val_upper in ["TOTAL", "TOTAL:", "TOTAL OF:", "TOTAL OF", "TOTAL：", "TOTAL AMOUNT", "TOTAL AMOUNT:", "TOTAL AMOUNT："] or val_upper.startswith("TOTAL OF") or val_upper.startswith("TOTAL AMOUNT"):
                        fallback_candidate = row
                        break
            if fallback_candidate:
                break

        if fallback_candidate:
            self.logger.warning(f"    Using strict TOTAL keyword fallback at row {fallback_candidate}.")
        return fallback_candidate

    def _has_adjacent_pair(self, sorted_cols: list) -> bool:
        """
        Check if a sorted list of column indices contains at least one adjacent pair.

        Args:
            sorted_cols: List of 1-based column indices (already in order from left-to-right scan).

        Returns:
            True if any two consecutive entries differ by exactly 1.
        """
        for i in range(len(sorted_cols) - 1):
            if sorted_cols[i + 1] - sorted_cols[i] == 1:
                return True
        return False

    def _get_cell_value(self, cell) -> Optional[str]:
        # If the cell is a part of a merged range, its value might be stored in the top-left cell.
        # openpyxl represents non-top-left cells as MergedCell which usually have .value == None
        if isinstance(cell, MergedCell) or cell.value is None:
            # We must check if this cell belongs to any merged range to pull the text from the anchor
            ws = cell.parent
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Get the top-left cell of the merged range
                    top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    if top_left_cell.value is not None:
                        return str(top_left_cell.value)
            return None
            
        return str(cell.value)

    def _should_record_empty_cell(self, ws: Worksheet, row: int, col: int) -> bool:
        """
        Check if an empty cell should be recorded based on non-default dimensions.
        
        An empty cell is worth recording if:
        - Its row has a height different from Excel's default (15.0 pt)
        - Its column has a width different from Excel's default (8.43 chars)
        
        This prevents recording thousands of empty cells with only default styling.
        """

        
        # Check row height
        if row in ws.row_dimensions:
            height = ws.row_dimensions[row].height
            if height is not None and height != self.DEFAULT_ROW_HEIGHT:
                return True
        
        # Check column width
        col_letter = get_column_letter(col)
        if col_letter in ws.column_dimensions:
            width = ws.column_dimensions[col_letter].width
            if width is not None and width != self.DEFAULT_COL_WIDTH:
                return True
        
        return False

    def _capture_cell_style(self, cell: Cell, is_empty: bool = False) -> Optional[Dict[str, Any]]:
        """
        Capture font, alignment, border, and fill styles from a cell.
        Returns None if strict mode is on and the cell has effectively default style.
        
        Args:
            cell: The cell to analyze.
            is_empty: If True, we ignore "setup" styles like Font Name/Size and only capture 
                      visible modifiers like Borders, Fills, Bold, etc.
        """
        style = {}
        has_significant_style = False
        
        # 1. Font
        if cell.font:
            # Filter out default font props to save space
            font_data = {}
            
            # If cell is empty, we DON'T care about Font Name/Size (User preference)
            # We only care if it's explicitly styled (Bold, Italic, Color)
            if not is_empty:
                if cell.font.name and cell.font.name not in ["Calibri", "Arial"]: # Capture non-standard fonts
                    font_data["name"] = cell.font.name
                if cell.font.size and cell.font.size not in [11.0, 11, 10.0, 10]: # Capture non-standard sizes
                    font_data["size"] = cell.font.size
            
            if cell.font.bold: font_data["bold"] = True
            if cell.font.italic: font_data["italic"] = True
            if cell.font.color and hasattr(cell.font.color, "rgb"): # Capture colors
                 color_val = self._serialize_color(cell.font.color)
                 if color_val and color_val != "00000000": # Skip black/auto
                     font_data["color"] = color_val

            if font_data:
                style["font"] = font_data
                has_significant_style = True
            
        # 2. Alignment
        if cell.alignment:
            align_data = {}
            # Only capture if NOT default (general/bottom)
            if cell.alignment.horizontal and cell.alignment.horizontal != 'general':
                align_data["horizontal"] = cell.alignment.horizontal
            if cell.alignment.vertical and cell.alignment.vertical != 'bottom': # bottom is default in Excel? usually.
                align_data["vertical"] = cell.alignment.vertical
            if cell.alignment.wrap_text:
                align_data["wrap_text"] = True
                
            if align_data:
                style["alignment"] = align_data
                has_significant_style = True
            
        # 3. Fill (Background)
        if cell.fill and hasattr(cell.fill, "start_color"):
             color_val = self._serialize_color(cell.fill.start_color)
             # Skip default "none" or white fills
             if color_val and color_val not in ["00000000", "FFFFFFFF", None]:
                 style["fill"] = {
                     "type": cell.fill.fill_type,
                     "color": color_val
                 }
                 has_significant_style = True
             
        # 4. Border
        if cell.border:
             border_data = {}
             # Only capture if there is an actual border style
             if cell.border.left and cell.border.left.style: border_data["left"] = cell.border.left.style
             if cell.border.right and cell.border.right.style: border_data["right"] = cell.border.right.style
             if cell.border.top and cell.border.top.style: border_data["top"] = cell.border.top.style
             if cell.border.bottom and cell.border.bottom.style: border_data["bottom"] = cell.border.bottom.style
             
             if border_data:
                 style["border"] = border_data
                 has_significant_style = True
             
        # 5. Number Format
        # Only relevant if content exists, usually? 
        # Actually user might pre-format a column for dates. 
        # But for "Pure Empty Junk" check, maybe skip? 
        # User said "modify more property like border or change width and height". 
        # Number format is invisible until data is typed.
        # Let's keep strictness: If empty, ignore number format too?
        # Safe bet: Capture it. It's rare to have Custom Num Format on junk cells.
        if cell.number_format and cell.number_format != "General":
            style["number_format"] = cell.number_format
            has_significant_style = True
        
        # If nothing significant was captured, return None to save massive JSON space
        return style if has_significant_style else None

    def _serialize_color(self, color) -> Optional[str]:
        """Try to extract RGB hex string from Color object."""
        if not color: return None
        if hasattr(color, "rgb") and color.rgb:
            # openpyxl rgb is usually "AARRGGBB" or "RRGGBB"
            # We treat it as string
            if isinstance(color.rgb, str):
                return color.rgb
        return None

