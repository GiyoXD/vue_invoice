import openpyxl
from datetime import datetime
from core.invoice_generator.builders.deep_sheet_builder import DeepSheetBuilder

def test_deepsheet_date():
    wb = openpyxl.Workbook()
    # Mock invoice data with ISO date string
    invoice_data = {
        "invoice_info": {
            "col_inv_no": "TEST-123",
            "col_inv_date": "2026-03-20",
            "col_inv_ref": "REF-456"
        }
    }
    
    # Build DeepSheet
    DeepSheetBuilder.build(wb, invoice_data)
    
    # Check DeepSheet
    ws = wb["DeepSheet"]
    date_label = ws.cell(row=3, column=1).value
    date_val = ws.cell(row=3, column=2).value
    date_format = ws.cell(row=3, column=2).number_format
    
    print(f"Label: {date_label}")
    print(f"Value Type: {type(date_val)}")
    print(f"Value: {date_val}")
    print(f"Number Format: {date_format}")
    
    assert date_label == "date"
    assert isinstance(date_val, (datetime, datetime.date))
    assert date_format == "yyyy-mm-dd"
    print("Verification SUCCESSFUL!")

if __name__ == "__main__":
    test_deepsheet_date()
