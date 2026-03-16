import openpyxl
import sys

def main():
    wb = openpyxl.load_workbook(r"c:\Users\JPZ031127\Desktop\project\vue_invoice_project\database\temp_uploads\Copy of CT&INV&PL JLFTLT-VC26008 FCA(1).xlsx", data_only=True)
    
    if "Contract" in wb.sheetnames:
        sheet = wb["Contract"]
    elif "CONTRACT" in wb.sheetnames:
        sheet = wb["CONTRACT"]
    else:
        print("Sheet not found")
        sys.exit(1)
        
    for row in range(15, min(sheet.max_row + 1, 100)):
        values = []
        for col in range(1, 15):
            val = sheet.cell(row=row, column=col).value
            if val is not None:
                values.append(f"Col {col}: '{val}'")
        if values:
            print(f"Row {row}: {' | '.join(values)}")

if __name__ == "__main__":
    main()
